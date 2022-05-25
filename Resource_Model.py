import shutil
import zipfile
import shutil
from tkscrolledframe import ScrolledFrame
from mplot import plot
from mplot.color import *
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg,NavigationToolbar2Tk)
import tkinter as tk
#import tkinter.tix as tx
from tkinter import messagebox
from tkinter import filedialog
from tkinter.simpledialog import askstring
from tkinter.messagebox import showinfo
from tkinter import ttk
import re
import numpy as np
from PIL import ImageTk, Image
import os
from scipy.stats import linregress
import random
import math
import copy
import openpyxl
import pickle
import csv
import mlatex
import datetime
import sys
import tempfile



RUNLATEX = True
LATEX = 'latex'
BIBTEX = 'bibtex'
REST = 'Rest'
CPPMODEL = "GeRSDeMo_csv.exe"
RECYCLENAME = "recycle"
DEMAND = "demandall"
MODELCOLOR = RED
DATASTYLE = '.'
DATALABEL = 'Historic'
RECYCLE = 'Recycle'
DATACOLOR = BLACK
DVIPS = 'DviPs'
PSPDF = 'ps2pdf'

'''
Need to ensure no duplicate production or URR scenario...
'''
TMPFOLDERTAG = 'TMP_'
PRODWORDS = ['data','prod']
URRWORDS = ['URR','resource']
MAINCLASSFILE = 'MainClass.pickle'
SAVEFRAMESFILE = 'Frames.pickle'
SAVEMAPPINGFILE = 'Mapping.pickle'
YHEIGHT = 10 # as a percentage of YMax.
SAVEFAILED = "Save Failed"
RUNSCENFAILED = 'Run Scenario Failed'
FAILEDOPENFILE = 'Open file failed'
LOADFAILED = 'Load Failed'
IMPORTFAILED = 'Import Failed'
STARTYEAR = 'Start Year:'
ENDYEAR = 'End Year:'
URRLAB = 'URR:'
MINCUM = 0.1 # as a percent
MAXVLINES = 10
RUNNINGSKIPS = 2
POPUPGEOM = '350x100+250+125'
DEFAULTOUTLIERLEVEL = 5
PADX = 5
PADY = 5
GRIDSIZE = (13,9)
PROGRAMTITLE = "Resource Modeller"
MINFRAMESIZE = (1000,500)
FRAMESIZE = (3000,3000)
MINROWSIZE = 40
MINCOLSIZE = 43
BADSYMBOLS = ['+','*']
MINVALUES = 10
LOADTEXT = 'Load'
NEWMODELTEXT = 'New Model'
#FRAMECOLOR = '#D3D3D3'
FRAMECOLOR = None
HLMESSAGE = 'Select start and end year'
HLDESC = "Hover over graph to find corresponding year"
MAINTITLETEXT = "Resource Modeller"
MAINDESCTEXT = ['Welcome to the Resource Modeller, here you will be able to project a geological resource using',
                'either GeRS-DeMo or other methods such as Hubbert curves. It is assumed that you have production data',
                'already collated into the correct format for the software.']
DESCLOC = (1,0)
HLPLOTSIZE = (6,4.5)
MAINTITLELOC = (0,0)
LOADMODELLOC = (10,3)
IMPORTMODELLOC = (10,2)
NEWMODELLOC = (10,4)
GERSDEMOLOC = (0.3, 0.5)
OTHERMODELSLOC = (0.7,0.5)
LOADPRODLOC = (8,3)
IMPORTTEXT = 'Import'
INTERACTIONTAB = 'Interaction'
MINETAB = 'Mine'
FIELDTAB = 'Field'
GENERALTAB = 'General_Info'
IMPORTTABS = [INTERACTIONTAB,MINETAB,FIELDTAB,GENERALTAB]


BOOSTABOVEMIN = 2.5


LOADPRODTEXT = 'Load Production Data'
LOADURRTEXT = 'Load URR Data (optional)'
LOADURRLOC = (9,3)
BIBTEXTEXT = 'Load bibtex file (optional)'
BIBTEXLOC = (10,3)
BLANKTEXT = "                                  "
FILEFONT = ("Helvetica",10)
FILECOLORBG = '#FFFFFF'
FILECOLOR = '#414141'
LOADFILELOC = (8,2)
TEXTURRLOC = (9,2)
LOADBIBTEXLOC = (10,2)

LATEXFIGURESFOLDER = 'Figures'
YES = 'Yes'
NO = 'No'

SUPERSCRIPTNUMS = {'0':u"\u2070",
                   '1':u"\u00B9",
                   '2':u"\u00B2",
                   '3':u"\u00B3",
                   '4':u"\u2074",
                   '5':u"\u2075",
                   '6':u"\u2076",
                   '7':u"\u2077",
                   '8':u"\u2078",
                   '9':u"\u2079"
                   }
SMALLFONT = ("Helvetica",10)
BUTTONFONT = ("Helvetica",12)
RADIOBUTTONFONT = ("Helvetica",12)
TITLEFONT = ("Helvetica",16)
DESCFONT = ("Helvetica",12)
TEXTFONT = ("Helvetica",12)
ACCEPTFONT = ("Helvetica",12)
TITLELOC = (0.5,0.05)
LOADPRODTITLETEXT = "Load Production Data"
LOADPRODDESCTEXT = ["The resource modelling software requires historic production data. Acceptable files are:",
                    "\u2022 Csv",
                    "\u2022 Excel.",
                    "You can optionally add in URR data as well. This is optional, and you can enter it in manually for each",
                    "individual projection if desired.",
                    "Please ensure that the production and URR data is in the correct structure.",
                    "See the example file for required layout."]
MODELTYPETITLETEXT = "Select modelling approach"
MODELTYPEDESCTEXT = ["The resource modelling softweare can either use GeRS-DeMo or Hubbert modelling approach."]
HUBBERTMODELBUTTONLOC = (8,5)
GERSDEMOBUTTONTEXT = 'GeRS-DeMo'
GERSDEMOBUTTONLOC = (8,4)
NEXTBUTTONTEXT = 'Next'
NEXTBUTTONLOC = (12,7)
SAVEBUTTONTEXT = 'Save'
SAVEBUTTONLOC = (12,6)
SAVEASBUTTONTEXT = 'Save As'
SAVEASBUTTONLOC = (12,5)


RUNSCENLOC = (12,2)
RUNSCENTEXT = 'Run Static Scenario'
RUNFULLTEXT = 'Run Model'
DEMANDKEY = '+DEMANDALL+DEMANDALL+DEMANDALL+ALL+DEMANDALL+DEMANDALL'
RECYCLEKEY = '+RECYCLE+RECYCLE+RECYCLE+ALL+RECYCLE+RECYCLE' 

EXPORTFAILED = "Export Failed"
EXPORTBUTTONLOC = (12,4)
EXPORTBUTTONTEXT = 'Export'
EXPORT = 'Export'
GENERICEXCELFILE = 'Generic_Excel_Template.xlsm'

EXPORTOPTIONMESSAGE = 'Export file format'
EXPORTTXT = 'Text Files'
EXPORTEXCEL = 'Excel'

BACKBUTTONTEXT = 'Back'
BACKBUTTONLOC = (12,0)
ACCEPTBUTTONLOC = (12,1)
EXCELBUTTONLOC = (4,0)
EXCELBUTTONTEXT = 'Write Results to Excel'
PDFBUTTONLOC = (6,0)
PDFBUTTONTEXT = ' Write Results to PDF'


ACCEPTBUTTONTEXT = 'Accept'
EXCELURRSTR = 'Excel_URR'
TITLECOLOR = 'black'
DESCCOLOR = 'black'
GERSDEMOMODELTITLE = 'GeRS-DeMo Introduction'
GERSDEMOMODELDESC = ["The easiest way to create a GERSDEMO model is to create the individual scenarios",
                     "in STATIC mode, and once finished with all the individual scenarios to input the",
                     "demand inputs then.",
                     "",
                     "As a result while building the scenarios the results shown will be STATIC mode results, and you will be",
                     "able to input the demand variables at the end."]
MINESBUTTONTEXT = 'New Mines'
FIELDSBUTTONTEXT = 'New Fields'
MINESBUTTONLOC = (12,2)
FIELDSBUTTONLOC = (12,3)
MINESTITLE = 'GeRS-DeMo Scenario'

PLACE = 'place'
PACK = 'pack'
GRID = 'grid'

CYCLEFRAME = 'CycleFrame'
SCROLLFRAME = 'ScrollFrame'
CYCLEBUTTON = 'CycleButton'
MAINDESC = 'Description'
SAVE = 'Save'
RUNSCEN = 'RunScen'

SAVEAS = 'SaveAs'
PRODIMG = 'ProdImg'
DIFFIMG = 'DiffImg'
CUMMIMG = 'CummImg'
BACK = 'Back'
EXCELTAG = 'exceltag'
PDFTAG = 'pdfout'
ACCEPT = 'Accept'
NEXT = 'Next'
LOAD = 'Load'
NEWMODEL = 'NewModel'
TITLE = 'Title'
LOADPROD = 'LoadProd'
LOADPRODLAB = 'LoadProdLab'
LOADURR = 'LoadURR'
LOADURRLAB = 'LoadURRLab'
LOADBIBTEX = 'LoadBibTex'
LOADBIBTEXLAB = 'LoadBibTexLab'
GLOBALINPUTSTITLETEXT = 'Global input'

NONEKEY = '+NONE+NONE+NONE+NONE+NONE+NONE'
POSTPROCESSORTITLETEXT = 'Post Processor Outputs'
POSTPROCESSORDESCTEXT = ''

FILENAMEDICT = {LOADPRODLAB:'prodfile',
                    LOADURRLAB:'urrfile',
                    LOADBIBTEXLAB:'bibtextfile'}

NEWMODELFRAME = 'newmodelframe'
MAINFRAME = 'mainframe'
SELECTMODELFRAME = 'selectmodelframe'
GERSDEMOFRAME = 'gersdemomainframe'
MINESFRAME = 'minesframe'
SCENARIOFRAME = 'scenarioframe'
EDITSCENARIOFRAME = 'editScenarioframe'
GLOBALINPUTSFRAME = 'globalinputsframe'
POSTPROCESSORFRAME = 'postprocessorframe'
WELLSFRAME = 'wellsframe'
HUBBERTFRAME = 'hubbertframe'
GOMPERTZFRAME = 'gompertzframe'
SIMPLEMODELSFRAME = 'HubbertGompertzmainframe'

EDITBUTTON = 'editbut'
EDITSCENARIOLABEL = 'Edit Scenario'
EDITSCENARIOLOC = (1,3)
ACCEPTBUTTON = 'acceptbut'
ACCEPTSCENARIOLABEL = 'Create Scenario'
ACCEPTSCENARIOLOC = (5,0)
LABEL = 'Label'
PRODIMGBUTTONLOC = (1,5)

PRODIMGBUTTONTEXT = "Production"
DIFFIMGBUTTONLOC = (1,6)
DIFFIMGBUTTONTEXT = "Difference"
CUMMIMGBUTTONLOC = (1,7)
CUMMIMGBUTTONTEXT = "Cumulative"
OUTPUTIMAGELOC = (2,5)
OUTPUTIMAGECOLSPAN = 4
OUTPUTIMAGEROWSPAN = 9
BLANKIMAGEFILE = 'test.png'

PROD = 'Prod'
CUMM = 'CUMM'
DIFF = 'Diff'
IMAGESIZE = (450,360)

URRCHOOSEMESSAGE = 'Please choose URR method'
URRINPUTMESSAGE = 'Please insert URR value'
URRINPUTERRORMESSAGE = 'Please insert a URR value (Numeric value)'
ENTERTEXT = 'Enter'

#URRBOX = 'URR_Input'
URRTITLE = 'URR_Title'
URRLABEL = 'URR_Label'
HLBUTTON = 'HLButton'
CURRURRTITLE = 'CurrURRLab'
CURRURRLABEL = 'Current URR:'
CURRURRVALUE = '0'
CURRURRVAL = 'CurrURR'

CURRURRVALUELOC = (6,3)
CURRURRTITLELOC = (6,2)


#URRBOXLOC = (6,1)
URRLABELLOC = (6,1)
URRBOXENTRIES = []
URRTITLELOC = (6,0)
URR = 'URR:'
HLBUTTONLOC = (6,2)

MANUALURRINPUT = 'Input URR'
HLTEXT = 'Apply HL'
HLBUTTONTEXT = 'Apply HL'
CYCLELOC = (7,0)
CYCLELOCSPAN = (3,3)
NEWCYCLELOC = (10,0)
CYCLESIZE = (150,250)




NEWCYCLEBUTTONLOC = (0.25,0.37)
NEWCYCLETEXT = "Add New Cycle"
CYCLEBUTTON = ' Cycle'
DELETEBUTTON = 'Delete '
DUPLICATEBUTTON = 'Duplicate'
SCENARIOBUTTON = ' New Scenario'
SCENARIOSLISTLOC = (3,2)
SCENARIOSLISTSPAN = (3,7)
CYCLETAG = 'Cycle'
SCENSCROLLSIZE = (150,600)
NEWSCENLOC = (6,2)
NEWSCENTEXT = "New Scenario"

SCENARIOCYCLETITLE = 'Cycle Inputs'
MINESCENARIOBUTTONLOC = (1,0)
WELLSCENARIOBUTTONLOC = (1,1)
MINE = 'Mine'
WELL = 'Well'
MINESCENARIOTEXT = 'Mine'
WELLSCENARIOTEXT = 'Field'
DISRUPTIONLOC = (8,2)
DISRUPTIONSPAN = (3,3)
NEWDISRUPTIONLOC = (11,2)
DISRUPTIONBUTTON = ' Disruption'
DISRUPTIONTAG = 'Disr'
DISRUPTIONSIZE = (150,250)
DISRUPTIONTEXT = "Add New Disruption"
DISRUPTIONTITLE = "Enter New Disruption"
SCENARIOTAG = 'scenario'
MAXFILEINPUTWIDTH = 75

class scenarioobj():
    def __init__(self,entityloc,titleloc,defaultfield,title,objtype):
        self.entityloc = entityloc
        self.titleloc = titleloc
        self.defaultfield = defaultfield
        self.title = title
        self.objtype = objtype

SCENTITLE = 'ScenTitle'
SCENARIO = 'Scen'

EDITSCENSHOW = ['scenarioScen','continentScen','countryScen','regionScen','subregionScen','mineralScen','submineralScen',
                'scenarioScenTitle','continentScenTitle','countryScenTitle','regionScenTitle','subregionScenTitle','mineralScenTitle','submineralScenTitle',
                'Title','acceptbut','Save','SaveAs','Back']



EDITSCENHIDE = ['scenarioScen','continentScen','countryScen','regionScen','subregionScen','mineralScen','submineralScen','acceptbut']


SCENARIOTERMS = {'scenario':scenarioobj((1,1),(1,0),'defaultscenarios','Scenario','auto'),
                 'continent':scenarioobj((2,1),(2,0),'defaultcontinents','Continent','auto'),
                 'country':scenarioobj((2,3),(2,2),'defaultcountrys','Country','auto'),
                 'region':scenarioobj((3,1),(3,0),'defaultregions','Region','auto'),
                 'subregion':scenarioobj((3,3),(3,2),'defaultsubregions','Sub-Region','auto'),
                 'mineral':scenarioobj((4,1),(4,0),'defaultminerals','Mineral','auto'),
                 'submineral':scenarioobj((4,3),(4,2),'defaultsubminerals','Sub-Mineral','auto')}

# why does disruptions act oddly.
#disruptionpage)
#scenariocyclepage)

GLOBALINPUTENTRYTAG = 'Inputentry_'
GLOBALINPUTTITLETAG = 'Inputtitle_'

INPUTTERMSENTRYTAG = 'Unitsentry_'
INPUTTERMSTITLETAG = 'Unitstitle_'


ENTRYTAGDISR = 'disrentry_'
TITLETAGDISR = 'disrtitle_'
### STEVE FIGURE OUT HERE




GLOBALINPUTSDESCTEXT = ['Input global parameters here']


STARTYEAR = 'StartYear'
MINEPRODLOW = 'MineProdLow'
MINEPRODHIGH = 'MineProdHigh'
MINELIFELOW = 'MineLifeLow'
MINELIFEHIGH = 'MineLifeHigh'
MINETIMECONST = 'MineTimeConstant'
MINERATECONST = 'MineRateConstant'
MINEONLINERATE = 'MinesonlineRate'
MINERAMPUPTIME = 'RampUpTime'
MINUPGRADETIME = 'MinUpgradeTime'
MINEURR = 'URR'
MINEABORT = 'Abort'
MAXCAPACITY = 'MaxCapacity'
YEARDYNCHANGE = 'YearDynamicChanges'
NEWMINEONLINERATE = 'NewMinesonlineRate'
NEWMAXCAPACITY = 'NewMaxCapacity'
WHENPARITY = 'WhenParity'
FIELDRAMPUPTIME = 'RampUpTime'
URRTOURRNAUGHT = 'URRtor0'
URRTOURRRED = 'URRtoURRr'
POWERNUM = 'powernum'
MAXFIELDSIZE = 'MaxFieldSize'
HOWMANYFIELD = 'HowManyFields'
HOWMANYOGRES = 'HowManyOGREs'
OGRERATE = 'OGRERate'
FIELDURR = 'URR'
FIELDABORT = 'Abort'
NEWWHENPARITY = 'NewWhenParity'
ENDYEAR = 'EndYear'
DISRRED = 'DisruptionRed'

SHUTOFFPERCENT = 'Shutoffpercent'
TIMEDELAY = 'timedelay'
INITIALDEMAND = 'initialdemand'
KD = 'kd'
MAXDEMAND = 'Maxdemand'
K1TERM = 'K1'
K2TERM = 'K2'
K3TERM = 'K3'
K4TERM = 'K4'
WHENFIELDS = 'Whenfields'
K5TERM = 'K5'
WHENUPGRADS = 'whenupgrads'
GAPDELAY = 'Gapdelay'
MAXPOP = 'Maxpop'
POPINIT = 'Popinit'
POPRATE = 'Poprate'
POPRATE = 'Poprate'
POPMIDYEAR = 'Popmidyear'
POPB = 'PopB'
POPGAMMA = 'PopGamma'
LIFEPRODUCT = 'LifeProduct'
MINRECYCLED = 'MinRecycled'
MAXRECYCLED = 'MaxRecycled'
RATERECYCLED = 'RateRecycled'
MIDYEAR = 'MidYear'

LOADUNITS = 'Units'

LOADUNITSTERM = {LOADUNITS:scenarioobj((6,6),(6,5),None,"Units",'Input')}



### STEVE PLONK THIS INTO THE TAB WITH BIBTEX TO GET UNITS INTO MODEL
UNITS = 'units'
INPUTUNITS = {UNITS:scenarioobj((11,3),(11,2),None,"Units:",'Input')}

DISRTERMS = {STARTYEAR:scenarioobj((2,1),(2,0),None,'Disruption Start Year','Input'),
             ENDYEAR:scenarioobj((3,1),(3,0),None,'Disruption End Year','Input'),
             DISRRED:scenarioobj((4,1),(4,0),None,'Disruption Reduction','Input')}

URREXCEL = 'ExcelURR'



MINE = 'Mine'
WELL = 'Well'
ENTRY = 'Entry_'
ENTRYTAG = {MINE:'Mine'+ENTRY,
            WELL:'Well'+ENTRY}
TITLETAG = {MINE:'MineTitle_',
            WELL:'WellTitle_'}
FAILTITLE = 'Failed to Read'
ERRORTITLE = 'Error reading file'
WARNINGCONTINUE = 'Use file despite warnings?'
WARNINGTITLE = 'Warnings reading file'

URRINPUTMETHOD = 'Manual_Input'
URRHLMETHOD = 'HL'
URRLOOKUP = 'supplied_File'
URRNONE = 'None'

EXCELSHEETS = {MINE:'Mine',
               WELL:'Field'}


SAVEFRAMELIST = ['hideelems','initialise','othervars']
SAVEELEMTYPES = ['autocomplete','entry','label','radiobutton']        



INPUTTERMS = {MINE:{STARTYEAR:scenarioobj((2,1),(2,0),None,'Start Year','Input'),
                    MINEPRODLOW:scenarioobj((2,3),(2,2),None,'Mine Prod Low','Input'),
                    MINEPRODHIGH:scenarioobj((2,5),(2,4),None,'Mine Prod High','Input'),
                    MINELIFELOW:scenarioobj((3,1),(3,0),None,'Mine Life Low','Input'),
                    MINELIFEHIGH:scenarioobj((3,3),(3,2),None,'Mine Life High','Input'),
                    MINETIMECONST:scenarioobj((3,5),(3,4),None,'Mine Time Constant','Input'),
                    MINERATECONST:scenarioobj((4,1),(4,0),None,'Mine Rate Constant','Input'),
                    MINEONLINERATE:scenarioobj((4,3),(4,2),None,'Mine Online Rate','Input'),
                    MINERAMPUPTIME:scenarioobj((4,5),(4,4),None,'Ramp Up Time','Input'),
                    MINUPGRADETIME:scenarioobj((5,1),(5,0),None,'Min Upgrade Time','Input'),
                    MINEURR:scenarioobj((5,3),(5,2),None,'URR','Input'),
                    MINEABORT:scenarioobj((5,5),(5,4),None,'Abort','Input'),
                    MAXCAPACITY:scenarioobj((6,1),(6,0),None,'Max Capacity','Input'),
                    YEARDYNCHANGE:scenarioobj((6,3),(6,2),None,'Year Dynamics Changes','Input'),
                    NEWMINEONLINERATE:scenarioobj((6,5),(6,4),None,'New Mine Online Rate','Input'),
                    NEWMAXCAPACITY:scenarioobj((7,1),(7,0),None,'New Max Capacity','Input')},
              WELL:{STARTYEAR:scenarioobj((2,1),(2,0),None,'Start Year','Input'),
                    WHENPARITY:scenarioobj((2,3),(2,2),None,'When Parity','Input'),
                    FIELDRAMPUPTIME:scenarioobj((2,5),(2,4),None,'Ramp Up Time','Input'),
                    URRTOURRNAUGHT:scenarioobj((3,1),(3,0),None,'URR To R0','Input'),
                    URRTOURRRED:scenarioobj((3,3),(3,2),None,'URR To URRr','Input'),
                    POWERNUM:scenarioobj((3,5),(3,4),None,'Power Number','Input'),
                    MAXFIELDSIZE:scenarioobj((4,1),(4,0),None,'Max Field Size','Input'),
                    HOWMANYFIELD:scenarioobj((4,3),(4,2),None,'How Many Fields','Input'),
                    HOWMANYOGRES:scenarioobj((4,5),(4,4),None,'How Many OGREs','Input'),
                    OGRERATE:scenarioobj((5,1),(5,0),None,'OGRE Rate','Input'),
                    FIELDURR:scenarioobj((5,3),(5,2),None,'URR','Input'),
                    FIELDABORT:scenarioobj((5,5),(5,4),None,'Abort','Input'),
                    MAXCAPACITY:scenarioobj((6,1),(6,0),None,'Max Capacity','Input'),
                    YEARDYNCHANGE:scenarioobj((6,3),(6,2),None,'Year Dynamics Changes','Input'),
                    NEWWHENPARITY:scenarioobj((6,5),(6,4),None,'New When Parity','Input'),
                    NEWMAXCAPACITY:scenarioobj((7,1),(7,0),None,'New Max Capacity','Input')}}
              
DEFAULTWIDTH = 20
DEFAULTINPUTWIDTH = 20
SCENARIOWIDTHS = 12
UNITWIDTH = 10
NCOLS = 2
INTERACTIONSFILE = 'Interaction.csv'
MINESFILE = 'Mine.csv'
FIELDSFILE = 'Field.csv'
CPPOUTPUTFILE = 'Raw_Summary.csv'

EXCELOUTPUTSHEET = 'Model_Results'

ATTRORDERIFBLANKORUNREADABLE = ['continent','country','region','subregion','mineral','submineral','unit']
REFTAG = 'reference:'
ALL = 'All'
ATTR = 'attributes'
REF = 'ref'
SCEN = 'URR'
SCENVARS = ['continent','country','region','subregion','mineral','submineral','scenario']

KEYORDER = ['continent','country','region','subregion','mineral','submineral']
FIRSTDATACOL = 9
NFIGS = 16
        
##Start Year	Mine Prod. Low	Mine Prod High	Mine Life Low	Mine Life High	Mine time constant	Mine rate constant	Minesonline rate	rampup time	Min upgrade time	URR	Abort	Max capacity	Year dynamic changes	new minesonline rate	new max capacity	Disruption Start	Disruption End	Disruption reduction	Disruption Start
##1941	3.00E-03	1.80E-02	30	70	1960	3.70E-02	6	4	10	0.88									
##
#### WELLS
##Start year	when parity 	ramp up time	URRtor0 	URRtoURRr 	powernum	Max field size 	how many fields 	how many OGREs	OGRE rate 	URR	Abort	Max Capacity	Year dynamics changes	new when parity	new maximum capacity	Disruption start year	Disruption end year	Disruption reduction


GOODEXCELNONALPHANUM = [' ','_','.']
PPDROPDOWNROW = 2
PPDROPDOWNCOL = 0 
THRESHOLD = 0.01
STEP = 25
YLABELSTART = 'Production ('
YLABELEND = ')'
N = 16

'''
disruptioncyclefunc
disruptionpage
scenariofunc
pagefunc
scenariocyclefunc
scenariocyclepage
'''

def stringy(val):
    if val is None:
        return None
    return str(val)

def floaty(val):
    if val is None:
        return None
    try:
        newval = float(val)
    except:
        newval = None
    return newval



GENERALORDER = [('Scenario',stringy),('Units',stringy)]

RUNSCENINPUTS = {SHUTOFFPERCENT:0.01,
                 TIMEDELAY:0,
                 INITIALDEMAND:0.083823,
                 KD:0.025,
                 MAXDEMAND:60,
                 K1TERM:0,
                 K2TERM:0,
                 K3TERM:0,
                 K4TERM:0,
                 WHENFIELDS:0.2,
                 K5TERM:0,
                 WHENUPGRADS:0.2,
                 GAPDELAY:20,
                 MAXPOP:10,
                 POPINIT:0.82,
                 POPRATE:0.023,
                 POPMIDYEAR:2007,
                 POPB:1.5,
                 POPGAMMA:2,
                 LIFEPRODUCT:0,
                 MINRECYCLED:0,
                 MAXRECYCLED:0,
                 RATERECYCLED:0,
                 MIDYEAR:0}

DEFAULTINTINPUTS = {SHUTOFFPERCENT:0.01,
                 TIMEDELAY:0,
                 INITIALDEMAND:0.083823,
                 KD:0.025,
                 MAXDEMAND:60,
                 K1TERM:0.15,
                 K2TERM:0.1,
                 K3TERM:0.01,
                 K4TERM:0.1,
                 WHENFIELDS:0.2,
                 K5TERM:0.1,
                 WHENUPGRADS:0.2,
                 GAPDELAY:20,
                 MAXPOP:10,
                 POPINIT:0.82,
                 POPRATE:0.023,
                 POPMIDYEAR:2007,
                 POPB:1.5,
                 POPGAMMA:2,
                 LIFEPRODUCT:0,
                 MINRECYCLED:0,
                 MAXRECYCLED:0,
                 RATERECYCLED:0,
                 MIDYEAR:0}


GLOBALINPUTTERMS = {SHUTOFFPERCENT:scenarioobj((2,1),(2,0),None,"Shut off Percent",'Input'),
                    TIMEDELAY:scenarioobj((2,3),(2,2),None,"Time Delay",'Input'),
                    INITIALDEMAND:scenarioobj((2,5),(2,4),None,"Initial Demand",'Input'),
                    KD:scenarioobj((3,1),(3,0),None,"Demand Rate",'Input'),
                    MAXDEMAND:scenarioobj((3,3),(3,2),None,"Max Demand",'Input'),
                    K1TERM:scenarioobj((3,5),(3,4),None,"K1",'Input'),
                    K2TERM:scenarioobj((4,1),(4,0),None,"K2",'Input'),
                    K3TERM:scenarioobj((4,3),(4,2),None,"K3",'Input'),
                    K4TERM:scenarioobj((4,5),(4,4),None,"K4",'Input'),
                    WHENFIELDS:scenarioobj((5,1),(5,0),None,"When Fields",'Input'),
                    K5TERM:scenarioobj((5,3),(5,2),None,"K5",'Input'),
                    WHENUPGRADS:scenarioobj((5,5),(5,4),None,"When Upgrade",'Input'),
                    GAPDELAY:scenarioobj((6,1),(6,0),None,"Gap Delay",'Input'),
                    MAXPOP:scenarioobj((6,3),(6,2),None,"Max Population",'Input'),
                    POPINIT:scenarioobj((6,5),(6,4),None,"Initial Population",'Input'),
                    POPRATE:scenarioobj((7,1),(7,0),None,"Population Rate",'Input'),
                    POPMIDYEAR:scenarioobj((7,3),(7,2),None,"Population Mid Year",'Input'),
                    POPB:scenarioobj((7,5),(7,4),None,"Population Const B",'Input'),
                    POPGAMMA:scenarioobj((8,1),(8,0),None,"Population Gamma",'Input'),
                    LIFEPRODUCT:scenarioobj((8,3),(8,2),None,"Life of Product",'Input'),
                    MINRECYCLED:scenarioobj((8,5),(8,4),None,"Min Recycled",'Input'),
                    MAXRECYCLED:scenarioobj((9,1),(9,0),None,"Max Recycled",'Input'),
                    RATERECYCLED:scenarioobj((9,3),(9,2),None,"Rate Recycled",'Input'),
                    MIDYEAR:scenarioobj((9,5),(9,4),None,"Mid Year",'Input')
                    }

ITERATIONFIELDSORDER = [(SHUTOFFPERCENT,floaty),(TIMEDELAY,floaty),(INITIALDEMAND,floaty),(KD,floaty),
                        (MAXDEMAND,floaty),(K1TERM,floaty),(K2TERM,floaty),(K3TERM,floaty),(K4TERM,floaty),(WHENFIELDS,floaty),
                        (K5TERM,floaty),(WHENUPGRADS,floaty),(GAPDELAY,floaty),(MAXPOP,floaty),(POPINIT,floaty),
                        (POPRATE,floaty),(POPMIDYEAR,floaty),(POPB,floaty),(POPGAMMA,floaty),(LIFEPRODUCT,floaty),
                        (MINRECYCLED,floaty),(MAXRECYCLED,floaty),(RATERECYCLED,floaty),(MIDYEAR,floaty)]

NUMBERSTR = 'number'
BASEEXCEL = [('continent',stringy),('country',stringy),('region',stringy),('mineral',stringy),('submineral',stringy),(NUMBERSTR,stringy)]
MINEEXCELSPECIFIC = [('StartYear',floaty),('MineProdLow',floaty),('MineProdHigh',floaty),('MineLifeLow',floaty),('MineLifeHigh',floaty),
             ('MineTimeConstant',floaty),('MineRateConstant',floaty),('MinesonlineRate',floaty),('RampUpTime',floaty),('MinUpgradeTime',floaty),
             ('URR',floaty),('Abort',floaty),('MaxCapacity',floaty),('YearDynamicChanges',floaty),('NewMinesonlineRate',floaty),('NewMaxCapacity',floaty)]
MINEORDER = BASEEXCEL + MINEEXCELSPECIFIC
EXCELSCENFIELDS = ['continent','country','region','mineral','submineral']
FIELDEXCELSPECIFIC = [('StartYear',floaty),('WhenParity',floaty),('RampUpTime',floaty),('URRtor0',floaty),('URRtoURRr',floaty),('powernum',floaty),
              ('MaxFieldSize',floaty),('HowManyFields',floaty),('HowManyOGREs',floaty),('OGRERate',floaty),('URR',floaty),('Abort',floaty),
              ('MaxCapacity',floaty),('YearDynamicChanges',floaty),('NewWhenParity',floaty),('NewMaxCapacity',floaty)]
EXCELCYCLEFIELDS = {MINE:MINEEXCELSPECIFIC,
                    WELL:FIELDEXCELSPECIFIC}

SCENARIOFIELDORDER = ['continent','country','region','subregion','mineral','submineral','unit']
UNITSTAB = 'UnitsTab'
FIELDORDER = BASEEXCEL + FIELDEXCELSPECIFIC
DISRUPTIONSNAME = 'Disruptions'
DISRUPTIONORDERS = [('StartYear',floaty),('EndYear',floaty),('DisruptionRed',floaty)]
SCENARIOTABS = 'Scen'
CONTINENTTERM = 'continent'
COUNTRYTERM = 'country'
MINERALTERM = 'mineral'
SUBMINERALTERM = 'submineral'
REGIONTERM = 'region'
SUBREGIONTERM = 'subregion'
URRTERM = 'urr'
DROPDOWNTAG = 'dropdown'
PPDDROPDOWNLABELROW = 1
PPDDROPDOWNLABELCOL = 0
PPDROPDOWNLABELTEXT = 'Select Scenario'
DROPDOWNLABELTAG = 'dropdownlabel'


class frame():
    def __init__(self,obj):
        self.clsname = 'frame'
        self.cls = None
        self.frame = obj
        self.frametype = None # use this to determine the scenarios.
        self.fromframe = None
        self.frombutton = None # use this to change the cyclebutton text
        self.elems = {}
        self.hide = False
        self.hideelems = [] ## elements to hide on openning
        self.initialise = True ### HERERERE STEVE USE THIS TO PARTIALLY SHOW FRAME.
        self.initfunc = None
        self.imageoptions = {}
        self.imageobj = None
        self.imgvar = None
        self.dropdownvar = None
        self.othervars = {}
        self.cycletypevar = None
        self.cycletypevarobjs = {}
        self.elemsincycle = 0
        self.cycleframeelems = {}
        self.scrollbox = {}
        self.buttonfunc = {}
        self.cyclename = {}
        self.scenariofunc = {}
        self.pagefunc = {}


class productioncls():
    def __init__(self):
        self.clsname = 'production'
        self.continent = ALL
        self.country = ALL
        self.region = ALL
        self.subregion = ALL
        self.mineral = ALL
        self.submineral = ALL
        self.unit = ALL
        self.years = []
        self.production = []
        self.cumulative = []
        self.prodovercumm = []
        self.outlierdist = []
        self.urr = 0
        self.peakyear = None
        self.color = None
        self.legend = None
        self.data = None

        

class URRcls():
    def __init__(self):
        self.clsname = 'URR'
        self.continent = ALL
        self.country = ALL
        self.region = ALL
        self.subregion = ALL
        self.mineral = ALL
        self.submineral = ALL
        self.scenario = None
        self.unit = ALL
        self.URR = None
        self.ref = None


class scenariocls():
    def __init__(self):
        self.clsname = 'scenario'
        self.scenario = ''
        self.continent = ''
        self.country = ''
        self.region = ''
        self.subregion = ''
        self.mineral = ''
        self.submineral = ''
        self.URR = None
        self.HLfrom = None
        self.HLto = None
        self.cycles = {}







class cyclecls():
    def __init__(self,row = 0,index = 0):
        self.clsname = 'cycle'
        self.row = row
        self.hide = False
        self.frame = None
        self.cycles = {}
        self.index = index
        self.delete = None
        self.button = None
        self.duplicate = None
    def grid(self):
        self.delete.elem.grid(row=self.row, column=self.delete.column,padx = self.delete.padx, pady = self.delete.pady, columnspan=self.delete.columnspan,rowspan = self.delete.rowspan, sticky=self.delete.sticky)
        self.button.elem.grid(row=self.row, column=self.button.column,padx = self.button.padx, pady = self.button.pady, columnspan=self.button.columnspan,rowspan = self.button.rowspan, sticky=self.button.sticky)
        self.duplicate.elem.grid(row=self.row, column=self.duplicate.column,padx = self.duplicate.padx, pady = self.duplicate.pady, columnspan=self.duplicate.columnspan,rowspan = self.duplicate.rowspan, sticky=self.duplicate.sticky)

    def grid_forget(self):
        self.delete.elem.grid_forget()
        self.button.elem.grid_forget()
        self.duplicate.elem.grid_forget()
    
class elemcls():
    # wrapper - just to make it easier to hide/show frames
    def __init__(self,frame = None,row = 0,column = 0, sticky = 'nsew',columnspan = 1,rowspan = 1,padx = 0,pady = 0,text = '',obj = None,
                 image = None,fg = DESCCOLOR,bg = FRAMECOLOR,font = DESCFONT,width = None,variable = None,lista = [],
                 anchor = 'center',justify = 'left',command = lambda a:a,height = 150,indicatoron = 0,value = '',widget = None):
        self.clsname = 'elem'
        self.text = text
        self.row = row
        self.column = column
        self.sticky = sticky
        self.columnspan = columnspan
        self.rowspan = rowspan
        self.padx = padx
        self.pady = pady
        self.obj = obj
        self.fg = fg
        self.bg = bg
        self.font = font
        self.width = width
        self.height = height
        self.justify = justify
        self.text = text
        self.command = command
        self.image = image
        self.indicatoron = indicatoron
        self.var = variable
        self.frame = frame
        if self.obj == 'label':
            self.elem = tk.Label(frame,image = image,text = text,fg = fg,bg = bg,font = font,width = width,justify = justify)
        elif self.obj == 'button':
            self.elem = tk.Button(frame,text = text,anchor = anchor,command = command,font = font,fg = fg,bg = bg,width = width,justify = justify)
        elif self.obj == 'frame':
            self.elem = tk.Frame(frame)
        elif self.obj == 'scrollframe':
            self.elem = ScrolledFrame(frame,height = height,width = width)
        elif self.obj == 'radiobutton':
            self.var.set(value)
            self.elem = tk.Radiobutton(frame,indicatoron = indicatoron,text = text,command = command,variable = variable,value=value,font = font)
        elif self.obj == 'autocomplete':
            # okoko need to send a textvar into here... not sure how.
            self.elem = AutocompleteEntry(lista = lista,master = frame,width = width)
        elif self.obj == 'entry':
            # need to add a textvar in here too.
            self.elem = tk.Entry(frame,width = width)
        elif self.obj == 'displaywidget':
            self.elem = widget(frame)
        elif self.obj == 'dropdown':
            self.elem = tk.OptionMenu(frame,variable, *lista)
        else:
            print(obj)
            raise Exception()


        
    def grid(self):
        self.elem.grid(row=self.row, column=self.column,padx = self.padx, pady = self.pady, columnspan=self.columnspan,rowspan = self.rowspan, sticky=self.sticky)


class imgcls():
    # wrapper - just to make it easier to hide/show frames
    def __init__(self,row = None,column = None,rowspan = None,columnspan = None,sticky = None,padx = None,pady = None,elem = None,img = None):
        self.clsname = 'img'
        self.row = row
        self.column = column
        self.sticky = sticky
        self.columnspan = columnspan
        self.rowspan = rowspan
        self.padx = padx
        self.pady = pady
        self.elem = elem
        self.img = img
    def grid(self):
        self.elem.grid(row=self.row, column=self.column,padx = self.padx, pady = self.pady, columnspan=self.columnspan,rowspan = self.rowspan, sticky=self.sticky)


class maincls():
    def __init__(self):
        self.clsname = 'main'
        self.production = {}
        self.modelresults = {} # model projection.
        self.outlierthreshold = {}
        self.URR = {}
        self.bibtextfile = None
        self.urrfile = None
        self.prodfile = None
        self.scenarios = {} ## scenariokey to scenario class
        self.scenarioframelookup = {} # this is the scenarios inputted into the model
        self.defaultscenarios = []    # this is a lookup between the scenarios and the corresponding frame.
        self.defaultcontinents = []
        self.defaultcountrys = []
        self.defaultregions = []
        self.defaultsubregions = []
        self.defaultminerals = []
        self.defaultsubminerals = []
        self.allcycles = 0
        self.savefile = None
        self.saveformat = None
        self.savefolder = None
        self.runscenfolder = basefolder()

def basefolder():
    folder = os.path.dirname(os.path.realpath('__file__'))
    paths = folder.split(os.sep)
    return paths[0]







class AutocompleteEntry(tk.Entry):
    '''
    This class is slightly modified from:
    http://code.activestate.com/recipes/578253-an-entry-with-autocompletion-for-the-tkinter-gui/
    '''
    def __init__(self, lista, *args, **kwargs):
        
        tk.Entry.__init__(self, *args, **kwargs)
        self.lista = lista        
        self.var = self["textvariable"]
        if self.var == '':
            self.var = self["textvariable"] = tk.StringVar()

        self.var.trace('w', self.changed)
        self.bind("<Right>", self.selection)
        self.bind("<Up>", self.up)
        self.bind("<Down>", self.down)
        self.bind("<Return>", self.selection)
        self.bind("<Tab>",self.selection)
        self.lb_up = False

    def changed(self, name, index, mode):  

        if self.var.get() == '':
            if 'lb' in dir(self):
                self.lb.destroy()
            self.lb_up = False
        else:
            words = self.comparison()
            if words:            
                if not self.lb_up:
                    self.lb = tk.Listbox()
                    self.lb.bind("<Double-Button-1>", self.selection)
                    self.lb.bind("<Right>", self.selection)
                    self.lb.place(x=self.winfo_x(), y=self.winfo_y()+self.winfo_height())
                    self.lb_up = True
                
                self.lb.delete(0, tk.END)
                for w in words:
                    self.lb.insert(tk.END,w)
            else:
                if self.lb_up:
                    self.lb.destroy()
                    self.lb_up = False

        
    def selection(self, event):
        if self.lb_up:
            self.var.set(self.lb.get(tk.ACTIVE))
            self.lb.destroy()
            self.lb_up = False
            self.icursor(tk.END)

    def up(self, event):

        if self.lb_up:
            if self.lb.curselection() == ():
                index = '-1'
            else:
                index = self.lb.curselection()[0]
            if index != '0':                
                self.lb.selection_clear(first=index)
                index = str(int(index)-1)                
                self.lb.selection_set(first=index)
                self.lb.activate(index) 

    def down(self, event):

        if self.lb_up:
            if self.lb.curselection() == ():
                index = '-1'
            else:
                index = self.lb.curselection()[0]
            if index != tk.END:
                self.lb.selection_clear(first=index)
                index = str(int(index)+1)        
                self.lb.selection_set(first=index)
                self.lb.activate(index) 

    def comparison(self):
        pattern = re.compile('.*' + self.var.get() + '.*')
        return [w for w in self.lista if re.match(pattern, w)]

##### GLOBALS ABOVE.

def scenariospecificdup(frames,oldframe,newframe):
    return     

def disrspecificdup(frames,oldframe,newframe):
    return

def cyclespecificdup(frames,oldframe,newframe):
    var = frames[oldframe].cycletypevar.get()
    frames[newframe].cycletypevar.set(var)


PAGESPECIFICDUPLICATION = {CYCLETAG:cyclespecificdup,
                           DISRUPTIONTAG:disrspecificdup,
                           SCENARIOTAG:scenariospecificdup}



def initialiseframe():
    root = tk.Tk()
    root.title(PROGRAMTITLE)
    #root.geometry("500x500+10+10")
    root.grid_rowconfigure(0,weight = 1)
    root.grid_columnconfigure(0,weight=1)

    
    mainclass = maincls()
    frames = {}

    

    container = tk.Frame(root)
    container.grid(sticky = 'nswe')
    container.grid_rowconfigure(0,weight = 1)
    container.grid_columnconfigure(0,weight = 1)
    #container.grid(sticky = 'nswe',columnspan = GRIDSIZE[1],rowspan = GRIDSIZE[0])
##    for row in range(0,GRIDSIZE[0]):
##        container.grid_rowconfigure(row, weight = 1, minsize=MINROWSIZE)
##    for col in range(0,GRIDSIZE[1]):
##        container.grid_columnconfigure(col, weight = 1,minsize = MINCOLSIZE)
##


    canvas = tk.Canvas(container)


    
    #root.minsize(MINFRAMESIZE[0],MINFRAMESIZE[1])
    #canvas = tk.Canvas(container,scrollregion=(0,0,3000,3000))
    for frame in [MAINFRAME,NEWMODELFRAME,SELECTMODELFRAME,GLOBALINPUTSFRAME,POSTPROCESSORFRAME]:
        frames[frame] = createemptyframe(root,container,FRAMECOLOR)
    createmainframe(root,container,canvas,frames,mainclass)
    newmodelpage(root,container,canvas,frames,mainclass)
    selectmodeltypepage(root,container,canvas,frames,mainclass)
    globalinputspage(root,container,canvas,frames,mainclass)
    postprocessorpage(root,container,canvas,frames,mainclass)
    loadframe(root,frames,mainclass,MAINFRAME,container)
    
    return root


def createemptyframe(root,container,color,fromframe = None,frombutton = None):
##    svbar = ttk.Scrollbar(container,orient="vertical")
##    shbar = ttk.Scrollbar(container,orient="horizontal")
##    svbar.grid(column = 9,rowspan = 13,sticky = 'ns')
##    shbar.grid(row = 13, columnspan = 9,sticky = 'ew')
    #newframe = tk.Frame(root,yscrollcommand = svbar.set,xscrollcommand = shbar.set)
    #newframe = tk.Frame(container)
    newframe = tk.Frame(container)
    #outerframe = tk.Frame(container)
    #scrollframe = ScrolledFrame(outerframe)
    #newframe = scrollframe.display_widget(tk.Frame)
    #outerframe = ScrolledFrame(container,height = 1000,width = 1000)
    #newframe = outerframe.display_widget(tk.Frame)
    #newframe = DoubleScrolledFrame(root)
    #fm = tk.Frame(root)
    ## HERERE at some point figure out how to make the canvas scrollable.
    #newframe = ScrolledFrame(fm,height = FRAMESIZE[0],width = FRAMESIZE[1])
    #newframe.bind_arrow_keys(fm)
    #newframe.bind_scroll_wheel(fm)
    #newframe.grid(in_=container,columnspan = GRIDSIZE[1],rowspan = GRIDSIZE[0],sticky = 'nsew')
    newframe.grid(in_=container,sticky = 'nswe')
    newframe.config(background = color)

    for row in range(0,GRIDSIZE[0]):
        newframe.grid_rowconfigure(row, weight = 1, minsize=MINROWSIZE)
    for col in range(0,GRIDSIZE[1]):
        newframe.grid_columnconfigure(col, weight = 1,minsize = MINCOLSIZE)
    newframecls = frame(newframe)
    newframecls.fromframe = fromframe
    newframecls.frombutton = frombutton
    return newframecls



def savemodel(root,frames,mainclass):
    if mainclass.savefile is None or mainclass.savefile == '':
        value = saveasmodel(root,frames,mainclass)
        if value == 0:
            return
    maindict = vars(mainclass)
    tmp = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
    tmpfolder = tmp.name
    if os.path.exists(mainclass.savefile):
        try:
            os.remove(mainclass.savefile)
        except:
            popupmessage(SAVEFAILED,'Failed to delete existing file')
            return
    try:
        mainfile = open(tmpfolder+os.sep+MAINCLASSFILE,'wb')
    except:
        popupmessage(SAVEFAILED,'Failed to create mainclass')
        return
    try:
        pickle.dump(mainclass,mainfile)
        mainfile.close()
    except:
        popupmessage(SAVEFAILED,'Failed to pickle mainclass')
        return
    saveframes = {}
    for key in frames:
        if frames[key].hide:
            continue
        saveframes[key] = {}
        for term in SAVEFRAMELIST:
            saveframes[key][term] = getattr(frames[key],term)
        saveframes[key]['elems'] = {}
        saveframes[key]['autoelems'] = {}
        saveframes[key]['labels'] = {}
        saveframes[key]['radiobutton'] = {}
        for k in frames[key].elems:
            if frames[key].elems[k].obj in SAVEELEMTYPES:
                if frames[key].elems[k].obj == 'autocomplete':
                    saveframes[key]['autoelems'][k] = frames[key].elems[k].elem.get()
                elif frames[key].elems[k].obj == 'label':
                    ## HEREREREERE issues doesn't work with label elements.
                    saveframes[key]['labels'][k] = frames[key].elems[k].elem.cget('text')
                elif frames[key].elems[k].obj == 'radiobutton':
                    saveframes[key]['radiobutton'][k] = frames[key].elems[k].var.get()
                else:
                    saveframes[key]['elems'][k] = frames[key].elems[k].elem.get()
        cycleframedict = {}
        for tag in frames[key].cycleframeelems.keys():
            cycleframedict[tag] = sorted(list(frames[key].cycleframeelems[tag].keys()))
        saveframes[key]['cycleframeelems'] = cycleframedict
    try:
        saveframesfile = open(tmpfolder+os.sep+SAVEFRAMESFILE,'wb')
    except:
        popupmessage(SAVEFAILED,'Failed to create frames file')
        return
    try:
        pickle.dump(saveframes,saveframesfile)
        saveframesfile.close()
    except:
        popupmessage(SAVEFAILED,'Failed to pickle frames')
        return
    try:
        mapping = getscenariomap(frames)
    except:
        popupmessage(SAVEFAILED,'Failed to create mapping file')
        return
    try:
        savemappingfile = open(tmpfolder+os.sep+SAVEMAPPINGFILE,'wb')
    except:
        popupmessage(SAVEFAILED,'Failed to create frames file')
        return
    try:
        pickle.dump(mapping,savemappingfile)
        savemappingfile.close()
    except:
        popupmessage(SAVEFAILED,'Failed to save mapping file')
        return
    try:
        createzip(tmpfolder,mainclass.savefile)
    except:
        popupmessage(SAVEFAILED,'Failed to create file')
        return
    

    
def saveasmodel(root,frames,mainclass):
    filename =  filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("GeRS-DeMo","*.GRS"),))
    if filename is None or filename == '':
        return 0
    path,ext = os.path.splitext(filename)
    if ext == '':
        filename = filename + '.GRS'
        ext = '.GRS'    
    mainclass.savefile = filename
    mainclass.saveformat = ext
    mainclass.savefolder = os.path.dirname(filename)
    savemodel(root,frames,mainclass)
    return 1


class exportversion(object):
    def __init__(self,root):
        self.option = None
        self.filetypes = None
        self.exportoption = tk.Toplevel(root)
        self.exportoption.geometry(POPUPGEOM)
        self.label = tk.Label(self.exportoption, text=EXPORTOPTIONMESSAGE)
        self.label.grid(padx=PADX, pady=PADY,sticky = 'nsw',columnspan = 2)
        
        button_txt = tk.Button(self.exportoption, text=EXPORTTXT, command = lambda : self.exportfunc('text',('*.zip',)),width = DEFAULTINPUTWIDTH)
        button_txt.grid(padx = PADX, pady = PADY,sticky = 'nswe',row = 1,column = 0)
        
        button_excel = tk.Button(self.exportoption, text=EXPORTEXCEL, command = lambda : self.exportfunc('excel',('*.xlsx','*.xls')),width = DEFAULTINPUTWIDTH)
        button_excel.grid(padx = PADX, pady = PADY,sticky = 'nswe',row = 1,column = 1)
        
    def exportfunc(self,exportopt,extensions):
        self.option = exportopt
        self.filetypes = []
        for elem in extensions:
            self.filetypes.append((exportopt,elem))
        self.exportoption.destroy()


def getexcellines(frames,mainclass):
    # create scenariomap
    scenmap = getscenariomap(frames)
    lines = {}
    for skey in scenmap:
        scen = frames[skey].elems['scenarioScen'].elem.var.get()
        basename,scenario = getbasename(frames[skey])
        if scenario not in lines:
            lines[scenario] = {MINE:[],WELL:[]}
        row = 0
        for ckey in scenmap[skey]:
            row += 1
            linetype,line = gettextline(basename,frames,frames[ckey],scenmap[skey][ckey],row)
            if linetype is None:
                continue
            lines[scenario][linetype].append(line)
    return lines


def exportmodel(root,frames,mainclass):
    lines = getexcellines(frames,mainclass)
    for scenario in lines:
        exportmodel_sub(root,frames,mainclass,lines[scenario],scenario)

def exportmodel_sub(root,frames,mainclass,lines,scenario):
    fname =  filedialog.asksaveasfilename(initialdir = "/",title = "Export to "+scenario,filetypes = (('excel','*.xlsm'),))
    path,ext = os.path.splitext(fname)
    if path == '':
        return
    if ext.upper() not in ('.XLSM',):
        answer = acceptlocation(root,"Incorrect file type, export to: "+path+'.xlsm')
        root.wait_window(answer.window)
        if answer.answer is None:
            return
        if answer.answer:
            response = exportexcel(frames,mainclass,path,'.xlsm',lines,scenario)
            popupmessage(EXPORTFAILED,response)
        return
    response = exportexcel(frames,mainclass,path,ext,lines,scenario)
    popupmessage(EXPORTFAILED,response)



def exportexcel(frames,mainclass,path,ext,lines,scenario):
    # delete folder/file if exists
    valid = safedelete(path+ext)
    if not valid:
        return "Failed to delete folder"
    try:

        shutil.copyfile(GENERICEXCELFILE,path+ext)
    except:
        return "Failed copy paste"
    try:
        wb = openpyxl.load_workbook(filename=path+ext, read_only=False, keep_vba=True)
    except:
        return "Failed to open"
    row = 0
    col = 2
    sht = wb[INTERACTIONTAB]
    for pair in ITERATIONFIELDSORDER:
        row += 1
        term,func = pair
        val = frames[GLOBALINPUTSFRAME].elems[term + GLOBALINPUTENTRYTAG].elem.get()
        sht.cell(row=row,column=col).value = val
    val = frames[NEWMODELFRAME].elems[LOADUNITS+INPUTTERMSENTRYTAG].elem.get()
    sht = wb[GENERALTAB]
    sht.cell(2,2).value = val
    sht.cell(1,2).value = scenario
    for key in lines:
        row = 1
        sht = wb[EXCELSHEETS[key]]
        for line in lines[key]:
            row += 1
            col = 0
            for elem in line:
                col += 1
                sht.cell(row=row,column = col).value = elem
    try:
        wb.save(path+ext)
    except:
        return "Failed to save"
    





def popupmessage(title,txt):
    if txt is None:
        return
    if txt == '':
        return
    messagebox.showwarning(title = title,message = txt)
    
    #### todo --> create a popup with the message txt

def safedelete(path):
    if not os.path.exists(path):
        return True
    if os.path.isdir(path):
        try:
            shutil.rmtree(path)
        except:
            print('badness')
            print(path)
            return False
        return True
    if os.path.isfile(path):
        try:
            os.unlink(path)
        except:
            print('badness2')
            print(path)
            return False
        return True

def safecreatedir(path):
    try:
        os.mkdir(path)
    except:
        return False
    return True

def safecreatefile(path):
    try:
        f = open(path,'w',endline = '')
    except:
        return None,False
    return f,True


def relkeys(frames,tag):
    keys = {}
    for key in frames:
        if not key.startswith(tag):
            continue
        if frames[key].hide:
            continue
        fromframe = frames[key].fromframe
        keys[key] = fromframe
    return keys


def getscenariomap(frames):
    scenkeys = relkeys(frames,SCENARIOTAG)
    cyckeys = relkeys(frames,CYCLETAG)
    disrkeys = relkeys(frames,DISRUPTIONTAG)
    
    mapping = {}
    for key in scenkeys:
        mapping[key] = {}
    for key in cyckeys:
        if cyckeys[key] not in mapping:
            continue
        mapping[cyckeys[key]][key] = []
    for dkey in disrkeys:
        ckey = disrkeys[dkey]
        if ckey not in cyckeys:
            continue
        skey = cyckeys[ckey]
        if skey not in mapping:
            continue
        if ckey not in mapping[skey]:
            continue
        mapping[skey][ckey].append(dkey)
    return mapping

def getcleanstr(term):
    cterm = term.replace('+','_')
    cterm = cterm.replace(',',';')
    return cterm

def getbasename(frame):
    scenario = frame.elems['scenarioScen'].elem.var.get()
    terms = []
    for key in ('continentScen','countryScen'):
        terms.append(frame.elems[key].elem.var.get())
    jointerms = []
    for key in ('regionScen','subregionScen'):
        jointerms.append(frame.elems[key].elem.var.get())
    terms.append('$'.join(jointerms))
    for key in ('mineralScen','submineralScen'):
        terms.append(frame.elems[key].elem.var.get())
    newterms = []
    for term in terms:
        cterm = getcleanstr(term)
        newterms.append(cterm)
    return newterms,scenario

def gettextline(basename,frames,frame,disruptions,row):
    '''
    Need to work on this - want to run checks on it?
    '''
    line = basename + [str(row)]
    termkey = frame.cycletypevar.get()
    if termkey == MINE:        
        terms = (STARTYEAR,MINEPRODLOW,MINEPRODHIGH,MINELIFELOW,MINELIFEHIGH,MINETIMECONST,MINERATECONST,MINEONLINERATE,
                 MINERAMPUPTIME,MINUPGRADETIME,MINEURR,MINEABORT,MAXCAPACITY,YEARDYNCHANGE,NEWMINEONLINERATE,NEWMAXCAPACITY)
    elif termkey == WELL:
        terms = (STARTYEAR,WHENPARITY,FIELDRAMPUPTIME,URRTOURRNAUGHT,URRTOURRRED,POWERNUM,MAXFIELDSIZE,HOWMANYFIELD,
                 HOWMANYOGRES,OGRERATE,FIELDURR,FIELDABORT,MAXCAPACITY,YEARDYNCHANGE,NEWWHENPARITY,NEWMAXCAPACITY)
    else:
        return None,None
    for term in terms:
        key = term + ENTRYTAG[termkey]
        line.append(frame.elems[key].elem.get())
    for dkey in disruptions:
        line.append(frames[dkey].elems[STARTYEAR+ENTRYTAGDISR].elem.get())
        line.append(frames[dkey].elems[ENDYEAR+ENTRYTAGDISR].elem.get())
        line.append(frames[dkey].elems[DISRRED+ENTRYTAGDISR].elem.get())
    return termkey,line
                    





        
class acceptlocation(object):
    def __init__(self,root,text):
        self.window = tk.Toplevel(root)
        self.window.geometry(POPUPGEOM)
    
        self.answer = None
        self.label = tk.Label(self.window, text=text)
        self.label.grid(padx=PADX, pady=PADY,sticky = 'nsw',columnspan = 2)
        
        button_yes = tk.Button(self.window, text=YES, command = lambda : self.answerfunc(True),width = DEFAULTINPUTWIDTH)
        button_yes.grid(padx = PADX, pady = PADY,sticky = 'nswe',row = 1,column = 0)
        
        button_no = tk.Button(self.window, text=NO, command = lambda : self.answerfunc(False),width = DEFAULTINPUTWIDTH)
        button_no.grid(padx = PADX, pady = PADY,sticky = 'nswe',row = 1,column = 1)
        
    def answerfunc(self,term):
        self.answer = term
        self.window.destroy()
    
    
        




def recursive_zip(zipf, directory, folder = ""):
   for item in os.listdir(directory):
      if os.path.isfile(directory + os.sep + item):
         zipf.write(directory + os.sep + item, folder + os.sep + item)
      elif os.path.isdir(directory + os.sep + item):
         recursive_zip(zipf, directory + os.sep + item, folder + os.sep + item)
         
def createzip(fname,filename):
    zipf = zipfile.ZipFile(filename,'w',compression = zipfile.ZIP_DEFLATED)
    recursive_zip(zipf,fname)
    zipf.close()

def reinitisialise(mainclass,mcls,frames,root,currframe,container,canvas):
    for term in vars(mainclass):
        setattr(mainclass,term,getattr(mcls,term))
    #    print(term)
    #    print(getattr(mcls,term))
    framekeys = list(frames.keys())

    forgetframe(root,frames,currframe)
    for frame in framekeys:
        if frame in [MAINFRAME,NEWMODELFRAME,GLOBALINPUTSFRAME,POSTPROCESSORFRAME]:
            continue
        del frames[frame]        
    if NEWMODELFRAME in frames:
        for key in FILENAMEDICT:
            if key in frames[NEWMODELFRAME].elems:
                frames[NEWMODELFRAME].elems[key].elem.config(text = getattr(mainclass,FILENAMEDICT[key]))
    frames[SELECTMODELFRAME] = createemptyframe(root,container,FRAMECOLOR)
    selectmodeltypepage(root,container,canvas,frames,mainclass)
    return frames
    
    
def loadmodelfile(root,container,canvas,frames,mainclass,currframe):
    filename = filedialog.askopenfilename(filetypes = (("GeRS-DeMo","*.GRS"),))
    if filename == '' or filename is None:
        return
    tmp = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
    tmpfolder = tmp.name
    
    if tmpfolder is None:
        popupmessage(LOADFAILED,'Could not create temporary folder')
        return
    try:
        zippy = zipfile.ZipFile(filename,'r')
        zippy.extractall(tmpfolder)
    except:
        popupmessage(LOADFAILED,'Could not unzip file')
        return
    try:
        mainfile = open(tmpfolder+os.sep+MAINCLASSFILE,'rb')
        mcls = pickle.load(mainfile)
        mainfile.close()
        frmfile = open(tmpfolder+os.sep+SAVEFRAMESFILE,'rb')
        frm = pickle.load(frmfile)
        frmfile.close()
        mpfile = open(tmpfolder+os.sep+SAVEMAPPINGFILE,'rb')
        mp = pickle.load(mpfile)
        mpfile.close()
    except:
        popupmessage(LOADFAILED,'Could not read files')
        return
    frames = reinitisialise(mainclass,mcls,frames,root,currframe,container,canvas)
    mplookup = {}
    for scenkey in mp:
        newframe = frames[SELECTMODELFRAME].elems[SCENARIOTAG+CYCLEBUTTON].command()
        mplookup[scenkey] = newframe
        for cyckey in mp[scenkey]:
            nextframe = frames[newframe].elems[CYCLETAG+CYCLEBUTTON].command()
            mplookup[cyckey] = nextframe
            for disrkey in mp[scenkey][cyckey]:
                lastframe = frames[nextframe].elems[DISRUPTIONTAG+CYCLEBUTTON].command()
                mplookup[disrkey] = lastframe
        
    #for frame in frm:
    #    if frame == GLOBALINPUTSFRAME:
    #        for key in frm[frame]['elems']:
    #            frames[mplookup[frame]].elems[key].elem.delete(0,tk.END)
    #            frames[mplookup[frame]].elems[key].elem.insert(0,frm[frame]['elems'][key])

    skipframes = [MAINFRAME,NEWMODELFRAME,SELECTMODELFRAME,GLOBALINPUTSFRAME,POSTPROCESSORFRAME]
    for frame in frm:
        if frame in skipframes:
            continue
        if frame not in mplookup:
            continue
        for term in SAVEFRAMELIST:
            setattr(frames[mplookup[frame]],term,frm[frame][term])
            ### HERERER issue seems to be that there is an autoelement and a label element, need to add the label elements.
        for key in frm[frame]['elems']:
            frames[mplookup[frame]].elems[key].elem.delete(0,tk.END)
            frames[mplookup[frame]].elems[key].elem.insert(0,frm[frame]['elems'][key])
        for key in frm[frame]['autoelems']:
            frames[mplookup[frame]].elems[key].elem.delete(0,tk.END)
            frames[mplookup[frame]].elems[key].elem.insert(0,frm[frame]['autoelems'][key])
            frames[mplookup[frame]].elems[key].elem.var.set(frm[frame]['autoelems'][key])
            if 'lb' in dir(frames[mplookup[frame]].elems[key].elem):
                frames[mplookup[frame]].elems[key].elem.lb.destroy()
        for key in frm[frame]['labels']:
            frames[mplookup[frame]].elems[key].elem.config(text = frm[frame]['labels'][key])
        for key in frm[frame]['radiobutton']:
            frames[mplookup[frame]].elems[key].var.set(frm[frame]['radiobutton'][key])

    for key in GLOBALINPUTTERMS:
        frames[GLOBALINPUTSFRAME].elems[key+GLOBALINPUTENTRYTAG].elem.delete(0,tk.END)
        frames[GLOBALINPUTSFRAME].elems[key+GLOBALINPUTENTRYTAG].elem.insert(0,frm[GLOBALINPUTSFRAME]['elems'][key+GLOBALINPUTENTRYTAG])

    frames[NEWMODELFRAME].elems[LOADUNITS+INPUTTERMSENTRYTAG].elem.delete(0,tk.END)
    frames[NEWMODELFRAME].elems[LOADUNITS+INPUTTERMSENTRYTAG].elem.insert(0,frm[NEWMODELFRAME]['elems'][LOADUNITS+INPUTTERMSENTRYTAG])
    #STEVIESTEVE
    for scenkey in mp:
        nframe = mplookup[scenkey]
        acceptscenario(root,frames,mainclass,nframe,container,skipurr=True,suppresswarnings = True)
        forgetframe(root,frames,nframe)
    for lab in (LOADPRODLAB,LOADURRLAB,LOADBIBTEXLAB):
        frames[NEWMODELFRAME].elems[lab].elem.config(text = frm[NEWMODELFRAME]['labels'][lab]) 
    loadframe(root,frames,mainclass,SELECTMODELFRAME,container)

def getexcelcolfunc(order,elem):
    i = 0
    for pair in order:
        i += 1
        if pair[0] == elem:
            return i,pair[1]

        
def getexceldata_sub(sht,row,order,scenfields):
    scen = []
    for elem in scenfields:
        col,func = getexcelcolfunc(order,elem)
        val = func(sht.cell(row,col).value)
        scen.append(val)
    if list(set(scen)) == [None]:
        return None,None
    scen = '|'.join(scen)
    record = {}
    i = 0
    for pair in order:
        elem,func = pair
        i += 1
        if i > sht.max_column:
            return None,None
        val = func(sht.cell(row,i).value)
        record[elem] = val
    i += 1
    record[DISRUPTIONSNAME] = []
    while True:
        if i+len(DISRUPTIONORDERS)-1 > sht.max_column:
            return scen,record
        disr = []
        for j in range(0,len(DISRUPTIONORDERS)):
            elem,func = DISRUPTIONORDERS[j]
            val = func(sht.cell(row,i+j).value)
            disr.append(val)
        if list(set(disr)) == [None]:
            return scen,record
        record[DISRUPTIONSNAME].append(disr)
        i = i + len(DISRUPTIONORDERS)
    
def getexceldata(sht,mineorwell,data,order,scenfields):
    for i in range(1,sht.max_row):
        scen,record = getexceldata_sub(sht,i+1,order,scenfields)
        if scen is None:
            continue
        if scen not in data:
            data[scen] = []
        data[scen].append((mineorwell,record))
    return data

def getdatafromexcel(wb):
    exceldata = {}
    intsht = wb[INTERACTIONTAB]
    gensht = wb[GENERALTAB]
    exceldata[INTERACTIONTAB] = {}
    i = 0
    for pair in ITERATIONFIELDSORDER:
        i += 1
        elem,func = pair
        val = func(intsht.cell(i,2).value)
        exceldata[INTERACTIONTAB][elem] = val
    exceldata[SCENARIOTABS] = {}
    exceldata[SCENARIOTABS] = getexceldata(wb[MINETAB],MINE,exceldata[SCENARIOTABS],MINEORDER,EXCELSCENFIELDS)
    exceldata[SCENARIOTABS] = getexceldata(wb[FIELDTAB],WELL,exceldata[SCENARIOTABS],FIELDORDER,EXCELSCENFIELDS)
    exceldata[UNITSTAB] = {}
    i = 0
    for pair in GENERALORDER:
        i += 1
        term,func = pair
        val = func(gensht.cell(i,2).value)
        exceldata[UNITSTAB][term] = val
    return exceldata

        
def importmodelfile(root,container,canvas,frames,mainclass,currframe):
    filename = filedialog.askopenfilename(filetypes = (("Excel","*.xlsm"),("Excel",'*.xlsx')))
    if filename == '' or filename is None:
        return
    try:
        wb = openpyxl.load_workbook(filename)
    except:
        popupmessage(IMPORTFAILED,'Could not read file')
        return
    #response = InputScenarioPopup(root,"Scenario")
    #root.wait_window(response.inputscenario)
    #scenario = response.scenario
    #FHERERERERERERER
    #sheets = wb.get_sheet_names()
    sheets = wb.sheetnames
    for sht in IMPORTTABS:
        if sht not in sheets:
            popupmessage(IMPORTFAILED,'Could not find sheet: '+sht)
            return

    ## OK HERE --> need to import the data correctly. 
    mcls = maincls()
    frames = reinitisialise(mainclass,mcls,frames,root,currframe,container,canvas)
    excelscens = getdatafromexcel(wb)
    #2) load all the data from excel into something temporary
    #3) load each scenario.
    # {'scenario0': {'Cycle1': ['Disr2']}}
    for term in vars(mainclass):
        setattr(mainclass,term,getattr(mcls,term))
    framekeys = list(frames.keys())
    forgetframe(root,frames,currframe)
    for frame in framekeys:
        if frame in [MAINFRAME,NEWMODELFRAME,GLOBALINPUTSFRAME,POSTPROCESSORFRAME]:
            continue
        del frames[frame]        
    if NEWMODELFRAME in frames:
        for key in FILENAMEDICT:
            if key in frames[NEWMODELFRAME].elems:
                frames[NEWMODELFRAME].elems[key].elem.config(text = getattr(mainclass,FILENAMEDICT[key]))       

    frames[NEWMODELFRAME].elems[LOADUNITS+INPUTTERMSENTRYTAG].elem.delete(0,'end')
    units = excelscens[UNITSTAB]['Units']
    if units is not None:
        frames[NEWMODELFRAME].elems[LOADUNITS+INPUTTERMSENTRYTAG].elem.insert(0,units)

    for key in GLOBALINPUTTERMS:
        term = excelscens[INTERACTIONTAB][key]
        frames[GLOBALINPUTSFRAME].elems[key + GLOBALINPUTENTRYTAG].elem.delete(0,'end')
        if term is not None:
            frames[GLOBALINPUTSFRAME].elems[key + GLOBALINPUTENTRYTAG].elem.insert(0,term)

    frames[SELECTMODELFRAME] = createemptyframe(root,container,FRAMECOLOR)
    selectmodeltypepage(root,container,canvas,frames,mainclass)    
    


    for scen in excelscens[SCENARIOTABS]:
        addscenario(excelscens[UNITSTAB]['Scenario'],scen,excelscens[SCENARIOTABS][scen],root,container,canvas,frames,mainclass,SELECTMODELFRAME,SELECTMODELFRAME)
    loadframe(root,frames,mainclass,SELECTMODELFRAME,container)
    return

def cleanstring(rawstr):
    string = []
    for char in rawstr:
        if char.isascii() and char not in BADSYMBOLS:
            string.append(char)
    return ''.join(string)


def addscenario(scenario,scenkey,scen,root,container,canvas,frames,mainclass,frame,fromframe):
    ### lots to do here - this is the dup function, need to hack this to make it work for adding a new scenario and utilising the scenario supplied.
    ###
    if len(scen) == 0:
        return
    index = mainclass.allcycles
    cycleframe = CYCLETAG+str(index)
    frames[frame].elemsincycle += 1
    mainclass.allcycles += 1

    ## this adds a new blank scenario in. 
    newscenarioframe = newcycle(root,container,canvas,frames,mainclass,SELECTMODELFRAME,frames[SELECTMODELFRAME].scrollbox[SCENARIOTAG],editmodelfunc,SCENARIOBUTTON,SCENARIOTAG,generalscenariofunc,editscenariopage)

    scenny = scen[0][1]
    frames[newscenarioframe].elems['scenario'+SCENARIO].elem.var.set(cleanstring(scenario))
    frames[newscenarioframe].elems['continent'+SCENARIO].elem.var.set(cleanstring(scenny['continent']))
    frames[newscenarioframe].elems['country'+SCENARIO].elem.var.set(cleanstring(scenny['country']))
    frames[newscenarioframe].elems['mineral'+SCENARIO].elem.var.set(cleanstring(scenny['mineral']))
    frames[newscenarioframe].elems['submineral'+SCENARIO].elem.var.set(cleanstring(scenny['submineral']))
    regionsub = scenny['region']
    if '$' in regionsub:
        reg = regionsub.split('$')[0]
        subreg = regionsub.split('$')[1]
    elif '+' in regionsub:
        reg = regionsub.split('+')[0]
        subreg = regionsub.split('+')[0]
    else:
        reg = regionsub
        subreg = ALL
    frames[newscenarioframe].elems['region'+SCENARIO].elem.var.set(cleanstring(reg))
    frames[newscenarioframe].elems['subregion'+SCENARIO].elem.var.set(cleanstring(subreg))
    totURR = 0
    
    for pair in scen:
        scendict = pair[1]
        totURR += scendict['URR']     
    frames[newscenarioframe].othervars[EXCELURRSTR] = totURR
    frames[newscenarioframe].othervars['urrmethod'] = ''
    acceptscenario(root,frames,mainclass,newscenarioframe,container,skipurr = True)
    for pair in scen:
        scendict = pair[1]
        mineorwell = pair[0]
        cycleframe = newcycle(root,container,canvas,frames,mainclass,newscenarioframe,frames[newscenarioframe].scrollbox[CYCLETAG],minecyclefunc,CYCLEBUTTON,CYCLETAG,scenariocyclefunc,scenariocyclepage)
        frames[cycleframe].cycletypevar.set(mineorwell)        
        for newpair in EXCELCYCLEFIELDS[mineorwell]:
            term = newpair[0] + ENTRYTAG[mineorwell]
            excelterm = newpair[0]
            if scendict[excelterm] is None:
                continue
            frames[cycleframe].elems[term].elem.insert(0,scendict[excelterm])
        for disruption in scendict['Disruptions']:
            disrframe = newcycle(root,container,canvas,frames,mainclass,cycleframe,frames[cycleframe].scrollbox[DISRUPTIONTAG],disruptionfunc,DISRUPTIONBUTTON,DISRUPTIONTAG,disruptioncyclefunc,disruptionpage)
            for i in range(0,len(DISRUPTIONORDERS)):
                excelterm = DISRUPTIONORDERS[i][0]
                term = excelterm + ENTRYTAGDISR
                frames[disrframe].elems[term].elem.insert(0,disruption[i])
            forgetframe(root,frames,disrframe)
        forgetframe(root,frames,cycleframe)
    forgetframe(root,frames,newscenarioframe)
    loadframe(root,frames,mainclass,newscenarioframe,container,additionalfunc = sumcurrentURRfunc)
    forgetframe(root,frames,newscenarioframe)
    return    
        
def findbestoption(sheets,terms):
    for sheet in sheets:
        upsheet = sheet.upper()
        for term in terms:
            if term.upper() in upsheet:
                return sheet
    return None

def getsheet(filename,terms):
    try:
        wb = openpyxl.load_workbook(filename)
    except:
        messagebox.showerror(title = FAILTITLE,message = 'Could not open file')
        return None
    sheets = wb.sheetnames
    bestguesssheet = findbestoption(sheets,terms)
    upsheets = {}
    for sht in sheets:
        upsheets[sht.upper()] = sht
    tab = askstring('Which Tab', 'What tab has the data?',initialvalue = bestguesssheet)
    if tab == '' or tab is None:
        return None    
    uptab = tab.upper()
    while uptab not in upsheets:
        tab = askstring('Tab not found', 'What tab has the data?',initialvalue = bestguesssheet)
        if tab == '' or tab is None:
            return None
        uptab = tab.upper()
    sheet = wb[upsheets[uptab]]
    return sheet



def makekey(cls,scenario=True):
    key = ''
    if scenario:
        key += str(cls.scenario)
    for term in KEYORDER:
        val = str(getattr(cls,term)).upper().replace('_','')
        key = ''.join((key,'+',val))
    return key


### extract production data section...

def cleanword(rawword):
    word = ''.join(filter(str.isascii,rawword))
    for char in BADSYMBOLS:
        word = word.replace(char,'')
    return word

def getattribute(row):
    if len(row) == 0:
        return None
    rawattr = row[0]
    if rawattr is None:
        return None
    rawattr = str(rawattr)
    try:
        attr = cleanword(rawattr).lower().strip()
    except:
        attr = None
    if attr is None or attr == '':
        return None
    return attr

def getyear(row):
    if len(row) == 0:
        return None
    try:
        year = float(row[0])
    except:
        year = None
    return year

def prodrowmismatch(production,attrs,index,warnings,errors,row):
    error = 'Row mismatchs data not extracted properly'
    if error not in errors:
        errors.append(error)
    return production,attrs,index,warnings,errors





def getoutlierdist(values):
    '''
    Method From:
    https://stackoverflow.com/questions/11686720/is-there-a-numpy-builtin-to-reject-outliers-from-a-list
    '''
    if values == []:
        return None
    try:
        med = np.median(values)
    except:
        med = 0.1
    dists = []
    for val in values:
        try:
            newval = abs(val-med)
        except:
            newval = 0
        dists.append(newval)
    norm = np.median(dists)
    if norm == 0:
        return None
    return norm


def getproductionyear(production,attrs,index,warnings,errors,row,year):
    if production == []:
        return production,attrs,index,warnings,errors
    if len(production) != len(row):
        return prodrowmismatch(production,attrs,index,warnings,errors,row)
    i = 0
    for elem in production:
        try:
            value = None
            val = row[i]
            if val is not None and val != '':
                value = float(val)
        except:
            value = None
            warning = 'Failed to read production value'
            if warning not in warnings:
                warnings.append(warning)
        i += 1
        if value is None:
            continue
        if value == 0:
            continue
        if value < 0:
            warning = 'Negative production values found'
            if warning not in warnings:
                warnings.append(warning)
            continue
        precumm = 0
        if len(elem.cumulative) > 0:
            precumm = elem.cumulative[-1]
        cumm = value + precumm
        try:
            prodoncumm = value/float(cumm)
        except:
            prodoncumm = 0
        elem.production.append(value)
        elem.years.append(year)
        elem.cumulative.append(cumm)
        elem.prodovercumm.append(prodoncumm)
    return production,attrs,index,warnings,errors

def getattryear(production,attrs,index,warnings,errors,row,attr):
    if production == []:
        for elem in row:
            production.append(productioncls())
    if len(production) != len(row):
        return prodrowmismatch(production,attrs,index,warnings,errors,row)
    i = 0
    for elem in production:
        try:
            value = str(row[i]).strip()
            if value.upper() == ALL.upper() or value == '':
                value = ALL
        except:
            value = None
            warning = 'Failed to read some attribute values'
            if warning not in warnings:
                warnings.append(warning)
        setattr(elem,attr,value)
        i += 1
    return production,attrs,index,warnings,errors


def addtoproduction(production,attrs,index,warnings,errors,row):
    if len(row) <= 1:
        return production,attrs,index,warnings,errors
    attr = getattribute(row)
    year = getyear(row)
    if len(attrs) == len(ATTRORDERIFBLANKORUNREADABLE) and year is not None:
        ## it's obviously a year run with it
        return getproductionyear(production,attrs,index,warnings,errors,row[1:],year)
    if year is not None:
        warning = 'Insufficient attributes found or production out of order'
        if warning not in warnings:
            warnings.append(warning)
        return getproductionyear(production,attrs,index,warnings,errors,row[1:],year)
    if attr in attrs:
        errors.append('Duplicate Attribute Found: '+attr)
        return production,attrs,index,warnings,errors
    if attr in ATTRORDERIFBLANKORUNREADABLE:
        index += 1
        attrs.append(attr)
        return getattryear(production,attrs,index,warnings,errors,row[1:],attr)
    if attr is not None and index < len(ATTRORDERIFBLANKORUNREADABLE):
        newattr = ATTRORDERIFBLANKORUNREADABLE[index]
        warnings.append('Could not interpret attribute: '+attr+' is not in list '+str(ATTRORDERIFBLANKORUNREADABLE)+', presuming it is: '+newattr)
        attrs.append(attr)
        index += 1
        return getattryear(production,attrs,index,warnings,errors,row[1:],newattr)
    if attr is not None:
        errors.append('Could not interpret attribute: '+attr+' is not in list '+str(ATTRORDERIFBLANKORUNREADABLE))
        index += 1
        return production,attrs,index,warnings,errors
    warning = 'Unsure if interpretting row correctly'
    if warning not in warnings:
        warnings.append(warning)
    attr = ATTRORDERIFBLANKORUNREADABLE[index]
    if attr in attrs:
        error = 'Reading rows incorrectly'
        if error not in errors:
            errors.append(error)
        return production,attrs,index,warnings,errors        
    index += 1
    attrs.append(attr)
    return getattryear(production,attrs,index,warnings,errors,row[1:],attr)

def processExcelProdFile(root,frames,mainclass,frame,labelelem,filename):
    sheet = getsheet(filename,PRODWORDS)
    if sheet is None:
        return
    production = []
    attrs = []
    warnings = []
    errors = []
    index = 0 
    for rawrow in sheet.rows:
        row = []
        for elem in rawrow:
            row.append(elem.value)
        production,attrs,index,warnings,errors = addtoproduction(production,attrs,index,warnings,errors,row)
    writeproduction(root,frames,mainclass,frame,production,warnings,errors,labelelem,filename)
    

def writeproduction(root,frames,mainclass,frame,production,warnings,errors,labelelem,filename):
    if not errorswarningspassed(warnings,errors):
        return
    ### HERERERE
    proddict = {}
    outlierdict = {}
    for prod in production:
        key = makekey(prod,False)
        if key == NONEKEY:
            continue
        if key in proddict:
            messagebox.showerror(title = FAILTITLE,message = 'Duplicate production record: '+key.replace('+',' '))
            return
        prod.outlierdist = getoutlierdist(prod.prodovercumm)
        proddict[key] = prod
        outlierdict[key] = DEFAULTOUTLIERLEVEL
    mainclass.production = proddict
    mainclass.outlierthreshold = outlierdict
    updatelists(mainclass)
    frames[frame].elems[labelelem].elem.config(text = filename)
    mainclass.prodfile = filename


def processCSVProdFile(root,frames,mainclass,frame,labelelem,filename):
    try:
        f = open(filename,'r')
        csvf = csv.reader(f,delimiter = ',',quotechar = '"')
    except:
        messagebox.showerror(title = FAILTITLE,message = 'Could not open file')
        return
    production = []
    attrs = []
    index = 0
    warnings = []
    errors = []
    for row in csvf:
        production,attrs,index,warnings,errors = addtoproduction(production,attrs,index,warnings,errors,row)
    writeproduction(root,frames,mainclass,frame,production,warnings,errors,labelelem,filename)


def getallproduction(proddict):
    histprodd = {}
    for key in proddict:
        cls = proddict[key]
        for i in range(0,len(cls.years)):
            year = cls.years[i]
            prod = cls.production[i]
            if year not in histprodd:
                histprodd[year] = 0
            histprodd[year] += prod
    histyears = sorted(list(histprodd.keys()))
    histprod = []
    histcum = []
    cum = 0 
    for year in histyears:
        prd = histprodd[year]
        histprod.append(prd)
        cum += prd
        histcum.append(cum)
    return histyears,histprod,histcum


    
def loadprodfile(root,frames,mainclass,frame,labelelem):
    frames[frame].elems[labelelem].elem.config(text = BLANKTEXT)
    filename = filedialog.askopenfilename()
    if filename == '' or filename is None:
        return
    meh,ext = os.path.splitext(filename)
    if ext.upper() in ('.XLSX','.XLS','.XLSM'):
        return processExcelProdFile(root,frames,mainclass,frame,labelelem,filename)
    if ext.upper() == '.CSV':
        return processCSVProdFile(root,frames,mainclass,frame,labelelem,filename)
    messagebox.showerror(title=FAILTITLE, message='Invalid Extension: '+ext+'\nFile needs to be excel (xlsx) or csv')
    return

##### load URR file...

def errorswarningspassed(warnings,errors):
    if len(errors) > 0:
        messagebox.showerror(title = ERRORTITLE,message = '\n'.join(errors))
        return False
    if len(warnings) > 0:
        warnings.append(WARNINGCONTINUE)
        warntext = '\n'.join(warnings)
        response = messagebox.askyesno(title=WARNINGTITLE, message=warntext)
        return response
    return True




def getattrtype(elem,index,headermap,warnings,errors):
    lowelem = elem.lower()
    if lowelem in ATTRORDERIFBLANKORUNREADABLE and lowelem not in headermap:
        return lowelem,ATTR,warnings,errors
    if lowelem in ATTRORDERIFBLANKORUNREADABLE:
        errors.append('Duplicate Attribute Found: '+lowelem)
        return None,None,warnings,errors
    if index < len(ATTRORDERIFBLANKORUNREADABLE):
        attr = ATTRORDERIFBLANKORUNREADABLE[index]
        if attr in headermap:
            error = 'Reading rows incorrectly'
            if error not in errors:
                errors.append(error)
            return None,None,warnings,errors
        warning = 'Unsure if interpretting row correctly'
        if warning not in warnings:
            warnings.append(warning)
        return attr,ATTR,warnings,errors
    if REFTAG in lowelem:
        index = lowelem.index(REFTAG)
        attr = elem[0:index] + elem[index + len(REFTAG):]
        if not attr in headermap:
            warning = 'Scenario Reference before/missing scenario: '+attr
            warnings.append(warning)
        return attr,REF,warnings,errors
    return elem,SCEN,warnings,errors


def getURRheader(row,warnings,errors):
    headermap = {ATTR:{}}
    cleanrow = []
    for elem in row:
        try:
            cleanelem = cleanword(elem)
        except:
            cleanelem = None
        cleanrow.append(cleanelem)
    i = -1
    for elem in cleanrow:
        ### function, return var, type (type is attr, scenario or reference)
        i += 1
        var,vartype,warnings,errors = getattrtype(elem,i,headermap,warnings,errors)
        if vartype == ATTR:
            headermap[ATTR][var] = i
        if vartype in (SCEN,REF):
            if var not in headermap:
                headermap[var] = {}
            headermap[var][vartype] = i
    return headermap,warnings,errors

def strfunc(default = ''):
    def strfunc_inner(string,warnings,errors):
        if string is None:
            return default,warnings,errors
        try:
            val = str(string)
        except:
            val = default
        if val.upper() == default.upper():
            return default,warnings,errors
        return val,warnings,errors
    return strfunc_inner

def floatfunc(flt,warnings,errors):
    if flt is None:
        return 0.0,warnings,errors
    try:
        val = float(flt)
    except:
        val = 0
        warning = 'Unfloatable value(s)'
        if warning not in warnings:
            warnings.append(warning)
    if val < 0:
        warning = 'Negative value(s)'
        if warning not in warnings:
            warnings.append(warning)
        val = 0
    return val,warnings,errors

def geturrval(index,lenny,row,warnings,errors,func):
    if index >= lenny:
        warning = 'Possible URR data missing'
        if warning not in warnings:
            warnings.append(warning)
        return func(None,warnings,errors)
    return func(row[index],warnings,errors)


def addtoURRs(URRs,headermap,warnings,errors,row):
    if errors != []:
        return URRs,warnings,errors
    if row == []:
        return URRs,warnings,errors    
    basecls = URRcls()
    lenny = len(row)
    for key in headermap[ATTR]:
        index = headermap[ATTR][key]
        val,warnings,errors = geturrval(index,lenny,row,warnings,errors,strfunc(ALL))
        setattr(basecls,key,val)
    for scen in headermap:
        if scen == ATTR:
            continue
        urrcl = copy.deepcopy(basecls)
        urrcl.scenario = scen
        if SCEN in headermap[scen]:
            val,warnings,errors = geturrval(headermap[scen][SCEN],lenny,row,warnings,errors,floatfunc)
            urrcl.URR = val
        if REF in headermap[scen]:
            val,warnings,errors = geturrval(headermap[scen][REF],lenny,row,warnings,errors,strfunc(''))
            urrcl.ref = val
        URRs.append(urrcl)
    return URRs,warnings,errors
    
def writeURRs(root,frames,mainclass,frame,URRs,warnings,errors,labelelem,filename):
    if not errorswarningspassed(warnings,errors):
        return
    ### HERERE AND PRODUCTION --> need to create as a dictionary...
    urrdict = {}
    for urr in URRs:
        key = makekey(urr,True)
        if key in urrdict:
            messagebox.showerror(title = FAILTITLE,message = 'Duplicate URR record: '+key.replace('+',' '))
            return
        urrdict[key] = urr
    mainclass.URR = urrdict
    updatelists(mainclass)
    frames[frame].elems[labelelem].elem.config(text = filename)
    mainclass.urrfile = filename

    
def processExcelURRFile(root,frames,mainclass,frame,labelelem,filename):
    sheet = getsheet(filename,URRWORDS)
    if sheet is None:
        return
    URRs = []
    warnings = []
    errors = []
    firstrow = True
    for rawrow in sheet.rows:
        row = []
        for elem in rawrow:
            row.append(elem.value)
        if list(set(row)) == [None]:
            continue
        if firstrow:
            headermap,warnings,errors = getURRheader(row,warnings,errors)
            firstrow = False
            continue
        URRs,warnings,errors = addtoURRs(URRs,headermap,warnings,errors,row)
    writeURRs(root,frames,mainclass,frame,URRs,warnings,errors,labelelem,filename)

    
def processCSVURRFile(root,frames,frame,mainclass,labelelem,filename):
    try:
        f = open(filename,'r')
        csvf = csv.reader(f,delimiter = ',',quotechar = '"')
    except:
        messagebox.showerror(title = FAILTITLE,message = 'Could not open file')
        return
    URRs = []
    warnings = []
    errors = []
    firstrow = True
    for row in csvf:
        if firstrow:
            headermap,warnings,errors = getURRheader(row,warnings,errors)
            firstrow = False
            continue
        URRs,warnings,errors = addtoURRs(URRs,headermap,warnings,errors,row)
    writeURRs(root,frames,mainclass,frame,URRs,warnings,errors,labelelem,filename)
    
def loadurrfile(root,frames,mainclass,frame,labelelem):
    frames[frame].elems[labelelem].elem.config(text = BLANKTEXT)
    filename = filedialog.askopenfilename()
    if filename == '' or filename is None:
        return
    meh,ext = os.path.splitext(filename)
    if ext.upper() in ('.XLSX','.XLS','.XLSM'):
        return processExcelURRFile(root,frames,mainclass,frame,labelelem,filename)
    if ext.upper() == '.CSV':
        return processCSVURRFile(root,frames,frame,mainclass,labelelem,filename)
    messagebox.showerror(title=FAILTITLE, message='Invalid Extension: '+ext+'\nFile needs to be excel (xlsx) or csv')
    return
    


def updatelists(mainclass):
    for var in SCENVARS:
        setattr(mainclass,''.join(('default',var,'s')),[])
    for term in ('URR','production'):
        dictterms = getattr(mainclass,term)
        for key in dictterms:
            cls = dictterms[key]
            for var in SCENVARS:
                if var == 'scenario' and term == 'production':
                    continue
                alist = getattr(mainclass,''.join(('default',var,'s')))
                alist.append(getattr(cls,var))
    for var in SCENVARS:
        redlist = sorted(list(set(getattr(mainclass,''.join(('default',var,'s'))))))
        setattr(mainclass,''.join(('default',var,'s')),redlist)
    


def loadbibtexfile(root,frames,mainclass,frame,labelelem):
    warnings = []
    errors = []
    filename = filedialog.askopenfilename()
    if filename == '' or filename is None:
        return
    meh,ext = os.path.splitext(filename)
    if ext.upper() != '.BIB':
        error = 'Unknown file extension: '+ext[1:]+' Expecting bib'
        errors.append(error)
    if not errorswarningspassed(warnings,errors):
        return
    frames[frame].elems[labelelem].elem.config(text = filename)
    mainclass.bibtextfile = filename
    
def hubbertmodelpage(root,canvas,frames):
    pass

def newminecycle(root,canvas,frames):
    pass


def loadimage(root,frames,mainclass,frame,container):
    for key in frames[frame].imageoptions:
        terms = frames[frame].imageoptions[key]
        if terms is None:
            continue
        can,wid,ptl,pfm = terms
        wid.grid_forget()
        pfm.grid_forget()

    imgkey = frames[frame].imgvar.get()
    if frames[frame].dropdownvar is not None:
        imgkey = frames[frame].dropdownvar.get()+'|'+imgkey
    if imgkey not in frames[frame].imageoptions:
        return
    imagedata = frames[frame].imageoptions[imgkey]
    if imagedata is None:
        #### HERERE --> make the image a blank image.
        return
    pltcanvas,pltwidget,plttool,pltframe = imagedata
    pltwidget.grid(row =OUTPUTIMAGELOC[0],column = OUTPUTIMAGELOC[1],columnspan = OUTPUTIMAGECOLSPAN,rowspan = OUTPUTIMAGEROWSPAN)
    pltcanvas.draw()
    pltframe.grid(row =11,column = 5,columnspan = 3)
    plttool.update()



def forgetframe(root,frames,frame):
    for key in frames[frame].elems:
        frames[frame].elems[key].elem.grid_forget()
    if frames[frame].imageobj is not None:
        frames[frame].imageobj.elem.grid_forget()
    for key in frames[frame].cycleframeelems:
        forgetcycleframes(frames[frame].cycleframeelems[key])
    frames[frame].frame.grid_forget()

def loadframe(root,frames,mainclass,frame,container,additionalfunc = None):
    exclude = frames[frame].hideelems
    #frames[frame].frame.grid(in_=root,columnspan = GRIDSIZE[1],rowspan = GRIDSIZE[0],sticky = 'nswe')
    frames[frame].frame.grid(in_=container,sticky = 'nswe')

##    for row in range(0,GRIDSIZE[0]):
##        newframe.grid_rowconfigure(row, weight = 1, minsize=MINROWSIZE)
##    for col in range(0,GRIDSIZE[1]):
##        newframe.grid_columnconfigure(col, weight = 1,minsize = MINCOLSIZE)

    for key in frames[frame].elems:
        if key not in exclude:
            frames[frame].elems[key].grid()
    for key in frames[frame].cycleframeelems:
        if key not in exclude:
            loadcycleframes(frames[frame].cycleframeelems[key])
    if frames[frame].imageobj is not None and 'imageobj' not in exclude:
        frames[frame].imageobj.grid()
    if frames[frame].initialise: ### HERERERE STEVE USE THIS TO PARTIALLY SHOW FRAME.
        if frames[frame].initfunc is not None:
            frames[frame].initfunc(root,frames,mainclass,frame,container)
    if additionalfunc is not None:
        additionalfunc(root,frames,mainclass,frame,container)
        #loadimage(root,frames,frame,container)
### Steve HERERERERERERERE PLUDONKIE

def forgetcycleframes(cycledict):
    for key in cycledict:
        cycledict[key].grid_forget()
        
def loadcycleframes(cycledict):
    for key in cycledict:
        if not cycledict[key].hide:
            cycledict[key].grid()


def navigateto(root,frames,mainclass,fromframe,toframe,container,tag='',additionalfunc=None):
    forgetframe(root,frames,fromframe)
    loadframe(root,frames,mainclass,toframe,container,additionalfunc=additionalfunc)

def duplicatenavigateto(root,frames,mainclass,fromframe,toframe,container,tag=''):
    ## herere --> frame:class    okoko
    forgetframe(root,frames,fromframe)
    loadframe(root,frames,mainclass,toframe,container)
    

def acceptbutton(root,frames,mainclass,fromframe,toframe,container,func = navigateto,tag = ''):
    ## okokoko
    #frames[frame].cycleframeelems[tag][index] = scenario
    acceptbuttonelem = elemcls(frame = frames[fromframe].frame,
                               row = ACCEPTBUTTONLOC[0],
                               column = ACCEPTBUTTONLOC[1],
                               padx = PADX,
                               pady = PADY,
                               text = ACCEPTBUTTONTEXT,
                               obj = 'button',
                               command = lambda : func(root,frames,mainclass,fromframe,toframe,container,tag),
                               font = ACCEPTFONT)
    acceptbuttonelem.grid()
    frames[fromframe].elems[ACCEPT] = acceptbuttonelem
    return root,frames



def backbutton(root,frames,mainclass,fromframe,toframe,container,func = navigateto,tag = '',backbuttontext = BACKBUTTONTEXT,additionalfunc = None):
    backbutton = elemcls(frame = frames[fromframe].frame,
                         obj = 'button',
                         row = BACKBUTTONLOC[0],
                         column = BACKBUTTONLOC[1],
                         padx = PADX,
                         pady = PADY,
                         text = backbuttontext,
                         command = lambda : func(root,frames,mainclass,fromframe,toframe,container,tag,additionalfunc=additionalfunc),
                         font = BUTTONFONT)
    backbutton.grid()
    frames[fromframe].elems[BACK] = backbutton
    return root,frames



def runfullmodelbutton(root,container,frames,mainclass,frame,toframe):
    runmodbut = elemcls(frame = frames[frame].frame,
                      obj = 'button',
                      row = RUNSCENLOC[0],
                      column = RUNSCENLOC[1],
                      padx = PADX,
                      pady = PADY,
                      text = RUNFULLTEXT,
                      command = lambda : runfullscenario(root,container,frames,mainclass,frame,toframe),
                      font = BUTTONFONT)
    runmodbut.grid()
    frames[frame].elems[RUNSCEN] = runmodbut
    return root,frames




def nextbutton(root,frames,mainclass,frame,toframe,container):    
    ## Next button
    nextbut = elemcls(frame = frames[frame].frame,
                      obj = 'button',
                      row = NEXTBUTTONLOC[0],
                      column = NEXTBUTTONLOC[1],
                      padx = PADX,
                      pady = PADY,
                      text = NEXTBUTTONTEXT,
                      command = lambda : navigateto(root,frames,mainclass,frame,toframe,container),
                      font = BUTTONFONT)
    nextbut.grid()
    frames[frame].elems[NEXT] = nextbut
    return root,frames






def cyclefunc(root,frames,mainclass,frame,buttonfunc,toframe,container):
    buttonfunc(root,frames,mainclass,frame,toframe,container)

def implementsaveback(root,frames,mainclass,fromframe,toframe,container,tag,additionalfunc=None):
    #frames[frame].elems[term+SCENARIO]        
    #frames[fromframe].cycleframeelems[tag] # what index in list... possibly make it a dictionary? 
    # how do I get that info to here. 
    navigateto(root,frames,mainclass,fromframe,toframe,container,additionalfunc=additionalfunc)
    ### ok when hitting the back button I need to find the scenario and write the data.
    ### cycle should be 
    ### only solution is to create a new frame with each button, then don't need to delete and hide.








def disruptionfunc(root,frames,mainclass,frame,toframe,container):
    navigateto(root,frames,mainclass,frame,toframe,container)


def minecyclefunc(root,frames,mainclass,frame,toframe,container):
    skipterms = []
    for key in INPUTTERMS:
        if frames[toframe].cycletypevar.get() != key:
            for term in INPUTTERMS[key]:
                skipterms.append(term + ENTRYTAG[key])
                skipterms.append(term + TITLETAG[key])
    frames[toframe].hideelems = skipterms
    #igateto(root,frames,mainclass,frame,scenario.loadframe,skipelems = skipterms)
    navigateto(root,frames,mainclass,frame,toframe,container)
        



def editmodelfunc(root,frames,mainclass,frame,toframe,container):  
    navigateto(root,frames,mainclass,frame,toframe,container)


def deletefunc(root,frames,mainclass,frame,tag,scenario):
    print('got here')
    #.scenarios
    print(mainclass.scenarios.keys())
    deleterow = scenario.row
    forgetcycleframes(frames[frame].cycleframeelems[tag])
    diction = frames[frame].cycleframeelems[tag]
    for scen in diction:
        if diction[scen].row < deleterow:
            continue
        if diction[scen].row == deleterow:
            #diction[scen].row = frames[frame].elemsincycle+1
            diction[scen].row = -1
            frames[diction[scen].frame].hide = True
            diction[scen].hide = True
        else:
            diction[scen].row = diction[scen].row -1
    loadcycleframes(diction)

def dupnamefunc(oldtext):
    return 'copy of ' + oldtext

def clone(widget):
    #https://stackoverflow.com/questions/46505982/is-there-a-way-to-clone-a-tkinter-widget
    parent = widget.nametowidget(widget.winfo_parent())
    cls = widget.__class__

    clone = cls(parent)
    for key in widget.configure():
        clone.configure({key: widget.cget(key)})
    return clone
    
def duplicatefunc(root,container,canvas,frames,mainclass,frame,scrollbox,buttonfunc,cyclename,origtag,scenariofunc,pagefunc,scenario,fromframe):
    '''
    These elements need to be found in the frame, for the given cycleelem:
    scrollbox,buttonfunc,cyclename,tag,scenariofunc,pagefunc,scenario

    then below we need to loop over the cycleelems and recall duplicatefunc (so that we can duplicate these components)

    Last. we need to call a frame specific function to implement stuff (notably the URR selection)
    
    '''
    index = mainclass.allcycles
    cycleframe = origtag+str(index)
    newscenario = cyclecls(row = frames[frame].elemsincycle,index = index)
    newscenario.frame = cycleframe
    newscenario.loadframe = cycleframe
    newscenario = makecyclebuttons(root,container,canvas,frames,mainclass,fromframe,scrollbox,buttonfunc,cyclename,origtag,scenariofunc,pagefunc,newscenario,newscenario.loadframe)
    #frames[cycleframe] = createemptyframe(root,container,FRAMECOLOR,frames[scenario.frame].fromframe,frombutton = newscenario.button)
    frames[cycleframe] = createemptyframe(root,container,FRAMECOLOR,fromframe,frombutton = newscenario.button)
    frames[cycleframe].scrollbox[origtag] = scrollbox
    frames[cycleframe].buttonfunc[origtag] = buttonfunc
    frames[cycleframe].cyclename[origtag] = cyclename
    frames[cycleframe].scenariofunc[origtag] = scenariofunc
    frames[cycleframe].pagefunc[origtag] = pagefunc
    
    
    newscenario.button.text = dupnamefunc(newscenario.button.text)
    pagefunc(root,container,canvas,frames,mainclass,cycleframe)

    
    frames[frame].elemsincycle += 1
    mainclass.allcycles += 1
    newelems = frames[newscenario.frame].elems
    oldelems = frames[scenario.frame].elems
    for ekey in oldelems:
        if oldelems[ekey].obj not in ('entry','autocomplete'):
            continue
        value = oldelems[ekey].elem.get()
        newelems[ekey].elem.delete(0,tk.END)
        newelems[ekey].elem.insert(0,value)
        if oldelems[ekey].obj  != 'autocomplete':
            continue
        newelems[ekey].elem.var.set(value)
        if 'lb' in dir(newelems[ekey].elem):
            newelems[ekey].elem.lb.destroy()
    frames[newscenario.frame].othervars = copy.deepcopy(frames[scenario.frame].othervars)
    
    PAGESPECIFICDUPLICATION[origtag](frames,scenario.frame,newscenario.frame)
    frames[fromframe].cycleframeelems[origtag][index] = newscenario
    frames[newscenario.loadframe].cls = newscenario
    for cycle in frames[scenario.frame].cycleframeelems:
        for term in frames[scenario.frame].cycleframeelems[cycle]:
            cframe = frames[scenario.frame].cycleframeelems[cycle][term].frame
            for tag in frames[cframe].pagefunc:
                duplicatefunc(root,container,canvas,frames,mainclass,cycleframe,frames[newscenario.frame].scrollbox[tag],
                              frames[cframe].buttonfunc[tag],frames[cframe].cyclename[tag],cycle,
                              frames[cframe].scenariofunc[tag],frames[cframe].pagefunc[tag],frames[cframe].cls,cycleframe)

    ### need to 'load' the new scenario.
    #newscenario.button.row = 10
    newscenario.grid()
    ## need to do more stuff - probably easier to do specific things for specific classes --> if on classes call function? 
    ## okokok
    
def makecyclebuttons(root,container,canvas,frames,mainclass,frame,scrollbox,buttonfunc,cyclename,tag,scenariofunc,pagefunc,scenario,toframe):
    scenario.button = elemcls(frame = scrollbox.elem,
                              row = frames[frame].elemsincycle,
                              column = 0,
                              anchor = 'w',
                              justify = 'left',
                              text = str(frames[frame].elemsincycle+1)+cyclename,
                              font = SMALLFONT,
                              obj = 'button',
                              command = lambda : cyclefunc(root,frames,mainclass,frame,buttonfunc,toframe,container))

    scenario.delete = elemcls(frame = scrollbox.elem,
                              row = frames[frame].elemsincycle,
                              column = 1,
                              text = DELETEBUTTON,
                              font = SMALLFONT,obj = 'button',
                              command = lambda : deletefunc(root,frames,mainclass,frame,tag,scenario))
    ##### STEVE
    scenario.duplicate = elemcls(frame = scrollbox.elem,
                                 row = frames[frame].elemsincycle,
                                 column = 2,
                                 text = DUPLICATEBUTTON,
                                 font = SMALLFONT,
                                 obj = 'button',
                                 command = lambda : duplicatefunc(root,container,canvas,frames,mainclass,frame,scrollbox,buttonfunc,cyclename,tag,scenariofunc,pagefunc,scenario,frame))
    return scenario


def newcycle(root,container,canvas,frames,mainclass,frame,scrollbox,buttonfunc,cyclename,tag,scenariofunc,pagefunc):
    index = mainclass.allcycles
    cycleframe = tag+str(index)
    scenario = cyclecls(row = frames[frame].elemsincycle,index = index)
    scenario.loadframe = cycleframe
    scenario = makecyclebuttons(root,container,canvas,frames,mainclass,frame,scrollbox,buttonfunc,cyclename,tag,scenariofunc,pagefunc,scenario,scenario.loadframe)    
    scenario.grid()
    scenario = scenariofunc(scenario)
    
    frames[frame].cycleframeelems[tag][index] = scenario
    frames[cycleframe] = createemptyframe(root,container,FRAMECOLOR,frame,frombutton = scenario.button)
    pagefunc(root,container,canvas,frames,mainclass,cycleframe)## here okoko
    frames[scenario.loadframe].cls = scenario
    scenario.frame = cycleframe
    frames[frame].elemsincycle += 1
    mainclass.allcycles += 1
    frames[cycleframe].scrollbox[tag] = scrollbox
    frames[cycleframe].buttonfunc[tag] = buttonfunc
    frames[cycleframe].cyclename[tag] = cyclename
    frames[cycleframe].scenariofunc[tag] = scenariofunc
    frames[cycleframe].pagefunc[tag] = pagefunc
    frames[cycleframe].othervars[EXCELURRSTR] = None
    return cycleframe
    
def makecycles(root,container,canvas,frames,mainclass,frame,func,cycleloc,cyclespan,newcycleloc,newcycletext,cyclename,tag,scenariofunc,scrollsize,pagefunc,whatcalledthis):
    # okoko
    
#scrollbox,buttonfunc,cyclename,tag,scenariofunc,pagefunc
    # frame to hold the scrollbox
    if tag not in frames[frame].cycleframeelems:
        frames[frame].cycleframeelems[tag] = {}
    
    cycleframe = elemcls(frame = frames[frame].frame,
                         row = cycleloc[0],
                         column = cycleloc[1],
                         rowspan = cyclespan[0],
                         columnspan = cyclespan[1],
                         sticky = 'nsew',
                         obj = 'frame')
    #for row in range(0,cyclespan[0]):
    #    cycleframe.elem.grid_rowconfigure(row, weight=1)
    #for col in range(0,cyclespan[1]):
    #    cycleframe.elem.grid_columnconfigure(col,weight=1)
    #cycleframe.elem.grid_propagate(False)
    cycleframe.grid()
    frames[frame].elems[tag+CYCLEFRAME] = cycleframe

    sf = elemcls(frame = cycleframe.elem,
                 sticky = 'nswe',
                 padx = PADX,
                 pady = PADY,
                 obj = 'scrollframe',
                 height = scrollsize[0],
                 width = scrollsize[1])
    sf.elem.bind_arrow_keys(cycleframe.elem)
    sf.elem.bind_scroll_wheel(cycleframe.elem)
    sf.elem._canvas.configure(bg='white')
    sf.grid()
    
    frames[frame].elems[tag+SCROLLFRAME] = sf
    frame_buttons = elemcls(obj = 'displaywidget',
                            frame = tk.Frame,
                            widget = sf.elem.display_widget)    
    frames[frame].scrollbox[tag] = frame_buttons
    newcyclebutton = elemcls(row = newcycleloc[0],
                             column = newcycleloc[1],
                             padx = PADX,
                             obj = 'button',
                             frame = frames[frame].frame,
                             pady = PADY,
                             text = newcycletext,
                             command = lambda : newcycle(root,container,canvas,frames,mainclass,frame,frame_buttons,func,cyclename,tag,scenariofunc,pagefunc))
    newcyclebutton.grid()
    frames[frame].elems[tag+CYCLEBUTTON] = newcyclebutton
    return root,frames

def scenariocyclefunc(scenario):
    return scenario

def disruptioncyclefunc(scenario):
    return scenario

### CONFUSED why nothing happens. 
#makecycles(root,container,canvas,frames,SELECTMODELFRAME,editmodelfunc,SCENARIOSLISTLOC,SCENARIOSLISTSPAN,NEWSCENLOC,NEWSCENTEXT,SCENARIOBUTTON,SCENARIOTAG,generalscenariofunc,SCENSCROLLSIZE,scencls,editscenariopage)
#makecycles(root,container,canvas,frames,frame,minecyclefunc,CYCLELOC,CYCLELOCSPAN,NEWCYCLELOC,NEWCYCLETEXT,CYCLEBUTTON,MINESTAG,scenariocyclefunc,CYCLESIZE,cyclecls,scenariocyclepage)

def loadsceninputs(root,frames,frame,key):
    skipterms = []
    for k in INPUTTERMS:
        for term in INPUTTERMS[k]:
            if k == key:
                frames[frame].elems[term + ENTRYTAG[k]].grid()
                frames[frame].elems[term + TITLETAG[k]].grid()
            else:
                frames[frame].elems[term + ENTRYTAG[k]].elem.grid_forget()
                frames[frame].elems[term + TITLETAG[k]].elem.grid_forget()
                skipterms.append(term + ENTRYTAG[k])
                skipterms.append(term + TITLETAG[k])
    frames[frame].hideelems = skipterms

def disruptionpage(root,container,canvas,frames,mainclass,frame):
    root,frames = maketitle(root,frames,frame,DISRUPTIONTITLE)
    ## Save buttons
    root,frames = savebuttons(root,frames,frame,mainclass)
    ### okokokok --> I think here is the issue, need to send the class into the create scenario, and link the widet to the class. 
    root,frames = createscenarios(root,frames,mainclass,frame,DISRTERMS,ENTRYTAGDISR,TITLETAGDISR)
    ## back button
    root,frames = acceptbutton(root,frames,mainclass,frame,frames[frame].fromframe,container,func=duplicatenavigateto)
    forgetframe(root,frames,frame)


def sumcurrentURRfunc(root,frames,mainclass,frame,container):
    toturr = 0
    for key in frames[frame].cycleframeelems:
        for num in frames[frame].cycleframeelems[key]:
            if frames[frame].cycleframeelems[key][num].hide:
                continue
            cycleframe = key+str(num)
            mineorwell = frames[cycleframe].cycletypevar.get()
            urr = floaty(frames[cycleframe].elems['URR'+mineorwell+ENTRY].elem.get())
            if urr is None:
                continue
            toturr += urr
    urrstring,digits = makeurr(toturr)
    frames[frame].elems['CurrURR'].elem['text'] = urrstring
    
def scenariocyclepage(root,container,canvas,frames,mainclass,frame):
    root,frames = maketitle(root,frames,frame,SCENARIOCYCLETITLE)
    cycletypevar = tk.StringVar()
    mineradio = elemcls(row = MINESCENARIOBUTTONLOC[0],
                        column = MINESCENARIOBUTTONLOC[1],
                        sticky = 'nswe',
                        obj = 'radiobutton',
                        frame = frames[frame].frame,
                        indicatoron = 0,
                        text = MINESCENARIOTEXT,
                        command = lambda : loadsceninputs(root,frames,frame,MINE),
                        variable=cycletypevar,
                        value=MINE,
                        font = RADIOBUTTONFONT)
    mineradio.grid()
    frames[frame].elems[MINE] = mineradio
    wellradio = elemcls(row = WELLSCENARIOBUTTONLOC[0],
                        column = WELLSCENARIOBUTTONLOC[1],
                        sticky = 'nswe',
                        frame = frames[frame].frame,
                        obj = 'radiobutton',
                        indicatoron = 0,
                        text = WELLSCENARIOTEXT,
                        command = lambda : loadsceninputs(root,frames,frame,WELL),
                        variable=cycletypevar,
                        value=WELL,
                        font = RADIOBUTTONFONT)
    wellradio.grid()
    frames[frame].elems[WELL] = wellradio
    cycletypevar.set(MINE)
    frames[frame].cycletypevar = cycletypevar
    root,frames = makecycles(root,container,canvas,frames,mainclass,frame,disruptionfunc,DISRUPTIONLOC,DISRUPTIONSPAN,NEWDISRUPTIONLOC,DISRUPTIONTEXT,DISRUPTIONBUTTON,DISRUPTIONTAG,disruptioncyclefunc,DISRUPTIONSIZE,disruptionpage,'disruption disruption disruption')

    root,frames = createscenarios(root,frames,mainclass,frame,INPUTTERMS[MINE],ENTRYTAG[MINE],TITLETAG[MINE])
    root,frames = createscenarios(root,frames,mainclass,frame,INPUTTERMS[WELL],ENTRYTAG[WELL],TITLETAG[WELL])
    
    ## Save buttons
    root,frames = savebuttons(root,frames,frame,mainclass)
    ## back button
    root,frames = backbutton(root,frames,mainclass,frame,frames[frame].fromframe,container,backbuttontext = ACCEPT,additionalfunc = sumcurrentURRfunc)

    forgetframe(root,frames,frame)
    
    ## DISRUPTIONS FRAME (CYCLES)
    
    ## DONE: RADIO BUTTON FOR MINE OR WELL INPUT
    ## THEN DISPLAY EITHER THE MINE OR WELL INPUT DATA
    ## THEN HAVE A LOAD BUTTON (or INPUT BUTTON) --> THIS BUTTON NEEDS TO TRIGGER RERUNNING THE MODEL. --> FOR THAT SCENARIO.


def editscenario(root,frames,mainclass,frame,container):
    forgetframe(root,frames,frame)
    exclude = []
    for key in frames[frame].elems:
        if key in EDITSCENSHOW:
            continue
        exclude.append(key)
    frames[frame].initialise = False
    frames[frame].hideelems = exclude
    loadframe(root,frames,mainclass,frame,container)
    frames[frame].initialise = True
    return root,frames

def createcls(root,frames,mainclass,frame):
    cls = scenariocls()
    for term in SCENARIOTERMS:
        val = frames[frame].elems[term+SCENARIO].elem.get()
        if not val.isascii():
            messagebox.showerror(title = FAILTITLE,message = 'Please stick to ASCII characters')
            return None
        for char in BADSYMBOLS:
            if char in val:
                messagebox.showerror(title = FAILTITLE,message = 'The character '+char+' is forbiden')
                return None
        if val.upper() == ALL.upper() or val == '':
            val = ALL
        setattr(cls,term,val)
    return cls

def relabelterms(frames,frame,cls):
    for term in SCENARIOTERMS:
        attr = term+LABEL
        frames[frame].elems[attr].elem.config(text = getattr(cls,term))

def getymax(production,threshold,ymin):
    production.outlierdist    
    if production.outlierdist is None:
        return 1
    try:
        maxy = production.outlierdist*threshold
    except:
        maxy = 1
    try:
        maxy = max(ymin * BOOSTABOVEMIN,maxy)
    except:
        pass
    return maxy

def getnumdeci(value):
    '''
    stupid function to convert:
    0.043 to 4,-2 i.e. 4x10^-2
    '''
    if value >= 1:
        return 1,0
    d = 0
    curr = value
    while curr < 1:
        curr = curr*10
        d = d - 1
    curr = int(curr)
    return curr,d
        
def nicestr(value):
    val = str(round(value, 1-int(math.floor(math.log10(abs(value))))))
    return val    

    
def makeyticks(ymin,ymax):
    #(locs,vals)
    locs = [0]
    vals = ['0']
    num,deci = getnumdeci(ymax)
    value = num*pow(10,deci)
    if num > 2:
        splits = num
    elif num == 2:
        splits = 4
    else:
        splits = 10
    for i in range(0,splits):
        loc = (i+1)*(value/splits)
        locs.append(loc)
        vals.append(nicestr(loc))
    if ymin == 0:
        return locs,vals
    num,mdeci = getnumdeci(ymin)
    if deci <= mdeci:
        return locs,vals
    for d in range(mdeci,deci):
        for i in range(1,10):
            loc = i*pow(10,d)
            if loc < ymax:
                locs.append(loc)
                vals.append('')
    return locs,vals
    
    
def getymin(values):
    if values == []:
        return 0
    ymin = 1
    for val in values:
        if val == 0:
            continue
        if val < ymin:
            ymin = val
    return ymin

def getscifizero(value):
    try:
        n = int(math.log10(value))
    except:
        return None,None
    d = 0
    charval = str(int(value))[::-1]
    for elem in charval:
        if elem != '0':
            return n,d
        d += 1
    return None,None

    
def makexticks(cum):
    locvals = plot.createvalticks(0,cum)
    locs = locvals[0]
    vals = locvals[1]
    if len(vals) <= 2:
        return locvals,''
    low = locs[1]
    high = locs[-1]
    if low < 10:
        return locvals,''
    nlow,zlow = getscifizero(low)
    nhigh,zhigh = getscifizero(high)
    divisor = nhigh
    digits = nhigh - zlow
    label = r'x10$^{'+str(divisor)+'}$'
    newvals = []
    for loc in locs:
        newval = f'{(loc/(pow(10,divisor))):.2f}'
        newvals.append(newval)
    return (locs,newvals),label

def getvlines(years,x,y):
    vlines = []
    xticks = [[],[]]
    runningaverage = RUNNINGSKIPS*2+1
    if len(x) <= runningaverage*2:
        return vlines,xticks
    try:
        rave = np.convolve(y,np.ones(runningaverage)/runningaverage,'valid')
    except:
        return vlines,xticks
    slopes = getslopes(x[RUNNINGSKIPS:-RUNNINGSKIPS],rave)
    xs = x[RUNNINGSKIPS*2:-RUNNINGSKIPS*2]
    yrs = years[RUNNINGSKIPS*2:-RUNNINGSKIPS*2]
    slopes = slopes[::-1]
    xs = xs[::-1]
    yrs = yrs[::-1]
    oldslope = None
    #MINCUM
    totcum = x[-1]
    for i in range(0,len(xs)):
        xval = xs[i]
        if xval/totcum < MINCUM:
            return vlines,xticks
        slope = slopes[i]
        if oldslope is None or slope is None:
            val = 0
        else:
            try:
                val = slope*oldslope
            except:
                val = 0
        if val < 0:
            yr = str(int(yrs[i]))
            p = plot.vline(xval,color = BLACK,style = '--',secondxaxis = True)
            vlines.append(p)
            xticks[0].append(xval)
            xticks[1].append(yr)
            if len(vlines) >= MAXVLINES:
                return vlines,xticks
        oldslope = slope        
    return vlines,xticks
    
    
def getslopes(x,aves):
    slopes = []
    for i in range(0,len(aves)-RUNNINGSKIPS*2):
        try:
            slope = linregress(x[i:RUNNINGSKIPS*2+1+i],aves[i:RUNNINGSKIPS*2+1+i]).slope
        except:
            slope = None
        slopes.append(slope)
    return slopes


def getxmax(xticks,lastcumval):
    try:
        val = max(xticks[0])
    except:
        val = lastcumval
    return val
    

def HLplot(production,outlierthreshold):
    if len(production.cumulative) == 0:
        return None
    ymin = getymin(production.prodovercumm)
    ymax = getymax(production,outlierthreshold,ymin)
    yticks = makeyticks(ymin,ymax)
    xticks,xlabel = makexticks(production.cumulative[-1])
    
    xmax = getxmax(xticks,production.cumulative[-1])
    vlines,sxticks = getvlines(production.years,production.cumulative,production.prodovercumm)
    
    phl = plot.line(production.cumulative,production.prodovercumm,color = BLACK,style = '-')
    plt = plot.display([phl]+vlines,None,save = False,yminmax = (0,ymax),xminmax = (0,xmax),second_xminmax = (0,xmax),yticks = yticks,xticks = xticks,xlabel = xlabel,second_xticks = sxticks,second_xrotation = 90,figsize = HLPLOTSIZE)
        
    annot = plt.ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",color = 'red',
                    bbox=dict(boxstyle="round", fc="w"),
                    arrowprops=dict(arrowstyle="->"))

    annot.set_visible(False)
    plt.annot = annot
    plt.bubbles = getbubbles(production.cumulative,production.prodovercumm,production.years,ymax)
    
    return plt

class bubble():
    def __init__(self,startx,endx,starty,endy,x,y,label):
        self.startx = startx
        self.endx = endx
        self.starty = starty
        self.endy = endy
        self.x = x
        self.y = y
        self.label = label

def getbubbles(xs,ys,years,ymax,ymin = 0):
    bubbles = []
    lastx = 0
    for i in range(0,len(xs)):
        x = xs[i]
        y = ys[i]
        yr = years[i]
        startx = (x + lastx)/2.0
        window = (max((ymin,y*(1-(YHEIGHT/100)))),min((ymax,y*(1+(YHEIGHT/100)))))
        starty = min(window)
        endy = max(window) ## if y is negative, start and end switch around.
        if i+1 < len(xs):
            nextx = xs[i+1]
        else:
            nextx = x + (x - lastx)
        endx = (nextx + x)/2.0
        label = str(int(yr))
        bub = bubble(startx,endx,starty,endy,x,y,label)
        bubbles.append(bub)
        lastx = x
    return bubbles


def getindi(vals,val):
    try:
        ind = vals.index(val)
    except:
        ind = None
    return ind 

def splitdata(xs,ys,vals,start,stop):
    relx = []
    rely = []
    starti = getindi(vals,start)
    endi = getindi(vals,stop)
    if starti is None or endi is None:
        return relx,rely
    relx = xs[starti:(endi+1)]
    rely = ys[starti:(endi+1)]
    return relx,rely




def getvarfromdict(dicty,key,default):
    if key not in dicty:
        return default
    if dicty[key] is None:
        return default
    return dicty[key]


class InputHL(object):
    def __init__(self,root,production,outlierthreshold,othervars):
        self.root = root
        self.production = production
        self.outlierthreshold = outlierthreshold
        self.URR = None
        ######### VAL ######
        self.URRstr = tk.StringVar()
        self.URRstr.set("")
        self.digits = tk.IntVar()
        self.digits.set(0)
        self.startyearstr = tk.StringVar()
        self.endyearstr = tk.StringVar()
        ### here do it.
        self.startyearstr.set(getvarfromdict(othervars,'startyear',''))
        self.startyearval = getvarfromdict(othervars,'startyearval',None)
        self.endyearstr.set(getvarfromdict(othervars,'endyear',''))
        self.endyearval = getvarfromdict(othervars,'endyearval',None)

        self.calculateHLURR()
        self.inputHL = tk.Toplevel(root)
        self.label = tk.Label(self.inputHL,text = HLMESSAGE,fg = TITLECOLOR,bg=FRAMECOLOR,font = TITLEFONT)
        self.label.grid(row = 0,column = 2,padx=PADX, pady=PADY,sticky = '',columnspan = 4)
        self.desc = tk.Label(self.inputHL,text = HLDESC,fg = DESCCOLOR,bg=FRAMECOLOR,font = DESCFONT)
        self.desc.grid(row = 1,column = 2,padx=PADX, pady=PADY,sticky = '',columnspan = 4)
        plt = HLplot(production,outlierthreshold)
        hlplot = FigureCanvasTkAgg(plt.fig,self.inputHL)
        hlplot.get_tk_widget().grid(row = 2,column = 3,columnspan = 3,rowspan = 10)
        hlplot.draw()
        toolframe = tk.Frame(self.inputHL)
        toolframe.grid(row = 12, column = 3,columnspan = 3,padx = PADX,pady = PADY,sticky = 'nswe')
        hltool = NavigationToolbar2Tk(hlplot,toolframe)
        hltool.update()
        hlplot.mpl_connect("motion_notify_event", hover(plt))

        self.startyearlab = tk.Label(self.inputHL,textvariable = self.startyearstr,fg = TITLECOLOR,bg=FRAMECOLOR,font = TITLEFONT)
        self.startyearlab.grid(row = 3,column = 0,padx=PADX, pady=PADY,sticky = '')

        self.endyearlab = tk.Label(self.inputHL,textvariable = self.endyearstr,fg = TITLECOLOR,bg=FRAMECOLOR,font = TITLEFONT)
        self.endyearlab.grid(row = 7,column = 0,padx=PADX, pady=PADY,sticky = '')

        
        self.startyear = tk.Label(self.inputHL,text = STARTYEAR,anchor = 'e',fg = DESCCOLOR,bg = FRAMECOLOR,justify = 'right',font = TEXTFONT)
        self.startyear.grid(row = 1,column = 0,padx=PADX,pady = PADY,sticky = 'w')
        self.startyearbut = tk.Button(self.inputHL,text = "Enter Start Year",command = lambda : self.InputStartYear(root,production.years),font = BUTTONFONT)
        self.startyearbut.grid(row = 2,column = 0,padx = PADX,pady = PADY,sticky = 'nswe')
        

        self.endyear = tk.Label(self.inputHL,text = ENDYEAR,anchor = 'e',fg = DESCCOLOR,bg = FRAMECOLOR,justify = 'right',font = TEXTFONT)
        self.endyear.grid(row = 5,column = 0,padx=PADX,pady = PADY,sticky = 'w')
        self.endyearbut = tk.Button(self.inputHL,text = "Enter End Year",command = lambda : self.InputEndYear(root,production.years),font = BUTTONFONT)
        self.endyearbut.grid(row = 6,column = 0,padx = PADX,pady = PADY,sticky = 'nswe')
        

        self.urrlab = tk.Label(self.inputHL,text = URRLAB,anchor = 'e',fg = DESCCOLOR,bg = FRAMECOLOR,justify = 'right',font = TEXTFONT)
        self.urrlab.grid(row = 9,column = 0,padx=PADX,pady = PADY,sticky = 'w')
        self.urrval = tk.Label(self.inputHL,textvariable = self.URRstr,fg = TITLECOLOR,bg=FRAMECOLOR,font = TITLEFONT)
        self.urrval.grid(row = 10,column = 0,padx=PADX, pady=PADY,sticky = 'nswe')
        self.button_ok = tk.Button(self.inputHL, text=ENTERTEXT, command=self.close,width = DEFAULTINPUTWIDTH)
        self.button_ok.grid(row = 13,column = 4,padx = PADX, pady = PADY,sticky = 'nswe')

    def close(self):
        self.inputHL.destroy()


    def InputStartYear(self,root,years):
        response = InputYearPopup(root,"Start Year",years)
        self.root.wait_window(response.inputYear)
        self.startyearval = response.year
        self.startyearstr.set(response.yearstr)
        self.calculateHLURR()
        
    def InputEndYear(self,root,years):
        response = InputYearPopup(root,"End Year",years)
        self.root.wait_window(response.inputYear)
        self.endyearval = response.year
        self.endyearstr.set(response.yearstr)
        self.calculateHLURR()

    def calculateHLURR(self):
        if self.startyearval is None or self.endyearval is None:
            self.URR = None
            self.URRstr.set("")
            return
        startyr = min((self.startyearval,self.endyearval))
        endyr = max((self.startyearval,self.endyearval))
        if startyr == endyr:
            self.URR = None
            self.URRstr.set("")
            return

        relcum,relprodovercum = splitdata(self.production.cumulative,self.production.prodovercumm,self.production.years,startyr,endyr)
        if relcum == [] or relprodovercum == []:
            self.URR = None
            self.URRstr.set("")
            return
        slope,intercept = plot.getlineofbestfit(relcum,relprodovercum)
        self.URR = -intercept/slope
        urrstring,digits = makeurr(self.URR)
        self.URRstr.set(urrstring)
        self.digits.set(digits)
        #https://stackoverflow.com/questions/12864294/adding-an-arbitrary-line-to-a-matplotlib-plot-in-ipython-notebook
        # try link to add the line of best fit to the plot...
        # update the URR string in a pretty way.
        
        
        #self.production
        #production.cumulative
        #production.prodovercumm

def round_to(x,n):
    digit = (n-1) - int(math.floor(math.log10(abs(x))))
    return round(x, digit),digit

def converttosuperscript(n):
    nstr = str(int(n))
    supern = ''
    for char in nstr:
        supern += SUPERSCRIPTNUMS[char]
    return supern
        
def makeurr(URR):
    if URR is None:
        return "",0
    if URR >= 1000000:
        urrval,digit = round_to(URR,6)
        inturr = str(int(urrval))
        n = len(inturr)-1
        rawstr = inturr[0]+'.'+inturr[1:6] + r'x10'+converttosuperscript(n)
        return rawstr,0
    
    if URR >= 1000:
        return str(int(URR)),0
    value,digits = round_to(URR,5)
    return str(value),digits
        


class InputYearPopup(object):
    def __init__(self,root,yeartype,years):
        self.years = years
        self.yeartype = yeartype
        self.year = None
        self.yearstr = ""
        self.inputYear = tk.Toplevel(root)
        self.inputYear.geometry(POPUPGEOM)
        self.label = tk.Label(self.inputYear, text=self.yearstrfunc(yeartype))
        self.label.grid(padx=PADX, pady=PADY,sticky = 'nsw',columnspan = 2)

        self.entry = tk.Entry(self.inputYear,width = DEFAULTINPUTWIDTH)
        self.entry.grid(padx = PADX,pady = PADY,sticky = 'nswe')
        button_close = tk.Button(self.inputYear, text=ENTERTEXT, command=self.checkinputYear,width = DEFAULTINPUTWIDTH)
        button_close.grid(padx = PADX, pady = PADY,sticky = 'nswe')

    def yearstrfunc(self,yeartype):
        return 'Enter '+yeartype

        
    def checkinputYear(self):
        try:
            year = int(self.entry.get())
            yearstr = str(year)
        except:
            year = None
            yearstr = ""
        if year not in self.years or year is None:
            self.label.config(text = 'Year not found, enter '+self.yeartype)
        else:
            self.year = year
            self.yearstr = yearstr
            self.inputYear.destroy()




class InputScenarioPopup(object):
    def __init__(self,root,message):
        self.scenario = ''
        self.inputscenario = tk.Toplevel(root)
        self.inputscenario.geometry(POPUPGEOM)
        self.label = tk.Label(self.inputscenario, text=self.messagefunc(message))
        self.label.grid(padx=PADX, pady=PADY,sticky = 'nsw',columnspan = 2)

        self.entry = tk.Entry(self.inputscenario,width = DEFAULTINPUTWIDTH)
        self.entry.grid(padx = PADX,pady = PADY,sticky = 'nswe')
        button_close = tk.Button(self.inputscenario, text=ENTERTEXT, command=self.checkscenario,width = DEFAULTINPUTWIDTH)
        button_close.grid(padx = PADX, pady = PADY,sticky = 'nswe')

    def messagefunc(self,message):
        return 'Enter '+message

        
    def checkscenario(self):
        try:
            self.scenario = str(self.entry.get())
        except:
            self.scenario = ''
        self.inputscenario.destroy()







        
class InputHLold(object):
    def __init__(self,root,production,outlierthreshold):
        self.root = root
        self.production = production
        self.outlierthreshold = outlierthreshold
        self.URR = None
        self.inputHL = tk.Toplevel(root)
        plt = HLplot(production,outlierthreshold)
        
        hlplot = FigureCanvasTkAgg(plt.fig,self.inputHL)
        hlplot.get_tk_widget().pack()
        hlplot.draw()
        #toolframe = tk.Frame(self.inputHL)
        #toolframe.grid(row = 11, column = 2,padx = PADX,pady = PADY,sticky = 'nswe')
        hltool = NavigationToolbar2Tk(hlplot,self.inputHL)
        hltool.update()
        hlplot.get_tk_widget().pack()
        hlplot.mpl_connect("motion_notify_event", hover(plt))


def update_annot(fig,bub):
    fig.annot.xy = (bub.x,bub.y)
    fig.annot.set_text(bub.label)
    #fig.annot.get_bbox_patch().set_facecolor(cmap(norm(c[ind["ind"][0]])))
    fig.annot.get_bbox_patch().set_alpha(0.4)

        
def hover(fig):
    def hover_inner(event):
        vis = fig.annot.get_visible()
        if event.xdata is None:
            fig.annot.set_visible(False)
            fig.fig.canvas.draw_idle()
            return
        for bub in fig.bubbles:
            if bub.startx > event.xdata:
                fig.annot.set_visible(False)
                fig.fig.canvas.draw_idle()
                continue
            if bub.endx < event.xdata:
                fig.annot.set_visible(False)
                fig.fig.canvas.draw_idle()
                continue
            if bub.starty > event.ydata:
                fig.annot.set_visible(False)
                fig.fig.canvas.draw_idle()
                continue
            if bub.endy < event.ydata:
                fig.annot.set_visible(False)
                fig.fig.canvas.draw_idle()
                continue
            update_annot(fig,bub)
            fig.annot.set_visible(True)
            fig.fig.canvas.draw_idle()
            return
    return hover_inner
        
class InputHLURR(object):
    def __init__(self,root,production,outlierthreshold,othervars):
        self.URR = None
        self.digits = None
        self.URRstr = None
        self.method = None
        self.startyear = None
        self.endyear = None
        self.startyearval = None
        self.endyearval = None
        self.root = root
        self.production = production
        self.outlierthreshold = outlierthreshold
        self.othervars = othervars
        self.inputHLURR = tk.Toplevel(root)
        self.inputHLURR.geometry(POPUPGEOM)
        self.label = tk.Label(self.inputHLURR, text=URRCHOOSEMESSAGE)
        self.label.grid(padx=PADX, pady=PADY,sticky = 'nsw',columnspan = 2)
        
        button_HL = tk.Button(self.inputHLURR, text=HLTEXT, command=self.HLpopup,width = DEFAULTINPUTWIDTH)
        button_HL.grid(row = 1,column = 0,padx = PADX, pady = PADY,sticky = 'nswe')

        button_input = tk.Button(self.inputHLURR, text = MANUALURRINPUT, command = self.manualinput,width = DEFAULTINPUTWIDTH)
        button_input.grid(row = 1,column = 1,padx = PADX, pady = PADY,sticky = 'nswe')


    def HLpopup(self):
        ## need to think how to do the HL trend --> plot HL graph, and calculate the estimated URR
        response = InputHL(self.root,self.production,self.outlierthreshold,self.othervars)
        self.root.wait_window(response.inputHL)
        self.URR = response.URR
        self.URRstr = response.URRstr
        self.method = URRHLMETHOD
        self.digits = response.digits
        self.startyear = response.startyearstr.get()
        self.endyear = response.endyearstr.get()
        self.startyearval = response.startyearval
        self.endyearval = response.endyearval
        self.inputHLURR.destroy()


    def manualinput(self):
        response = InputURR(self.root,self.othervars)
        self.root.wait_window(response.inputURR)
        self.URR = response.URR
        urrstring,digits = makeurr(self.URR)
        self.URRstr = urrstring
        self.method = URRINPUTMETHOD
        self.digits = digits
        self.inputHLURR.destroy()


        

class InputURR(object):
    def __init__(self,root,othervars):
        self.URR = None
        self.inputURR = tk.Toplevel(root)
        self.inputURR.geometry(POPUPGEOM)
        self.label = tk.Label(self.inputURR, text=URRINPUTMESSAGE)
        self.label.grid(padx=PADX, pady=PADY,sticky = 'nsw',columnspan = 2)

        self.entry = tk.Entry(self.inputURR,width = DEFAULTINPUTWIDTH)
        if 'urrmethod' in othervars:
            if othervars['urrmethod'] == URRINPUTMETHOD:
                self.entry.insert(tk.END,othervars['urrval'])
        self.entry.grid(padx = PADX,pady = PADY,sticky = 'nswe')
        button_close = tk.Button(self.inputURR, text=ENTERTEXT, command=self.checkinputURR,width = DEFAULTINPUTWIDTH)
        button_close.grid(padx = PADX, pady = PADY,sticky = 'nswe')
        
    def checkinputURR(self):
        try:
            urr = float(self.entry.get())
        except:
            urr = None
        #if urr is None:
        #    return
        self.URR = urr
        if urr is not None:
            self.inputURR.destroy()
        else:
            self.label.config(text = URRINPUTERRORMESSAGE)
    
def acceptscenario(root,frames,mainclass,frame,container,skipurr = False,suppresswarnings = False):
    cls = createcls(root,frames,mainclass,frame)
    if cls is None:
        return
    urrkey = makekey(cls,True)
    if urrkey in mainclass.scenarios:
        if not suppresswarnings:
            messagebox.showwarning(title = FAILTITLE,message = 'Scenario already exists')
            #return
    mainclass.scenarios[urrkey] = frame
    prodkey = makekey(cls,False)
    urr = None
    hideelems = copy.deepcopy(EDITSCENHIDE)
    if not skipurr:
        if urrkey not in mainclass.URR and prodkey not in mainclass.production and frames[frame].othervars[EXCELURRSTR] is None:
            response = messagebox.askyesno(title=WARNINGTITLE, message='Cannot locate URR or production for scenario, continue?')
            if response == False:
                return
            response = InputURR(root,frames[frame].othervars)
            root.wait_window(response.inputURR)
            if response.URR is None:
                return
            URRstring,digits = makeurr(response.URR)
            frames[frame].elems[URRLABEL].elem.config(text = URRstring)
            frames[frame].othervars['digits'] = digits
            frames[frame].othervars['urrval'] = response.URR
            frames[frame].othervars['urrmethod'] = URRINPUTMETHOD
            frames[frame].othervars['startyear'] = None
            frames[frame].othervars['endyear'] = None
            frames[frame].othervars['startyearval'] = None
            frames[frame].othervars['endyearval'] = None
            
            #hideelems.append(URRBOX)
            #hideelems.append(HLBUTTON)
            ### Load manual input option only.
        elif urrkey not in mainclass.URR and frames[frame].othervars[EXCELURRSTR] is None:
            #### GOT THIS ONE.
            response = InputHLURR(root,mainclass.production[prodkey],mainclass.outlierthreshold[prodkey],frames[frame].othervars)
            root.wait_window(response.inputHLURR)
            if response is None:
                return
            URRstring,digits = makeurr(response.URR)
            frames[frame].elems[URRLABEL].elem.config(text = URRstring)
            frames[frame].othervars['digits'] = digits
            frames[frame].othervars['urrval'] = response.URR
            frames[frame].othervars['urrmethod'] = response.method
            frames[frame].othervars['startyear'] = response.startyear
            frames[frame].othervars['endyear'] = response.endyear
            frames[frame].othervars['startyearval'] = response.startyearval
            frames[frame].othervars['endyearval'] = response.endyearval
            
            ## production exists, URR not, warn, but allow HL or manual input...
            #hideelems.append(URRLABEL)        
        elif frames[frame].othervars[EXCELURRSTR] is None: 
            urr = mainclass.URR[urrkey]
            URRstring,digits = makeurr(urr.URR)
            #hideelems.append(URRBOX)
            #hideelems.append(HLBUTTON)
            frames[frame].elems[URRLABEL].elem.config(text = URRstring)
            frames[frame].othervars['digits'] = digits
            frames[frame].othervars['urrval'] = urr.URR
            frames[frame].othervars['urrmethod'] = URRLOOKUP
            frames[frame].othervars['startyear'] = None
            frames[frame].othervars['endyear'] = None
            frames[frame].othervars['startyearval'] = None
            frames[frame].othervars['endyearval'] = None
            ## Just have a label with the dedicated URR value.
        else:
            urr = frames[frame].othervars[EXCELURRSTR]
            URRstring,digits = makeurr(urr)
            frames[frame].elems[URRLABEL].elem.config(text = URRstring)
            frames[frame].othervars['digits'] = digits
            frames[frame].othervars['urrval'] = urr
            frames[frame].othervars['urrmethod'] = URREXCEL
            frames[frame].othervars['startyear'] = None
            frames[frame].othervars['endyear'] = None
            frames[frame].othervars['startyearval'] = None
            frames[frame].othervars['endyearval'] = None
                
    relabelterms(frames,frame,cls)
    #.ref
    frames[frame].frombutton.elem.config(text = urrkey)
    frames[frame].initialise = False
    frames[frame].hideelems = hideelems
    forgetframe(root,frames,frame)
    loadframe(root,frames,mainclass,frame,container)
##    return root,frames

def editscenariopage(root,container,canvas,frames,mainclass,frame):
    ## title
    root,frames = maketitle(root,frames,frame,MINESTITLE)
    ## Scenario
    root,frames = createscenarios(root,frames,mainclass,frame,SCENARIOTERMS,SCENARIO,SCENTITLE,width = SCENARIOWIDTHS)
    ## scenario labels
    
    ## accept scenario button
    acceptbutton = elemcls(row = ACCEPTSCENARIOLOC[0],
                           column = ACCEPTSCENARIOLOC[1],
                           padx = PADX,
                           pady = PADY,
                           text = ACCEPTSCENARIOLABEL,
                           obj = 'button',
                           frame = frames[frame].frame,
                           command = lambda : acceptscenario(root,frames,mainclass,frame,container),
                           font = BUTTONFONT)
    acceptbutton.grid()
    frames[frame].elems[ACCEPTBUTTON] = acceptbutton
    ## Edit scenario button
    editbutton = elemcls(row = EDITSCENARIOLOC[0],
                         column = EDITSCENARIOLOC[1],
                         padx = PADX,
                         pady = PADY,
                         text = EDITSCENARIOLABEL,
                         obj = 'button',
                         frame = frames[frame].frame,
                         command = lambda : editscenario(root,frames,mainclass,frame,container),
                         font = BUTTONFONT)
    editbutton.grid()
    frames[frame].elems[EDITBUTTON] = editbutton
    ## Save button
    root,frames = savebuttons(root,frames,frame,mainclass)
    ## back button
    root,frames = backbutton(root,frames,mainclass,frame,SELECTMODELFRAME,container,implementsaveback,SCENARIOTAG,backbuttontext = ACCEPT)
    ## back button
    root,frames = runscenariobutton(root,container,frames,mainclass,frame)
    root,frames = HLbutton(root,frames,frame)
    ## OutputImage
    root,frames = outputimages(root,frames,mainclass,frame,container)
    ## initfunc
    frames[frame].initfunc = editscenario
    ## cyclelist
    root,frames = makecycles(root,container,canvas,frames,mainclass,frame,minecyclefunc,CYCLELOC,CYCLELOCSPAN,NEWCYCLELOC,NEWCYCLETEXT,CYCLEBUTTON,CYCLETAG,scenariocyclefunc,CYCLESIZE,scenariocyclepage,'scen scen scen scen scen')
    forgetframe(root,frames,frame)


def scrollfunc(event):
    canvas.configure(scrollregion=canvas.bbox('all'),width=200,height=200)

def fieldspage(root,canvas,frames):
    pass


def makedesc(root,frames,frame,desc):
    for r in range(0,len(desc)):
        maindesc = elemcls(row = DESCLOC[0]+r,
                           column = DESCLOC[1],
                           sticky = 'w',
                           columnspan = GRIDSIZE[1],
                           padx = PADX,
                           obj = 'label',
                           frame = frames[frame].frame,
                           text = desc[r],
                           fg = DESCCOLOR,
                           justify = 'left',
                           anchor = 'w',
                           bg = FRAMECOLOR,
                           font = DESCFONT)
        maindesc.grid()
        frames[frame].elems[MAINDESC+str(r)] = maindesc
    return root,frames

    
def generalscenariofunc(scenario):
    return scenario

def excelpostprocessorbutton(root,frames,mainclass,frame,container):
    excelbutton = elemcls(frame = frames[frame].frame,
                         obj = 'button',
                         row = EXCELBUTTONLOC[0],
                         column = EXCELBUTTONLOC[1],
                         padx = PADX,
                         pady = PADY,
                         text = EXCELBUTTONTEXT,
                         command = lambda : excelpostprocess(root,frames,mainclass,frame,container),
                         font = BUTTONFONT)
    excelbutton.grid()
    frames[frame].elems[EXCELTAG] = excelbutton
    return root,frames

def pdfpostprocessorbutton(root,frames,mainclass,frame,container):
    pdfbutton = elemcls(frame = frames[frame].frame,
                         obj = 'button',
                         row = PDFBUTTONLOC[0],
                         column = PDFBUTTONLOC[1],
                         padx = PADX,
                         pady = PADY,
                         text = PDFBUTTONTEXT,
                         command = lambda : pdfpostprocess(root,frames,mainclass,frame,container),
                         font = BUTTONFONT)
    pdfbutton.grid()
    frames[frame].elems[PDFTAG] = pdfbutton
    return root,frames


def getallyearsfromprojections(mainclass):
    minyear = None
    maxyear = None
    for scen in mainclass.modelresults:
        for key in mainclass.modelresults[scen]:
            minyear,maxyear = updateminmax(mainclass.modelresults[scen][key].years,minyear,maxyear)
            if key in mainclass.production:
                minyear,maxyear = updateminmax(mainclass.production[key].years,minyear,maxyear)
    years = list(range(int(minyear),int(maxyear+1)))
    return years


def updateminmax(years,minyear,maxyear):
    if len(years) == 0:
        return minyear,maxyear
    mini = min(years)
    maxi = max(years)
    if minyear is None:
        minyear = mini
    if maxyear is None:
        maxyear = maxi
    if mini < minyear:
        minyear = mini
    if maxi > maxyear:
        maxyear = maxi
    return minyear,maxyear

    

def excelpostprocess(root,frames,mainclass,frame,container):
    filename =  filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("Excel","*.xlsx"),))
    if filename is None or filename == '':
        return 0
    if not filename.upper().endswith('.XLSX'):
        filename = filename + '.xlsx'
    wb = openpyxl.Workbook()
    
    try:
        wb.save(filename)
    except:
        popupmessage(SAVEFAILED,'Failed to delete existing file')
        return

    sheetmap = {}
    goodsheetnames = []
    for scen in mainclass.modelresults.keys():
        wsscen = makecleanexcelsheetname(scen,wb.sheetnames)
        ws = wb.create_sheet(wsscen)
        sheetmap[scen] = wsscen
        goodsheetnames.append(wsscen)
    sheets = wb.sheetnames
    for sht in sheets:
        if sht in goodsheetnames:
            continue
        del wb[sht]

    for scen in mainclass.modelresults:
        wsscen = sheetmap[scen]
        excelpostprocess_sub(wb[wsscen],mainclass,scen)
    wb.save(filename)
    
def excelpostprocess_sub(ws,mainclass,scenario):    
    allyears = getallyearsfromprojections(mainclass)
    allkeys = list(mainclass.modelresults[scenario].keys())    
    allkeys = sorted(allkeys)
    if DEMANDKEY in allkeys:
        allkeys.remove(DEMANDKEY)
        allkeys.append(DEMANDKEY)
    
    for i in range(0,len(SCENARIOFIELDORDER)):
        ws.cell(row = i+1,column = 1).value = SCENARIOFIELDORDER[i]
    ws.cell(row = len(SCENARIOFIELDORDER)+1,column = 1).value = 'Model/Data'
    row = len(SCENARIOFIELDORDER)+1
    yeardict = {}
    for year in allyears:
        row += 1
        yeardict[year] = row
        ws.cell(row = row,column = 1).value = year
    col = 1
    for key in allkeys:
        if key in mainclass.production:
            col += 1
            writeexcel(ws,mainclass.production[key],col,'Production',yeardict)
        col += 1
        writeexcel(ws,mainclass.modelresults[scenario][key],col,'Model',yeardict)


def combinedata(datadict,yearprods):
    '''
    input is a list of dictionaries.
    dictionaries are in the form of year:val
    '''
    if len(yearprods)==0:
        return datadict
    for pair in yearprods:
        years,prods = pair
        for i in range(0,len(years)):
            year = years[i]
            prod = prods[i]
            if year not in datadict:
                datadict[year] = 0
            datadict[year] += prod
    return datadict

def getcoloralpha(c,rawalpha,rawcolor,colorlist):
    if rawcolor == None:
        c += 1
        if c >= len(colorlist):
            c = 0
            rawalpha = rawalpha - ALPHASTEP
        color = colorlist[c]
        alpha = rawalpha
    else:
        color = rawcolor
        alpha = 1
    return color,alpha,c,rawalpha

def getxyscenario(years,prods):
    xs = []
    ys = []
    if years == []:
        return xs,ys
    lastyr = int(min(years))-1 
    for i in range(0,len(years)):
        yr = years[i]
        prod = prods[i]
        for y in range(lastyr+1,int(yr)):
            xs.append(y)
            ys.append(0)
        xs.append(yr)
        ys.append(prod)
        lastyr = int(yr)
    return xs,ys



def graphN(relscenarios,name,n=None,xticks = None,legs = [MINERALTERM,SUBMINERALTERM,REGIONTERM,SUBREGIONTERM],colorlist = COLORLIST,orderform = None,ylabel = ''):
    order = getorder(orderform,relscenarios)
    topn,scen = gettopn(relscenarios,order,n)
    plots = []    
    c = -1
    rawalpha = 1
    histdict = {}
    histdatax = []
    histdatay = []
    totproj = {}
    for j in range(0,len(topn)):
        legend = getlegend(scen[topn[j][0]],legs)
        prod = scen[topn[j][0]].production
        years = scen[topn[j][0]].years
        totproj = combinedata(totproj,[(years,prod)])
        color,alpha,c,rawalpha = getcoloralpha(c,rawalpha,scen[topn[j][0]].color,colorlist)
        x,y = getxyscenario(years,prod)
        p = plot.area(x = x,y = y,legend = legend, color = color,alpha = alpha)
        plots.append(p)
        # STEVE FIX HERE
        histdatax,histdatay,histdict = addtohistdata(scen[topn[j][0]].data,histdatax,histdatay,histdict)
    if len(histdatax) > 0:
        pd = plot.line(x = histdatax,y = histdatay,color = DATACOLOR,style = DATASTYLE,legend = DATALABEL)
        plots.append(pd)
    firstx,lastx = getminmax(totproj)    
    plot.display(plots,name,xlabel = 'Year',ylabel = ylabel,ext = 'eps',xminmax = (firstx,lastx),ncols = NCOLS,yzero = True)
    plot.display(plots,name,xlabel = 'Year',ylabel = ylabel,ext = 'png',xminmax = (firstx,lastx),ncols = NCOLS,yzero = True)



def addtohistdata(data,histdatax,histdatay,histdict):
    if data is None:
        return histdatax,histdatay,histdict
    for key in data:
        if key not in histdict:
            histdict[key] = 0
        histdict[key] += data[key]
        histdatax.append(key)
        val = copy.deepcopy(histdict[key])
        histdatay.append(val)
    return histdatax,histdatay,histdict
        
    



def getfirst(xs,maxxy,proj):
    for x in xs:
        y = proj[x]
        if y > maxxy*THRESHOLD:
            return x
    
def getfirstlastx(proj,maxxy):
    xs = sorted(list(proj.keys()))
    firstx = getfirst(xs,maxxy,proj)
    xs = sorted(xs,reverse = True)
    lastx = getfirst(xs,maxxy,proj)
    return firstx,lastx

def getminmax(proj):
    vals = list(proj.values())
    if len(vals) == 0:
        return None
    maxy = max(vals)
    if maxy < 0:
        return None
    firstx,lastx = getfirstlastx(proj,maxy)
    firstx = STEP*math.floor(firstx/float(STEP))
    lastx = STEP*math.ceil(lastx/float(STEP))
    return (firstx,lastx)


        

def writeexcel(ws,prodcls,col,tag,yeardict):
    for r in range(0,len(SCENARIOFIELDORDER)):
        ws.cell(row = r + 1,column = col).value = getattr(prodcls,SCENARIOFIELDORDER[r])
    ws.cell(row = len(SCENARIOFIELDORDER)+1,column = col).value = tag
    for i in range(0,len(prodcls.years)):
        year = int(prodcls.years[i])
        row = yeardict[year]
        prod = prodcls.production[i]
        ws.cell(row = row,column = col).value = prod
        
    

def pdfpostprocess(root,frames,mainclass,frame,container):
    for scen in mainclass.modelresults:
        pdfpostprocess_sub(root,frames,mainclass,frame,container,scen)

def pdfpostprocess_sub(root,frames,mainclass,frame,container,scenario):
    if not shutil.which(LATEX):
        popupmessage(SAVEFAILED,'Need to install LaTeX')
        return    
    filename =  filedialog.asksaveasfilename(initialdir = "/",title = "Select file for "+scenario,filetypes = (("pdf","*.pdf"),("tex","*.tex")))
    if filename is None or filename == '':
        return 0
    filenoext, ext = os.path.splitext(filename)
    if ext == 'PDF':
        filename = filenoext + '.tex'
    if not filename.upper().endswith('.TEX'):
        filename = filename + '.tex'
    folder,meh = os.path.splitext(filename)
    figfolder =  os.path.basename(folder)
    folderpath = os.path.dirname(folder)
    tname = figfolder+'.tex'
    ## filename (tex) folder, filenoext
    passes = makelatexfolders(folder)
    if not passes:
        return
    units = frames[NEWMODELFRAME].elems[LOADUNITS+INPUTTERMSENTRYTAG].elem.get()
    ylabel = YLABELSTART + units + YLABELEND
    
    minerals,minorder = sortscenarios(mainclass.modelresults[scenario],MINERALTERM,SUBMINERALTERM,order = 'URR',skipkeys = [DEMANDKEY,RECYCLEKEY])
    continents,contorder = sortscenarios(mainclass.modelresults[scenario],CONTINENTTERM,COUNTRYTERM,skipkeys = [DEMANDKEY,RECYCLEKEY])
    recycle = None
    if RECYCLEKEY in mainclass.modelresults[scenario]:
        recycle = mainclass.modelresults[scenario][RECYCLEKEY]
    tex = mlatex.initialise(filename,dclass = 'report',packages = ['graphicx','booktabs',('hyperref','hidelinks'),('geometry','total={7in,9in},top=1in, left=0.7in'),'natbib'])
    maketitlepage(tex,scenario)
    tex.write('\\setcounter{tocdepth}{3}\n')
    tex.write('\\tableofcontents\n')
    if len(minorder)>1:
        makeurrtable(tex,frames,mainclass,scenario,continents,contorder,minerals,minorder)
        makecountries(mainclass,scenario,continents,contorder,minerals,minorder,folder,figfolder,tex,'Overall',recycle,ylabel = ylabel)
    for miner in minorder:
        subminerals = {}
        subminerals[miner] = minerals[miner]
        makeurrtable(tex,frames,mainclass,scenario,continents,contorder,subminerals,[miner],mineraltoselect = miner)
        makecountries(mainclass,scenario,continents,contorder,subminerals,[miner],folder,figfolder,tex,miner,recycle,ylabel = ylabel,mineraltoselect = miner)
    if mainclass.bibtextfile is not None:
        fixbibtex(mainclass.bibtextfile,tex,folder,figfolder)
    mlatex.end(tex,figfolder)
    createreport(folderpath,figfolder)


def makeurrtable(tex,frames,mainclass,scenario,continents,contorder,minerals,minorder,mineraltoselect = None):
    header = [[('Region','b'),('URR','b'),('Comment','b')]]
    data = []
    toturr = 0
    for continent in contorder:
        countries = sorted(continents[continent])
        for country in countries:    
            relevantscenarios,regions = getspecificscenarios(mainclass.modelresults[scenario],{'continent':continent,'country':country},mineraltoselect)
            if relevantscenarios == {}:
                continue
            order = getorder(None,relevantscenarios)
            for pair in order:
                key = pair[0]                
                legend = getlegend(relevantscenarios[key],['mineral','submineral','country','region','subregion'],ignoreleg = True)
                urr = nicestr(relevantscenarios[key].urr)
                ref = getref(mainclass,frames,scenario+key)
                data.append([legend,urr,ref])
                toturr += relevantscenarios[key].urr
    data.append([('Total','b'),(nicestr(toturr),'b'),''])
    mlatex.table(tex,data,header,'lrr',caption = 'URR values',borders = 'Minimal')        
    mlatex.clear(tex)

def getref(mainclass,frames,key):
    if key not in mainclass.scenarios:
        print('Oh I thought this wasnt possible')
        return ''
    frame = mainclass.scenarios[key]
    if 'urrmethod' not in frames[frame].othervars:
        print('Oh I thought this too wasnt possible')
        return ''
    method = frames[frame].othervars['urrmethod']
    if method == URRLOOKUP:
        if key not in mainclass.URR:
            return ''
        if mainclass.URR[key] is None:
            return ''
        return mainclass.URR[key]
    if method is None:
        return ''
    return method







    
def fixbibtex(bibtex,tex,folder,basefolder):
    bibname = os.path.basename(bibtex)
    shutil.copy(bibtex,folder+os.sep+bibname)
    tex.write('\\bibliographystyle{apa}')
    tex.write('\\bibliography{'+basefolder+'/'+bibname+'}')



def regionalsection(tex,folder,figfolder,tag,continent,country,reg,relevantscenarios,minerals,minorder,colorlist,mineraltoselect):
    mlatex.subsubsection(tex,clean(reg))
    regscenarios,ignore = getspecificscenarios(relevantscenarios,{'continent':continent,'country':country,'region':reg},mineraltoselect)
    graphN(regscenarios,folder+os.sep+tag+country+'_'+reg+'_region',n=N,legs = ['mineral','submineral','subregion'])
    mlatex.figure(tex,[figfolder+'/'+tag+country+'_'+reg+'_region.eps'],clean(country) + ' - '+clean(reg) + ' projections capped at '+str(N))
    tabularise(tex,regscenarios,'Peak years - '+ALL)
    subtypescenarios = getsubtypescenarios(regscenarios,minerals,colorlist)
    graphN(subtypescenarios,folder+os.sep+tag+country+'_'+reg+'minerals',n=N,legs = ['mineral','submineral','subregion'],orderform = [('mineral',minorder),('submineral',minerals)])
    mlatex.figure(tex,[figfolder+'/'+tag+country+'_'+reg+'minerals.eps'],clean(country) + ' - ' + clean(reg) + ' projection by mineral type')
    tabularise(tex,subtypescenarios,'Peak years - Minerals',orderform = [('mineral',minorder),('submineral',minerals)])
    regionscenario = getregionscenario(regscenarios,reg)
    mlatex.clear(tex)
    return regionscenario


def getregionscenario(regscenarios,reg):
    keys = list(regscenarios.keys())
    if len(keys) == 0:
        return
    regionscenario = copyscenario(regscenarios[keys[0]])
    for i in range(1,len(keys)):
        key = keys[i]
        regionscenario = combinescenarios(regionscenario,regscenarios[key],joindata=True)
    regionscenario.legend = reg
    regionscenario.color = None
    return regionscenario  


def createreport(folder,filename):
    if RUNLATEX:        
        current = os.getcwd()
        os.chdir(folder)
        for i in range(0,3):
            os.system(LATEX +' --draftmode '+ filename+'.tex')
            os.system(BIBTEX + ' ' + filename)
        os.system(LATEX + ' ' + filename+'.tex')
        os.system(BIBTEX + ' ' + filename)
        os.system(DVIPS + ' ' + filename+'.dvi')
        os.system(PSPDF + ' ' + filename+'.ps')
        os.chdir(current)
        
def nicestr(floaty,decimals = 2):
    val = floaty*pow(10,decimals)
    intval = int(val)
    remainder = val - intval 
    if remainder >=0.5:
        intval += 1
    if intval == 0 and remainder < 0.5:
        return '--'
    return str(intval/float(pow(10,decimals)))
        
    

def tabularise(tex,scenarios,caption,orderform = None,legs = ['mineral','submineral','region','subregion']):
    order = getorder(orderform,scenarios)
    tabledata = []
    header = [[('Name','b'),('URR','b'),('Peak Year','b'),('Peak Rate','b')]]
    totURR = 0
    production = {}
    for pair in order:
        key = pair[0]
        legend = getlegend(scenarios[key],legs)
        totURR += scenarios[key].urr
        URR = nicestr(scenarios[key].urr)
        peakyear = str(scenarios[key].peakyear)
        peakrate = nicestr(scenarios[key].peakrate)
        years = scenarios[key].years
        prod = scenarios[key].production
        production = combinedata(production,[(years,prod)])
        tabledata.append([legend,URR,peakyear,peakrate])
    peakyear,peakrate = getpeakdata(production)
    tabledata.append([('Total','b'),(nicestr(totURR),'b'),(str(peakyear),'b'),(nicestr(peakrate),'b')])
    mlatex.table(tex,tabledata,header,'lrrr',caption = caption,borders = 'Minimal')

def gethistoricdata(data1,data2,joindata):
    if not joindata:
        data = None
        if data1!=None:
            data = data1
        if data2!=None:
            data = data2
        return data
    datas = []
    if data1!= None:
        datas.append(data1)
    if data2!= None:
        datas.append(data2)
    return combinedatadicts(datas)


def combinedatadicts(datas):
    '''
    input is a list of dictionaries.
    dictionaries are in the form of year:val
    '''
    if len(datas)==0:
        return None
    years = []
    for data in datas:
        years += data.keys()
    years = list(set(years))
    combineddata = {}
    for year in years:
        val = 0
        for data in datas:
            if year in data:
                val += data[year]
        combineddata[year] = val
    return combineddata



def combinescenarios(scenario1,scenario2,joindata = False,color = None,legend = None,subreg = None):
    data = gethistoricdata(scenario1.data,scenario2.data,joindata)
    urr = scenario1.urr + scenario2.urr
    totproj = {}
    totproj = combinedata(totproj,[(scenario1.years,scenario1.production)])    
    totproj = combinedata(totproj,[(scenario2.years,scenario2.production)])
    peakyear,peakrate = getpeakdata(totproj)
    scenario = productioncls()
    for term in ['continent','country','region','mineral','submineral']:
        setattr(scenario,term,getattr(scenario1,term))
    syears = sorted(list(totproj.keys()))
    prods = []
    for year in syears:
        prods.append(totproj[year])
    scenario.color = color
    scenario.legend = legend
    scenario.urr = urr
    scenario.peakyear = peakyear
    scenario.peakrate = peakrate
    scenario.subregion = subreg
    scenario.years = syears
    scenario.production = prods
    scenario.data = data
    return scenario


def getpeakdata(projection):
    peakyear = None
    peakrate = None
    for year in projection.keys():
        if peakrate == None:
            peakyear = year
            peakrate = projection[year]
        elif peakrate < projection[year]:
            peakyear = year
            peakrate = projection[year]
    return peakyear,peakrate

def getsubtypescenarios(scenarios,minerals,colorlist):
    relscenarios = {}
    for mine in minerals.keys():
        for sub in minerals[mine]:
            temp = []
            for k in scenarios.keys():
                if scenarios[k].mineral == mine and scenarios[k].submineral == sub:
                    temp.append(scenarios[k])
            if len(temp) == 0:
                continue
            scen = copyscenario(temp[0])
            for i in range(1,len(temp)):
                scen = combinescenarios(scen,temp[i],joindata = True)
            scen.region = ALL
            scen.subregion = ALL
            scen.color = colorlist[mine][sub]
            relscenarios[mine+','+sub] = scen
            
    return relscenarios


def updatecontinentscenarios(subtypescenarios,countryscen,mineralscen,country,colorlist,mint = None):
    scen = None
    for mintype in subtypescenarios.keys():
        if mintype == RECYCLEKEY:
            if mint != None:
                color = colorlist[RECYCLEKEY][RECYCLEKEY]
                mineralscen[RECYCLEKEY] = subtypescenarios[mintype]
                legend = RECYCLE
                mineralscen[RECYCLEKEY].legend = RECYCLE
        elif not mintype in mineralscen:
            mineralscen[mintype] = copyscenario(subtypescenarios[mintype])
            mine,sub = mintype.split(',')
            mineralscen[mintype].color = colorlist[mine][sub]
            bits = mintype.split(',')
            if len(bits) == 2:
                if bits[0] == bits[1]:
                    leg = bits[0]
                else:
                    leg = mintype.replace(',',' ')
            else:
                leg = mintype.replace(',',' ')
            mineralscen[mintype].legend = mintype.replace(',',' ')
        else:
            color = mineralscen[mintype].color
            legend = mineralscen[mintype].legend
            mineralscen[mintype] = combinescenarios(mineralscen[mintype],subtypescenarios[mintype],joindata = True,color = color,legend = legend)
        leg = country
        if mintype == RECYCLEKEY:
            leg = RECYCLE
        if scen == None:
            scen = copyscenario(subtypescenarios[mintype])
            scen.color = None
            scen.legend = leg
        else:
            scen = combinescenarios(scen,subtypescenarios[mintype],joindata=True,legend = leg)
    if scen!=None:
        countryscen[country] = scen
    return countryscen,mineralscen        

    
def makelatexfolders(folder):
    if os.path.exists(folder):
        msg = messagebox.askyesnocancel(title = 'Warning', message = 'Ok to write over files in: '+folder)
        if msg is None:
            return False
        if msg == False:
            return False
    else:
        try:
            os.mkdir(folder)
        except:
            popupmessage(SAVEFAILED,'Failed to create Figure folder')
            return False
    return True


def getlegend(scenario,legs,ignoreleg = False):
    if scenario.legend != None and not ignoreleg:
        return clean(scenario.legend)
    names = []
    for key in legs:
        elem = getattr(scenario,key)
        if elem != ALL and elem not in names:
            names.append(elem)
    if len(names)==0:
        return None
    legend = names[0]
    for i in range(1,len(names)):
        legend += ' ' + names[i]
    return clean(legend)


def gettopn(relscenarios,order,n):
    if n == None:
        topn = order
        scen = relscenarios
    elif len(order) <=n:
        topn = order
        scen = relscenarios
    else:        
        topn = order[0:n]
        others = order[n:]
        scen = {}
        for pair in topn:
            scen[pair[0]] = relscenarios[pair[0]]
        rest = relscenarios[others[0][0]]
        recy = []
        for i in range(1,len(others)):
            if others[i][0] == RECYCLEKEY:
                recy = [(RECYCLEKEY,None)]
                scen[RECYCLEKEY] = relscenarios[RECYCLEKEY]
                continue
            rest = combinescenarios(rest,relscenarios[others[i][0]],True)
            for term in ['continent','country','region','subregion','mineral','submineral']:
                setattr(rest,term,REST)
        scen[REST] = rest
        topn.append((REST,0))
        topn += recy
    return topn,scen





def makecountries(mainclass,scenario,continents,contorder,minerals,minorder,folder,figfolder,tex,tag,recycle,ylabel = '',mineraltoselect = None):
    # latexname = tag...
    colorlist = getcolorlist(minerals,minorder,recycle)
    continentscen = {}
    worldmineralscen = {}
    worldbycountries = {}
    for continent in contorder:
        countries = sorted(continents[continent])
        countryscen = {}
        mineralscen = {}
        inside = True
        for country in countries:    
            if inside:
                print(continent)
                if mineraltoselect is None:
                    mlatex.chapter(tex,clean(continent))
                else:
                    mlatex.chapter(tex,clean(mineraltoselect)+' ' +clean(continent))
                inside = False
            print('\t'+country)
            relevantscenarios,regions = getspecificscenarios(mainclass.modelresults[scenario],{'continent':continent,'country':country},mineraltoselect)
            if relevantscenarios == {}:
                continue
            mlatex.section(tex,clean(country))
            mlatex.subsection(tex,ALL+' Projections')
            graphN(relevantscenarios,folder+os.sep+tag+country+'_'+ALL+'_Projections',n=NFIGS,ylabel = ylabel)
            mlatex.figure(tex,[figfolder+'/'+tag+country+'_'+ALL+'_Projections.eps'],clean(country) + ' projections capped at '+str(NFIGS))
            tabularise(tex,relevantscenarios,'Peak years - '+ALL)
            mlatex.clear(tex)
            mlatex.subsection(tex,'By Mineral')
            subtypescenarios = getsubtypescenarios(relevantscenarios,minerals,colorlist)
            countryscen,mineralscen = updatecontinentscenarios(subtypescenarios,countryscen,mineralscen,country,colorlist)
            graphN(subtypescenarios,folder+os.sep+tag+country+'_By_Minerals',n=N,orderform = [('mineral',minorder),('submineral',minerals)],ylabel = ylabel)
            mlatex.figure(tex,[figfolder+'/'+tag+country+'_By_Minerals.eps'],clean(country) + ' projection by mineral type')
            tabularise(tex,subtypescenarios,'Peak years - Minerals',orderform = [('mineral',minorder),('submineral',minerals)])        
            examineregions(tex,folder,figfolder,tag,continent,country,regions,relevantscenarios,minerals,minorder,colorlist,mineraltoselect)
            mlatex.clear(tex)
        worldbycountries = updatecountries(worldbycountries,countryscen)
        if countryscen == {}:
            continue
        if mineraltoselect is None:
            mlatex.section(tex,'Total')
        else:
            mlatex.section(tex,clean(mineraltoselect))
        mlatex.subsection(tex,'By country')
        graphN(countryscen,folder+os.sep+tag+continent+'_ByCountry',n=NFIGS,ylabel = ylabel)
        mlatex.figure(tex,[figfolder+'/'+tag+continent+'_ByCountry.eps'],clean(continent) + ' projections by country')
        tabularise(tex,countryscen,'Peak years - '+ALL)
        mlatex.subsection(tex,'By mineral')
        graphN(mineralscen,folder+os.sep+tag+continent+'_bymineral',n=NFIGS,orderform = [('mineral',minorder),('submineral',minerals)],ylabel = ylabel)
        mlatex.figure(tex,[figfolder+'/'+tag+continent+'_bymineral.eps'],clean(continent) + ' projection by mineral type')
        tabularise(tex,mineralscen,'Peak years - Minerals',orderform = [('mineral',minorder),('submineral',minerals)])
        continentscen,worldmineralscen = updatecontinentscenarios(mineralscen,continentscen,worldmineralscen,continent,colorlist)
    worldprojections(tex,folder,figfolder,tag,continentscen,worldmineralscen,worldbycountries,minorder,minerals,colorlist,recycle)
    
    ### STEVE YOU ARE HERE --> need to work on getting the post processor script copies over.
def updatecountries(worldbycountries,countryscen):
    for country in countryscen.keys():
        worldbycountries[country] = copyscenario(countryscen[country])
        if country is not RECYCLEKEY:
            worldbycountries[country].legend = country
        else:
            worldbycountries[country].legend = RECYCLE
    return worldbycountries

def worldprojections(tex,folder,figfolder,typetag,continentscen,worldmineralscen,worldbycountries,minorder,minerals,colorlist,recycle):
    worldprojections_sub(tex,folder,figfolder,typetag,continentscen,worldmineralscen,worldbycountries,minorder,minerals,'Total','')
    if recycle == None:
        return
    worldbycountries = updatecountries(worldbycountries,{RECYCLEKEY:recycle})
    wmkeys = list(worldmineralscen.keys())
    if len(wmkeys) == 0:
        return 
    mint = wmkeys[0]
    continentscen,worldmineralscen = updatecontinentscenarios({RECYCLEKEY:recycle},continentscen,worldmineralscen,RECYCLEKEY,colorlist,mint= mint)
    recymins = {}
    for key in minerals:
        recymins[key] = minerals[key]
    recymins[RECYCLEKEY] = [RECYCLEKEY]
    worldprojections_sub(tex,folder,figfolder,typetag,continentscen,worldmineralscen,worldbycountries,minorder+[RECYCLEKEY],recymins,'Total with recycle','_WR')
    
def worldprojections_sub(tex,folder,figfolder,typetag,continentscen,worldmineralscen,worldbycountries,minorder,minerals,chapname,tag):
    mlatex.chapter(tex,chapname)
    ## by continent
    mlatex.section(tex,'By continent')
    graphN(continentscen,folder+os.sep+typetag+'World_bycontinent'+tag,n=N)
    mlatex.figure(tex,[figfolder+'/'+typetag+'World_bycontinent'+tag+'.eps'],'Total projections by continent')
    tabularise(tex,continentscen,'Peak years - '+ALL)
    mlatex.clear(tex)
    ## by mineral
    mlatex.section(tex,'By mineral')
    graphN(worldmineralscen,folder+os.sep+typetag+'world_bymineral'+tag,n=N,orderform = [('mineral',minorder),('submineral',minerals)])
    mlatex.figure(tex,[figfolder+'/'+typetag+'world_bymineral'+tag+'.eps'],'Total projection by mineral type')
    tabularise(tex,worldmineralscen,'Peak year by mineral')
    mlatex.clear(tex)
    ## by country
    mlatex.section(tex,'By Country')
    graphN(worldbycountries,folder+os.sep+typetag+'world_bycountry'+tag,n=N)
    mlatex.figure(tex,[figfolder+'/'+typetag+'world_bycountry'+tag+'.eps'],'Total projection by country')
    tabularise(tex,worldbycountries,'Peak year - Country')
            



def examineregions(tex,folder,figfolder,tag,continent,country,regions,relevantscenarios,minerals,minorder,colorlist,mineraltoselect):
    if regions==[ALL]:
        return
    mlatex.subsection(tex,'Regional Projections')    
    mlatex.clear(tex)
    projectionsbyregion = {}
    for reg in regions:
        print('\t\t'+reg)
        projectionsbyregion[reg] = regionalsection(tex,folder,figfolder,tag,continent,country,reg,relevantscenarios,minerals,minorder,colorlist,mineraltoselect)
    mlatex.subsection(tex,'Projection by region')
    graphN(projectionsbyregion,folder+os.sep+tag+country+'_complete_by_region',n=N)
    mlatex.figure(tex,[figfolder+'/'+tag+country+'_complete_by_region.eps'],clean(country) + ' by region projections capped at '+str(N))
    tabularise(tex,projectionsbyregion,'Peak years - ' + ALL)


def getorder(orderform,relscenarios):
    order = []
    if orderform in [None,'','rank']:
        recy = []
        for key in relscenarios.keys():
            if key == RECYCLEKEY:
                recy.append((key,None))
                continue
            order.append((key,relscenarios[key].urr))
        order = sorted(order,key = lambda a:a[1],reverse = True)
        order += recy
        return order
    if type(orderform) == type([]):
        if len(orderform)==1:
            return orderform
        key = orderform[0][0]
        priority = orderform[0][1]
        order = []
        for prior in priority:
            subkey = orderform[1][0]
            subpriority = orderform[1][1][prior]
            for subprior in subpriority:
                for name in relscenarios.keys():
                    if getattr(relscenarios[name],key)==prior and getattr(relscenarios[name],subkey) == subprior:
                        order.append((name,'blank'))
        return order    
    raise Exception()




def getspecificscenarios(scenarios,dic,mineraltoselect):
    relscenarios = {}
    subregions = []
    for name in scenarios:
        keep = True
        for key in dic:
            if getattr(scenarios[name],key) != dic[key]:
                keep = False
                break
        if not keep:
            continue
        if mineraltoselect == None or mineraltoselect == getattr(scenarios[name],MINERALTERM):
            relscenarios[name] = copyscenario(scenarios[name])
            subregions.append(getattr(scenarios[name],REGIONTERM))
    subregions = sorted(list(set(subregions)))
    return relscenarios,subregions
            
def copyscenario(scenario1,color = None,legend = None):
    scenario = copy.deepcopy(scenario1)
    scenario.color = color
    scenario.legend = legend
    return scenario


def maketitlepage(tex,scenario):
    now = datetime.datetime.now()
    tex.write('$~$\\\\[3cm]\n')
    tex.write('\\begin{center}\n')
    tex.write('\\Huge{GeRS-DeMo Post Processor}\\\\[8cm]\n')
    tex.write('\\end{center}')
    tex.write('\\Large{Compilation ran on scenario: '+clean(scenario)+'}\\\\[2cm]\n')
    tex.write('\\Large{Date: '+str(now.day)+'/'+str(now.month)+'/'+str(now.year)+'}\\\\[2cm]\n')

def clean(rawstring):
    string = rawstring
    if '_' in string:
        string = string.replace('_',' ')
    return string


def sortscenarios(scenarios,key,link,order= 'alpha',skipkeys = []):
    sort = {}
    othersort = {}
    for k in scenarios.keys():
        if k in skipkeys:
            continue
        keyname = getattr(scenarios[k],key)
        linkname = getattr(scenarios[k],link)
        if not keyname in sort:
            sort[keyname] = {}
            othersort[keyname] = 0
        if not linkname in sort[keyname]:
            sort[keyname][linkname] = 0
        sort[keyname][linkname] += getattr(scenarios[k],URRTERM)
        othersort[keyname] += getattr(scenarios[k],URRTERM)
    if order == 'alpha':
        keyorder = sorted(sort.keys())
    elif order == 'URR':
        temp = []
        for key in sort.keys():
            temp.append((key,othersort[key]))
        skeyorder = sorted(temp,key = lambda a:a[1],reverse = True)
        keyorder = []
        for skey in skeyorder:
            keyorder.append(skey[0])
    else:
        raise Exception()
    for keyname in sort.keys():
        if order == 'alpha':
            sort[keyname] = sorted(sort[keyname].keys())
        elif order == 'URR':
            temp = []
            for key in sort[keyname].keys():
                temp.append((key,sort[keyname][key]))
            temp = sorted(temp,key = lambda a:a[1],reverse = True)
            sort[keyname] = []
            for t in temp:
                sort[keyname].append(t[0])
        else:
            raise Exception()
    return sort,keyorder

def getcolorlist(minerals,minorder,recycle):
    if len(minerals) <= len(SPECIFICCOLORS):
        colors = {}
        for i in range(0,len(minorder)):
            mineral = minorder[i]
            colors[mineral] = {}
            colorlist = SPECIFICCOLORS[i]
            for j in range(0,len(minerals[mineral])):
                colors[mineral][minerals[mineral][j]] = colorlist[j]
        if recycle != None:
            colors[RECYCLEKEY] = {RECYCLEKEY:DARK_GREY}
        return colors
    j = 0
    colors = {}
    lenny = len(COLORLIST)
    for mineral in minorder:
        colors[mineral] = {}
        for sub in minerals[mineral]:
            colors[mineral][sub] = COLORLIST[j]
            j+=1
            if j >= lenny:
                j = 0
    if recycle != None:
        colors[RECYCLEKEY] = {RECYCLEKEY:DARK_GREY}
    return colors


def postprocessorpage(root,container,canvas,frames,mainclass):
    ## Title
    root,frames = maketitle(root,frames,POSTPROCESSORFRAME,POSTPROCESSORTITLETEXT)
    ## Desc
    root,frames = makedesc(root,frames,POSTPROCESSORFRAME,POSTPROCESSORDESCTEXT)    
    ## back button
    #w = OptionMenu(master, variable, "one", "two", "three")
    dropdownlabel = elemcls(row = PPDDROPDOWNLABELROW,
                            column = PPDDROPDOWNLABELCOL,
                            text = PPDROPDOWNLABELTEXT,
                            font = DESCFONT,
                            padx = PADX,
                            obj = 'label',
                            frame = frames[POSTPROCESSORFRAME].frame,
                            fg = DESCCOLOR,
                            justify = 'left',
                            anchor = 'w',
                            bg = FRAMECOLOR,
                            )
    frames[POSTPROCESSORFRAME].elems[DROPDOWNLABELTAG] = dropdownlabel
    ddvar = tk.StringVar()
    dropdown = elemcls(row = PPDROPDOWNROW,
                       column = PPDROPDOWNCOL,
                       padx = PADX,
                       pady = PADY,
                       columnspan = 1,
                       sticky = 'nswe',
                       obj = 'dropdown',
                       lista = ['None'],
                       variable = ddvar,
                       frame = frames[POSTPROCESSORFRAME].frame,
                       fg = DESCCOLOR,
                       bg = FRAMECOLOR,
                       font = TEXTFONT)
    #https://stackoverflow.com/questions/28412496/updating-optionmenu-from-list
    dropdown.grid()
    frames[POSTPROCESSORFRAME].elems[DROPDOWNTAG] = dropdown
    frames[POSTPROCESSORFRAME].dropdownvar = ddvar    
    frames[POSTPROCESSORFRAME].dropdownvar.trace('w', dropdownvarupdated(root,frames,mainclass,POSTPROCESSORFRAME,container))
    root,frames = backbutton(root,frames,mainclass,POSTPROCESSORFRAME,GLOBALINPUTSFRAME,container)
    ## Save button
    root,frames = savebuttons(root,frames,POSTPROCESSORFRAME,mainclass)
    ## OutputImage
    root,frames = outputimages(root,frames,mainclass,POSTPROCESSORFRAME,container)
    root,frames = excelpostprocessorbutton(root,frames,mainclass,POSTPROCESSORFRAME,container)
    root,frames = pdfpostprocessorbutton(root,frames,mainclass,POSTPROCESSORFRAME,container)
    forgetframe(root,frames,POSTPROCESSORFRAME)

def dropdownvarupdated(root,frames,mainclass,frame,container):
    def inner(var, index, mode):
        if frames[frame].dropdownvar.get() != '':
            loadimage(root,frames,mainclass,frame,container)
    return inner





def globalinputspage(root,container,canvas,frames,mainclass):
    '''
    for frame in [MAINFRAME,NEWMODELFRAME,SELECTMODELFRAME,GLOBALINPUTSFRAME]:
        frames[frame] = createemptyframe(root,container,FRAMECOLOR)
    createmainframe(root,container,canvas,frames,mainclass)
    newmodelpage(root,container,canvas,frames,mainclass)
    selectmodeltypepage(root,container,canvas,frames,mainclass)
    '''
    ## Title
    root,frames = maketitle(root,frames,GLOBALINPUTSFRAME,GLOBALINPUTSTITLETEXT)
    ## Desc
    root,frames = makedesc(root,frames,GLOBALINPUTSFRAME,GLOBALINPUTSDESCTEXT)    
    ## back button
    root,frames = backbutton(root,frames,mainclass,GLOBALINPUTSFRAME,SELECTMODELFRAME,container)
    ## Save button
    root,frames = savebuttons(root,frames,GLOBALINPUTSFRAME,mainclass)
    root,frames = createscenarios(root,frames,mainclass,GLOBALINPUTSFRAME,GLOBALINPUTTERMS,GLOBALINPUTENTRYTAG,GLOBALINPUTTITLETAG)

    for key in DEFAULTINTINPUTS:
        val = str(DEFAULTINTINPUTS[key])
        frames[GLOBALINPUTSFRAME].elems[key+GLOBALINPUTENTRYTAG].elem.delete(0,tk.END)
        frames[GLOBALINPUTSFRAME].elems[key+GLOBALINPUTENTRYTAG].elem.insert(0,val)
    root,frames = runfullmodelbutton(root,container,frames,mainclass,GLOBALINPUTSFRAME,POSTPROCESSORFRAME)
    forgetframe(root,frames,GLOBALINPUTSFRAME)
        

def selectmodeltypepage(root,container,canvas,frames,mainclass):
    ## Title
    root,frames = maketitle(root,frames,SELECTMODELFRAME,MODELTYPETITLETEXT)
    ## Desc
    root,frames = makedesc(root,frames,SELECTMODELFRAME,MODELTYPEDESCTEXT)    
    ## Hubbert etc models
    root,frames = makecycles(root,container,canvas,frames,mainclass,SELECTMODELFRAME,editmodelfunc,SCENARIOSLISTLOC,SCENARIOSLISTSPAN,NEWSCENLOC,NEWSCENTEXT,SCENARIOBUTTON,SCENARIOTAG,generalscenariofunc,SCENSCROLLSIZE,editscenariopage,'edit edit edit edit')
    ## Save button
    root,frames = savebuttons(root,frames,SELECTMODELFRAME,mainclass)
    ## back button
    root,frames = backbutton(root,frames,mainclass,SELECTMODELFRAME,NEWMODELFRAME,container)
    ## next page
    root,frames = nextbutton(root,frames,mainclass,SELECTMODELFRAME,GLOBALINPUTSFRAME,container)
    ## HERERERERERERERERE
    forgetframe(root,frames,SELECTMODELFRAME)


def HLbutton(root,frames,frame):
##    urrbox = elemcls(row = URRBOXLOC[0],
##                     column = URRBOXLOC[1],
##                     padx = PADX,
##                     pady = PADY,
##                     columnspan = 1,
##                     sticky = 'nswe',
##                     elem = tk.Entry(frames[frame].frame,width = DEFAULTINPUTWIDTH))
##    urrbox.grid()
##    frames[frame].elems[URRBOX] = urrbox


    urrlabel = elemcls(row = URRLABELLOC[0],
                       column = URRLABELLOC[1],
                       padx = PADX,
                       pady = PADY,
                       columnspan = 1,
                       sticky = 'nswe',
                       obj = 'label',
                       frame = frames[frame].frame,
                       text = "",
                       fg = DESCCOLOR,
                       bg = FRAMECOLOR,
                       font = TEXTFONT)
    urrlabel.grid()
    frames[frame].elems[URRLABEL] = urrlabel
    frames[frame].othervars['digits'] = None
    frames[frame].othervars['urrval'] = None

    currurrtitle = elemcls(row = CURRURRTITLELOC[0],
                           column = CURRURRTITLELOC[1],
                           padx = PADX,
                           pady = PADY,
                           columnspan = 1,
                           sticky = 'e',
                           obj = 'label',
                           frame = frames[frame].frame,
                           text = CURRURRLABEL,
                           fg = DESCCOLOR,
                           bg = FRAMECOLOR,
                           font = TEXTFONT)
    currurrtitle.grid()
    frames[frame].elems[CURRURRTITLE] = currurrtitle

    currurrvalue = elemcls(row = CURRURRVALUELOC[0],
                           column = CURRURRVALUELOC[1],
                           padx = PADX,
                           pady = PADY,
                           columnspan = 1,
                           sticky = 'w',
                           obj = 'label',
                           frame = frames[frame].frame,
                           text = CURRURRVALUE,
                           fg = DESCCOLOR,
                           bg = FRAMECOLOR,
                           font = TEXTFONT)
    currurrvalue.grid()
    frames[frame].elems[CURRURRVAL] = currurrvalue

    
    urrtitle = elemcls(row = URRTITLELOC[0],
                       column = URRTITLELOC[1],
                       columnspan = 1,
                       sticky = 'e',
                       padx = PADX,
                       obj = 'label',
                       frame = frames[frame].frame,
                       text = URR,
                       fg = DESCCOLOR,
                       bg = FRAMECOLOR,
                       font = TEXTFONT)
    urrtitle.grid()
    frames[frame].elems[URRTITLE] = urrtitle
##    hlbutton = elemcls(row = HLBUTTONLOC[0],
##                       column = HLBUTTONLOC[1],
##                       padx = PADX,
##                       pady = PADY,
##                       text = HLBUTTONTEXT,
##                       elem = tk.Button(frames[frame].frame,text = HLBUTTONTEXT,command = lambda : HLfunction(root,frames,frame),font = BUTTONFONT))
##    hlbutton.grid()
##    frames[frame].elems[HLBUTTON] = hlbutton
## HERERER - STEVE TODO HERERERE
    return root,frames

def HLfunction(root,frames,frame):
    pass

def createscenarios(root,frames,mainclass,frame,terms,entrytag,titletag,width = DEFAULTINPUTWIDTH):
    for term in terms:
        cl = terms[term]
        if cl.objtype == 'auto':
            scenario = elemcls(row = cl.entityloc[0],
                               column = cl.entityloc[1],
                               padx = PADX,
                               pady = PADY,
                               sticky = 'w',
                               obj = 'autocomplete',
                               lista = getattr(mainclass,cl.defaultfield),
                               frame = frames[frame].frame,
                               width = width)
            scenariolabel = elemcls(row = cl.entityloc[0],
                                    column = cl.entityloc[1],
                                    padx = PADX,
                                    pady = PADY,
                                    sticky = 'w',
                                    obj = 'label',
                                    frame = frames[frame].frame,
                                    width = width,
                                    anchor = 'w',
                                    text = '',
                                    fg = DESCCOLOR,
                                    bg = FRAMECOLOR,
                                    justify = 'right',
                                    font = TEXTFONT)
            frames[frame].elems[term+LABEL] = scenariolabel
            scenariolabel.grid()
        elif cl.objtype == 'Input':
            scenario = elemcls(row = cl.entityloc[0],
                               column = cl.entityloc[1],
                               padx = PADX,
                               pady = PADY,
                               sticky = 'w',
                               obj = 'entry',
                               frame = frames[frame].frame,
                               width = width)
        else:
            raise Exception()
        scenario.grid()
        scenariotitle = elemcls(row = cl.titleloc[0],
                                column = cl.titleloc[1],
                                padx = PADX,
                                pady = PADY,
                                sticky = 'e',
                                obj = 'label',
                                frame = frames[frame].frame,
                                width = width,
                                anchor = 'e',
                                text = cl.title+':',
                                fg = DESCCOLOR,
                                bg = FRAMECOLOR,
                                justify = 'right',
                                font = TEXTFONT)
        ### HERERERERERER
        scenariotitle.grid()
        frames[frame].elems[term+entrytag] = scenario
        frames[frame].elems[term+titletag] = scenariotitle        
        
    return root,frames

def makeproddict(years,prod):
    pdict = {}
    for i in range(0,len(years)):
        pdict[years[i]] = prod[i]
    return pdict

def getdiffs(myears,mprod,dyears,dprod):
    mdict = makeproddict(myears,mprod)
    ddict = makeproddict(dyears,dprod)
    syears = sorted(list(mdict.keys()))
    dyears = []
    diffs = []
    for year in syears:
        if year not in ddict:
            continue
        dyears.append(year)
        diffs.append(mdict[year] - ddict[year])
    return dyears,diffs 

def safemax(alist,default = 0):
    if alist is None:
        return default
    if len(alist) == 0:
        return default
    return max(alist)


def initrunfiles(mainclass):
    tmp = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
    tmpfolder = tmp.name
    intf = open(tmpfolder+os.sep+INTERACTIONSFILE,'w',newline = '')
    minef = open(tmpfolder+os.sep+MINESFILE,'w',newline = '')
    fieldf = open(tmpfolder+os.sep+FIELDSFILE,'w',newline = '')
    intfcsv = csv.writer(intf,delimiter = ',',quotechar = '"')
    minefcsv = csv.writer(minef,delimiter = ',',quotechar = '"')
    fieldfcsv = csv.writer(fieldf,delimiter = ',',quotechar = '"')
    try:
        intf = open(tmpfolder+os.sep+INTERACTIONSFILE,'w',newline = '')
        minef = open(tmpfolder+os.sep+MINESFILE,'w',newline = '')
        fieldf = open(tmpfolder+os.sep+FIELDSFILE,'w',newline = '')
        intfcsv = csv.writer(intf,delimiter = ',',quotechar = '"')
        minefcsv = csv.writer(minef,delimiter = ',',quotechar = '"')
        fieldfcsv = csv.writer(fieldf,delimiter = ',',quotechar = '"')
    except:
        popupmessage(FAILEDOPENFILE,'Failed to create csv files')
        return
    return tmpfolder,intf,minef,fieldf,intfcsv,minefcsv,fieldfcsv

def makecleanexcelsheetname(name,existingnames):
    clean = name
    if len(clean) > 30:
        clean = clean[0:30]
    better = ''
    for char in clean:
        if char.isalnum() or char in GOODEXCELNONALPHANUM:
            better += char
    if better == '':
        for i in range(0,100):
            tmp = 'NoScenarioName_'+str(i)
            if tmp not in existingnames:
                return tmp
        return None
    if better not in existingnames:
        return better
    for i in range(0,100):
        tmp = better+str(i)
        if tmp not in existingnames:
            return tmp
    return None
    

def runfullscenario(root,container,frames,mainclass,frame,toframe):
    alllines = makelines(frames)
    frames[toframe].imageoptions = {}
    
    for scen in alllines:
        try:
            runfullscenario_sub(root,container,frames,mainclass,frame,toframe,scen,alllines[scen])
        except:
            popupmessage(FAILEDOPENFILE,'Failed to generate scenario')
            return

    navigateto(root,frames,mainclass,frame,toframe,container)
    frames[toframe].dropdownvar.set('')
    frames[toframe].elems[DROPDOWNTAG].elem['menu'].delete(0,'end')
    
    
    keys = sorted(list(alllines.keys()))
    if len(keys) > 0:
         frames[toframe].dropdownvar.set(keys[0])
    loadimage(root,frames,mainclass,toframe,container)

    return

def runfullscenario_sub(root,container,frames,mainclass,frame,toframe,scen,lines):

    frames[POSTPROCESSORFRAME].elems[DROPDOWNTAG].elem['menu'].add_command(label = scen,
                                    command=tk._setit(frames[POSTPROCESSORFRAME].dropdownvar,scen))    
    filesnfolders = initrunfiles(mainclass)
    if filesnfolders is None:
        return 

    tmpfolder,intf,minef,fieldf,intfcsv,minefcsv,fieldfcsv = filesnfolders
    for pair in ITERATIONFIELDSORDER:
        term,func = pair
        val = frames[GLOBALINPUTSFRAME].elems[term+GLOBALINPUTENTRYTAG].elem.get()
        row = [term,val]
        intfcsv.writerow(row)
        
    writecsvminefieldfile(minefcsv,MINEEXCELSPECIFIC,lines[MINE])
    writecsvminefieldfile(fieldfcsv,FIELDEXCELSPECIFIC,lines[WELL])

    intf.close()
    minef.close()
    fieldf.close()

    os.system(CPPMODEL + ' ' + tmpfolder)
    f = open(tmpfolder+os.sep+CPPOUTPUTFILE,'r')
    csvf = csv.reader(f,delimiter = ',',quotechar = '"')
    units = frames[NEWMODELFRAME].elems[LOADUNITS+INPUTTERMSENTRYTAG].elem.get()
    years,production,cumulative = gettotalcsvmodelresults(csvf,mainclass,scen,units,updatemainclass = True)
    f.close()
    histyears,histprod,histcum = getallproduction(mainclass.production)
    tmpimages = makeinteractiveplots(frames,toframe,years,production,cumulative,histyears,histprod,histcum)
    for key in tmpimages:
        frames[toframe].imageoptions[scen+'|'+key] = tmpimages[key]

    




def makelines(frames,frame = None):
    scenmap = getscenariomap(frames)
    lines = {}
    if frame is not None:
        lines = makelines_sub(lines,scenmap,frames,frame)
        return lines
    for frm in scenmap:
        lines = makelines_sub(lines,scenmap,frames,frm)
    return lines

def makelines_sub(lines,scenmap,frames,frame):
    basename,scenario = getbasename(frames[frame])
    row = 0
    if scenario not in lines:
        lines[scenario] = {MINE:[],WELL:[]}
    for ckey in scenmap[frame]:
        row += 1
        linetype,line = gettextline(basename,frames,frames[ckey],scenmap[frame][ckey],row)
        if linetype is None:
            continue
        lines[scenario][linetype].append(line)
    return lines
        

def runscenario(root,container,frames,mainclass,frame):
    filesnfolders = initrunfiles(mainclass)
    if filesnfolders is None:
        return
    tmpfolder,intf,minef,fieldf,intfcsv,minefcsv,fieldfcsv = filesnfolders

    for pair in ITERATIONFIELDSORDER:
        term,func = pair
        row = [term,RUNSCENINPUTS[term]]
        intfcsv.writerow(row)

    lines = makelines(frames,frame = frame)
    for scen in lines:
        writecsvminefieldfile(minefcsv,MINEEXCELSPECIFIC,lines[scen][MINE])
        writecsvminefieldfile(fieldfcsv,FIELDEXCELSPECIFIC,lines[scen][WELL])
    try:
        intf.close()
        minef.close()
        fieldf.close()
    except:
        popupmessage(FAILEDOPENFILE,'Failed to close temporary folder')
        return

    try: 
        os.system(CPPMODEL + ' ' + tmpfolder)
    except:
        popupmessage(FAILEDOPENFILE,'Failed to run model')
        return
    try:
        f = open(tmpfolder+os.sep+CPPOUTPUTFILE,'r')
        csvf = csv.reader(f,delimiter = ',',quotechar = '"')
    except:
        popupmessage(FAILEDOPENFILE,'Failed to read model results')
        return
    units = frames[NEWMODELFRAME].elems[LOADUNITS+INPUTTERMSENTRYTAG].elem.get()
    try:
        years,production,cumulative = gettotalcsvmodelresults(csvf,mainclass,'',units,updatemainclass = False)
    except:
        popupmessage(FAILEDOPENFILE,'Failed to get data')
        return
    try:
        f.close()
    except:
        popupmessage(FAILEDOPENFILE,'Failed to close file')
        return
    fcls = createcls(root,frames,mainclass,frame)
    if fcls is None:
        popupmessage(FAILEDOPENFILE,'Failed to create key')
        return    
    framekey = makekey(fcls,scenario = False)
    datayears = None
    dataprod = None
    datacum = None
    if framekey in mainclass.production:
        prodcls = mainclass.production[framekey]
        datayears = prodcls.years
        dataprod = prodcls.production
        datacum = prodcls.cumulative
    try: 
        frames[frame].imageoptions = makeinteractiveplots(frames,frame,years,production,cumulative,datayears,dataprod,datacum)
    except:
        popupmessage(FAILEDOPENFILE,'Failed to create plots')

    loadimage(root,frames,mainclass,frame,container)
    return



def makeinteractiveplots(frames,frame,years,production,cumulative,datayears = None,dataprod = None,datacum = None):
    pdict = {}
    pdict[CUMM] = makeinteractiveplot(frames,frame,years,cumulative,datayears,datacum)
    pdict[PROD] = makeinteractiveplot(frames,frame,years,production,datayears,dataprod)
    if datayears is not None and dataprod is not None and datayears != [] and dataprod != []:
        dyears,diffs = getdiffs(years,production,datayears,dataprod)
        pdict[DIFF] = makeinteractiveplot(frames,frame,dyears,diffs,None,None,False)    
    else:
        pdict[DIFF] = None
    return pdict

def makeinteractiveplot(frames,frame,years,model,histyears,historic,yminiszero = True):
    plots = []
    pm = plot.line(years,model,color = MODELCOLOR,style = '-')
    plots.append(pm)
    bubbleyears = copy.deepcopy(years)
    bubbleys = copy.deepcopy(model)
    if historic is not None and histyears is not None and historic != [] and histyears != []:
        pd = plot.line(histyears,historic,color = BLACK,style = '.')
        plots.append(pd)    
        bubbleyears += copy.deepcopy(histyears)
        bubbleys += copy.deepcopy(historic)

    pplot = plot.display(plots,None,save = False,figsize = HLPLOTSIZE)
    
    annot = pplot.ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",color = 'red',
                    bbox=dict(boxstyle="round", fc="w"),
                    arrowprops=dict(arrowstyle="->"))
    
    annot.set_visible(False)
    pplot.annot = annot

    ymax = 0
    ymin = 0
    if bubbleys != []:
        ymax = max(bubbleys)
        mini = min(bubbleys)
        if mini < ymin:
            ymin = mini
    pplot.bubbles = getbubbles(bubbleyears,bubbleys,bubbleyears,ymax,ymin = ymin)

    pcanvas = FigureCanvasTkAgg(pplot.fig,frames[frame].frame)
    pwidget = pcanvas.get_tk_widget()
    ptlframe = tk.Frame(master=frames[frame].frame)
    ptool = NavigationToolbar2Tk(pcanvas,ptlframe)
    ptool.update()
    pcanvas.mpl_connect("motion_notify_event", hover(pplot))
    return (pcanvas,pwidget,ptool,ptlframe)



def gethindex(header):
    hindex = {}
    for i in range(0,len(header)):
        hindex[header[i]] = i
    return hindex

def getyears(header):
    years = []
    if len(header) < FIRSTDATACOL:
        return years
    for y in range(FIRSTDATACOL,len(header)):
        years.append(float(header[y]))
    return years



def csvlinetoprodcls(line,units):
    if len(line) < 5:
        return None
    prodcls = productioncls()
    prodcls.continent = line[0]
    prodcls.country = line[1]
    regsub = line[2]
    if '$' in regsub:
        prodcls.region = regsub.split('$')[0]
        prodcls.subregion = regsub.split('$')[1]
    else:
        prodcls.region = regsub
    prodcls.mineral = line[3]
    prodcls.submineral = line[4]
    prodcls.unit = units
    return prodcls




def gettotalcsvmodelresults(csvf,mainclass,scen,units,updatemainclass = False):
    header = next(csvf)
    years = getyears(header)
    prod = []
    cumm = []
    proddict = {}
    for elem in years:
        prod.append(0)
        cumm.append(0)
    for line in csvf:
        prodcls = csvlinetoprodcls(line,units)
        pkey = makekey(prodcls,scenario = False)
        if pkey not in proddict:
            proddict[pkey] = ({},prodcls)
        for y in range(FIRSTDATACOL,len(line)):
            val = floaty(line[y])
            if val is None:
                continue
            year = years[y - FIRSTDATACOL]
            if year not in proddict[pkey][0]:
                proddict[pkey][0][year] = 0
            proddict[pkey][0][year] += val
            if line[0] in [RECYCLENAME,DEMAND]:
                continue
            prod[y - FIRSTDATACOL] += val
    tot = 0
    for i in range(0,len(prod)):
        tot += prod[i]
        cumm[i] = tot
    if not updatemainclass:
        return years,prod,cumm

    properproddict = {}
    for key in proddict:
        properproddict[key] = proddict[key][1]
        syears = sorted(proddict[key][0].keys())
        properproddict[key].years = syears
        properproddict[key].production = []
        properproddict[key].urr = 0
        peakyear = None
        peakamount = 0
        for y in syears:
            val = proddict[key][0][y]
            if val > peakamount:
                peakamount = val
                peakyear = y
            properproddict[key].urr += val
            properproddict[key].production.append(val)
        properproddict[key].peakyear = peakyear
        properproddict[key].peakrate = peakamount
        properproddict[key].data = None
        if key in mainclass.production:
            properproddict[key].data = makedatadict(mainclass.production[key])
    mainclass.modelresults[scen] = properproddict
    return years,prod,cumm

def makedatadict(prod):
    data = {}
    for i in range(0,len(prod.years)):
        yr = prod.years[i]
        prd = prod.production[i]
        data[yr] = prd
    return data

    
def writecsvminefieldfile(fcsv,order,lines):
    header = []
    for field in BASEEXCEL + order:
        header.append(field[0])
    fcsv.writerow(header)
    for line in lines:
        fcsv.writerow(line)

    
    '''
    NUMBERSTR = 'number'
    BASEEXCEL = [('continent',stringy),('country',stringy),('region',stringy),('mineral',stringy),('submineral',stringy),(NUMBERSTR,stringy)]
    MINEORDER = BASEEXCEL + MINEEXCELSPECIFIC
    FIELDORDER = BASEEXCEL + FIELDEXCELSPECIFIC
    '''
    


    ## STASTASTA
    ### need to create the base terms then the order row for relevant files.    
    ### need to make sure to ignore hidden cycles --> look to a save/export model function for guidance.


    
def runscenariobutton(root,container,frames,mainclass,frame):
    runscenbutton = elemcls(row = RUNSCENLOC[0],
                         column = RUNSCENLOC[1],
                         padx = PADX,
                         pady = PADY,
                         text = RUNSCENTEXT,
                         obj = 'button',
                         frame = frames[frame].frame,
                         command = lambda : runscenario(root,container,frames,mainclass,frame),
                         font = BUTTONFONT)
    runscenbutton.grid()
    frames[frame].elems[RUNSCEN] = runscenbutton

    return root,frames

def savebuttons(root,frames,fromframe,mainclass):
    ## Save buttons
    savebutton = elemcls(row = SAVEBUTTONLOC[0],
                         column = SAVEBUTTONLOC[1],
                         padx = PADX,
                         pady = PADY,
                         text = SAVEBUTTONTEXT,
                         obj = 'button',
                         frame = frames[fromframe].frame,
                         command = lambda : savemodel(root,frames,mainclass),
                         font = BUTTONFONT)
    savebutton.grid()
    frames[fromframe].elems[SAVE] = savebutton
    ## Save As button
    saveasbutton = elemcls(row = SAVEASBUTTONLOC[0],
                           column = SAVEASBUTTONLOC[1],
                           padx = PADX,
                           pady = PADY,
                           text = SAVEASBUTTONTEXT,
                           obj = 'button',
                           frame = frames[fromframe].frame,
                           command = lambda : saveasmodel(root,frames,mainclass),
                           font = BUTTONFONT)
    saveasbutton.grid()
    frames[fromframe].elems[SAVEAS] = saveasbutton

    exportbutton = elemcls(row = EXPORTBUTTONLOC[0],
                           column = EXPORTBUTTONLOC[1],
                           padx = PADX,
                           pady = PADY,
                           text = EXPORTBUTTONTEXT,
                           obj = 'button',
                           frame = frames[fromframe].frame,
                           command = lambda : exportmodel(root,frames,mainclass),
                           font = BUTTONFONT)
    exportbutton.grid()
    frames[fromframe].elems[EXPORT] = saveasbutton

    return root,frames


def outputimages(root,frames,mainclass,frame,container):
    imgvar = tk.StringVar()
    
    prodradio = elemcls(row = PRODIMGBUTTONLOC[0],
                        column = PRODIMGBUTTONLOC[1],
                        sticky = 'nswe',
                        obj = 'radiobutton',
                        frame = frames[frame].frame,
                        indicatoron = 0,
                        text = PRODIMGBUTTONTEXT,
                        command = lambda : loadimage(root,frames,mainclass,frame,container),
                        variable=imgvar,
                        value=PROD,
                        font = RADIOBUTTONFONT)
    prodradio.grid()
    frames[frame].elems[PRODIMG] = prodradio
    diffradio = elemcls(row = DIFFIMGBUTTONLOC[0],
                        column = DIFFIMGBUTTONLOC[1],
                        sticky = 'nswe',
                        obj = 'radiobutton',
                        frame = frames[frame].frame,
                        indicatoron = 0,
                        text = DIFFIMGBUTTONTEXT,
                        command = lambda : loadimage(root,frames,mainclass,frame,container),
                        variable=imgvar,
                        value=DIFF,
                        font = RADIOBUTTONFONT)
    diffradio.grid()
    frames[frame].elems[DIFFIMG] = diffradio
    cummradio = elemcls(row = CUMMIMGBUTTONLOC[0],
                        column = CUMMIMGBUTTONLOC[1],
                        sticky = 'nswe',
                        obj = 'radiobutton',
                        frame = frames[frame].frame,
                        indicatoron = 0,
                        text = CUMMIMGBUTTONTEXT,
                        command = lambda : loadimage(root,frames,mainclass,frame,container),
                        variable=imgvar,
                        value=CUMM,
                        font = RADIOBUTTONFONT)
    cummradio.grid()
    frames[frame].elems[CUMMIMG] = cummradio    
    
    frames[frame].imgvar = imgvar
    return root,frames
    
    




    
def newmodelpage(root,container,canvas,frames,mainclass):
    ## Title
    root,frames = maketitle(root,frames,NEWMODELFRAME,LOADPRODTITLETEXT)
    ## Desc
    root,frames = makedesc(root,frames,NEWMODELFRAME,LOADPRODDESCTEXT)
    ## load button
    loadprodbutton = elemcls(row = LOADPRODLOC[0],
                             column = LOADPRODLOC[1],
                             sticky = 'w',
                             text = LOADPRODTEXT,
                             obj = 'button',
                             frame = frames[NEWMODELFRAME].frame,
                             command = lambda : loadprodfile(root,frames,mainclass,NEWMODELFRAME,LOADPRODLAB),
                             font = BUTTONFONT)
    loadprodbutton.grid()
    frames[NEWMODELFRAME].elems[LOADPROD] = loadprodbutton
    loadprodfilelabel = elemcls(row = LOADFILELOC[0],
                                column = LOADFILELOC[1],
                                sticky = 'e',
                                padx = PADX,
                                obj = 'label',
                                frame = frames[NEWMODELFRAME].frame,
                                text = BLANKTEXT,
                                font = FILEFONT,
                                bg=FILECOLORBG,
                                fg=FILECOLOR,
                                width = MAXFILEINPUTWIDTH,
                                justify='right',
                                anchor = 'e')
    loadprodfilelabel.grid()
    frames[NEWMODELFRAME].elems[LOADPRODLAB] = loadprodfilelabel
    ## URR button
    loadURRbutton = elemcls(row = LOADURRLOC[0],
                            column = LOADURRLOC[1],
                            sticky = 'w',
                            text = LOADURRTEXT,
                            obj = 'button',
                            frame = frames[NEWMODELFRAME].frame,
                            command = lambda : loadurrfile(root,frames,mainclass,NEWMODELFRAME,LOADURRLAB),
                            font = BUTTONFONT)
    loadURRbutton.grid()
    frames[NEWMODELFRAME].elems[LOADURR] = loadURRbutton
    loadURRfilelabel = elemcls(row = TEXTURRLOC[0],
                               column = TEXTURRLOC[1],
                               sticky = 'e',
                               padx = PADX,
                               obj = 'label',
                               frame = frames[NEWMODELFRAME].frame,
                               text = BLANKTEXT,
                               font = FILEFONT,
                               bg=FILECOLORBG,
                               fg=FILECOLOR,
                               width = MAXFILEINPUTWIDTH,
                               justify='right',
                               anchor = 'e')
    loadURRfilelabel.grid()
    frames[NEWMODELFRAME].elems[LOADURRLAB] = loadURRfilelabel
    ## BIBTEX
    loadbibtexbutton = elemcls(row = BIBTEXLOC[0],
                               column = BIBTEXLOC[1],
                               sticky = 'w',
                               text = BIBTEXTEXT,
                               obj = 'button',
                               frame = frames[NEWMODELFRAME].frame,
                               command = lambda : loadbibtexfile(root,frames,mainclass,NEWMODELFRAME,LOADBIBTEXLAB),
                               font = BUTTONFONT)
    loadbibtexbutton.grid()
    frames[NEWMODELFRAME].elems[LOADBIBTEX] = loadbibtexbutton
    loadbibtextfilelabel = elemcls(row = LOADBIBTEXLOC[0],
                                   column = LOADBIBTEXLOC[1],
                                   sticky = 'e',
                                   padx = PADX,
                                   obj = 'label',
                                   frame = frames[NEWMODELFRAME].frame,
                                   text = BLANKTEXT,
                                   font = FILEFONT,
                                   bg=FILECOLORBG,
                                   fg=FILECOLOR,
                                   width = MAXFILEINPUTWIDTH,
                                   justify='right',
                                   anchor = 'e')
    loadbibtextfilelabel.grid()
    frames[NEWMODELFRAME].elems[LOADBIBTEXLAB] = loadbibtextfilelabel
    ### 

    
    root,frames = nextbutton(root,frames,mainclass,NEWMODELFRAME,SELECTMODELFRAME,container)
    ## Save buttons
    root,frames = savebuttons(root,frames,NEWMODELFRAME,mainclass)
    ## back button
    root,frames = backbutton(root,frames,mainclass,NEWMODELFRAME,MAINFRAME,container)

    root,frames = createscenarios(root,frames,mainclass,NEWMODELFRAME,LOADUNITSTERM,INPUTTERMSENTRYTAG,INPUTTERMSTITLETAG,width = UNITWIDTH)
    
    forgetframe(root,frames,NEWMODELFRAME)

def maketitle(root,frames,frame,title):
    maintitle = elemcls(columnspan = GRIDSIZE[1],
                        row = MAINTITLELOC[0],
                        column = MAINTITLELOC[1],
                        sticky = '',
                        obj = 'label',
                        frame = frames[frame].frame,
                        text = title,
                        fg = TITLECOLOR,
                        bg=FRAMECOLOR,
                        font = TITLEFONT)
    maintitle.grid()
    frames[frame].elems[TITLE] = maintitle
    return root,frames

    
def createmainframe(root,container,canvas,frames,mainclass):
    # TITLE
    root,frames = maketitle(root,frames,MAINFRAME,MAINTITLETEXT)
    # desc
    root,frames = makedesc(root,frames,MAINFRAME,MAINDESCTEXT)
    # New model button
    newmodelbutton = elemcls(row = NEWMODELLOC[0],
                             column = NEWMODELLOC[1],
                             columnspan = 2,
                             sticky='e',
                             text = NEWMODELTEXT,
                             obj = 'button',
                             frame = frames[MAINFRAME].frame,
                             command = lambda : navigateto(root,frames,mainclass,MAINFRAME,NEWMODELFRAME,container),
                             font = BUTTONFONT)
    newmodelbutton.grid()
    frames[MAINFRAME].elems[NEWMODEL] = newmodelbutton
    # Load button
    loadbutton = elemcls(row = LOADMODELLOC[0],
                         column = LOADMODELLOC[1],
                         columnspan = 2,
                         sticky = 'w',
                         text=LOADTEXT,
                         obj = 'button',
                         frame = frames[MAINFRAME].frame,
                         command = lambda : loadmodelfile(root,container,canvas,frames,mainclass,MAINFRAME),
                         font = BUTTONFONT)
    loadbutton.grid()
    frames[MAINFRAME].elems[LOAD] = loadbutton


    # Import button
    importbutton = elemcls(row = IMPORTMODELLOC[0],
                         column = IMPORTMODELLOC[1],
                         columnspan = 2,
                         sticky = 'w',
                         text=IMPORTTEXT,
                         obj = 'button',
                         frame = frames[MAINFRAME].frame,
                         command = lambda : importmodelfile(root,container,canvas,frames,mainclass,MAINFRAME),
                         font = BUTTONFONT)
    importbutton.grid()
    frames[MAINFRAME].elems[LOAD] = importbutton

    forgetframe(root,frames,MAINFRAME)
    
    

def main():
    root = initialiseframe()
    root.mainloop()
    




main()
